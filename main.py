import os
import sqlite3
import asyncio
from datetime import datetime
from typing import Optional, List
from contextlib import asynccontextmanager

import io
import json
import uuid

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, BackgroundTasks, HTTPException, Response, Request
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
import uvicorn

try:
    import httpx
    HTTPX_AVAILABLE = True
except ImportError:
    HTTPX_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

DB_PATH      = os.path.join(os.getcwd(), "drunks.db")
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
HOST         = "0.0.0.0"
PORT         = int(os.getenv("PORT", 8000))

# ── Auto-updater ──────────────────────────────
try:
    import updater as _updater
    APP_VERSION = _updater.get_current_version()
    _updater.start_background_check()
except Exception:
    _updater = None  # type: ignore
    APP_VERSION = "0.0.0"

# ─────────────────────────────────────────────
# FRONTEND — VENDEDOR
# ─────────────────────────────────────────────
VENDEDOR_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>Drunks · Vendedor</title>
<script>document.documentElement.setAttribute('data-theme',localStorage.getItem('drunksPalette')||'dark')</script>
<style>
html[data-theme=light] body{background:#f0ebff!important;color:#1e1b4b!important}
html[data-theme=light] [class*=bg-gray-95]{background:#eee9ff!important}
html[data-theme=light] [class*=bg-gray-90]{background:#fff!important}
html[data-theme=light] [class*=bg-gray-8]{background:#ede9fe!important}
html[data-theme=light] [class*=bg-gray-7]{background:#ddd6fe!important}
html[data-theme=light] [class*=border-gray-8]{border-color:#c4b5fd!important}
html[data-theme=light] [class*=border-gray-7]{border-color:#ddd6fe!important}
html[data-theme=light] [class*=bg-black]{background:rgba(80,50,180,.55)!important}
html[data-theme=light] .text-white{color:#1e1b4b!important}
html[data-theme=light] [class*=text-gray-]{color:#5b21b6!important}
html[data-theme=ocean] body{background:#020b18!important}
html[data-theme=ocean] [class*=bg-purple-6]{background:#2563eb!important}
html[data-theme=ocean] [class*=bg-purple-7]{background:#1d4ed8!important}
html[data-theme=ocean] [class*=bg-purple-8]{background:#1e40af!important}
html[data-theme=ocean] [class*=bg-purple-9]{background:#1e3a8a!important}
html[data-theme=ocean] [class*=text-purple]{color:#60a5fa!important}
html[data-theme=ocean] [class*=border-purple]{border-color:#3b82f6!important}
html[data-theme=ocean] [class*=from-purple]{--tw-gradient-from:#1e3a8a!important}
html[data-theme=ocean] [class*=via-purple]{--tw-gradient-via:#1d4ed8!important}
html[data-theme=ocean] [class*=glow-purple]{box-shadow:0 0 18px rgba(37,99,235,.4)!important}
html[data-theme=emerald] body{background:#020f0b!important}
html[data-theme=emerald] [class*=bg-purple-6]{background:#0d9488!important}
html[data-theme=emerald] [class*=bg-purple-7]{background:#0f766e!important}
html[data-theme=emerald] [class*=bg-purple-8]{background:#115e59!important}
html[data-theme=emerald] [class*=bg-purple-9]{background:#134e4a!important}
html[data-theme=emerald] [class*=text-purple]{color:#2dd4bf!important}
html[data-theme=emerald] [class*=border-purple]{border-color:#14b8a6!important}
html[data-theme=emerald] [class*=from-purple]{--tw-gradient-from:#134e4a!important}
html[data-theme=emerald] [class*=via-purple]{--tw-gradient-via:#0f766e!important}
html[data-theme=emerald] [class*=glow-purple]{box-shadow:0 0 18px rgba(13,148,136,.4)!important}
html[data-theme=amber] body{background:#0d0800!important}
html[data-theme=amber] [class*=bg-purple-6]{background:#d97706!important}
html[data-theme=amber] [class*=bg-purple-7]{background:#b45309!important}
html[data-theme=amber] [class*=bg-purple-8]{background:#92400e!important}
html[data-theme=amber] [class*=bg-purple-9]{background:#78350f!important}
html[data-theme=amber] [class*=text-purple]{color:#fbbf24!important}
html[data-theme=amber] [class*=border-purple]{border-color:#f59e0b!important}
html[data-theme=amber] [class*=from-purple]{--tw-gradient-from:#78350f!important}
html[data-theme=amber] [class*=via-purple]{--tw-gradient-via:#b45309!important}
html[data-theme=amber] [class*=glow-purple]{box-shadow:0 0 18px rgba(217,119,6,.4)!important}
html[data-theme=neon] body{background:#040804!important}
html[data-theme=neon] [class*=bg-purple-6]{background:#16a34a!important}
html[data-theme=neon] [class*=bg-purple-7]{background:#15803d!important}
html[data-theme=neon] [class*=bg-purple-8]{background:#166534!important}
html[data-theme=neon] [class*=bg-purple-9]{background:#14532d!important}
html[data-theme=neon] [class*=text-purple]{color:#4ade80!important}
html[data-theme=neon] [class*=border-purple]{border-color:#22c55e!important}
html[data-theme=neon] [class*=from-purple]{--tw-gradient-from:#14532d!important}
html[data-theme=neon] [class*=glow-purple]{box-shadow:0 0 18px rgba(74,222,128,.4)!important}
html[data-theme=rose] body{background:#0f0208!important}
html[data-theme=rose] [class*=bg-purple-6]{background:#e11d48!important}
html[data-theme=rose] [class*=bg-purple-7]{background:#be123c!important}
html[data-theme=rose] [class*=bg-purple-8]{background:#9f1239!important}
html[data-theme=rose] [class*=bg-purple-9]{background:#881337!important}
html[data-theme=rose] [class*=text-purple]{color:#fb7185!important}
html[data-theme=rose] [class*=border-purple]{border-color:#f43f5e!important}
html[data-theme=rose] [class*=from-purple]{--tw-gradient-from:#881337!important}
html[data-theme=rose] [class*=glow-purple]{box-shadow:0 0 18px rgba(244,63,94,.4)!important}
html[data-theme=violet] body{background:#06030f!important}
html[data-theme=violet] [class*=bg-purple-6]{background:#6d28d9!important}
html[data-theme=violet] [class*=bg-purple-7]{background:#5b21b6!important}
html[data-theme=violet] [class*=bg-purple-8]{background:#4c1d95!important}
html[data-theme=violet] [class*=bg-purple-9]{background:#3b0764!important}
html[data-theme=violet] [class*=text-purple]{color:#a78bfa!important}
html[data-theme=violet] [class*=border-purple]{border-color:#7c3aed!important}
html[data-theme=violet] [class*=from-purple]{--tw-gradient-from:#3b0764!important}
html[data-theme=violet] [class*=glow-purple]{box-shadow:0 0 18px rgba(124,58,237,.5)!important}
html[data-theme=slate] body{background:#0a0c10!important}
html[data-theme=slate] [class*=bg-purple-6]{background:#475569!important}
html[data-theme=slate] [class*=bg-purple-7]{background:#334155!important}
html[data-theme=slate] [class*=bg-purple-8]{background:#1e293b!important}
html[data-theme=slate] [class*=bg-purple-9]{background:#0f172a!important}
html[data-theme=slate] [class*=text-purple]{color:#94a3b8!important}
html[data-theme=slate] [class*=border-purple]{border-color:#64748b!important}
html[data-theme=slate] [class*=from-purple]{--tw-gradient-from:#0f172a!important}
html[data-theme=slate] [class*=glow-purple]{box-shadow:0 0 18px rgba(100,116,139,.3)!important}
html[data-theme=gold] body{background:#0d0900!important}
html[data-theme=gold] [class*=bg-purple-6]{background:#ca8a04!important}
html[data-theme=gold] [class*=bg-purple-7]{background:#a16207!important}
html[data-theme=gold] [class*=bg-purple-8]{background:#854d0e!important}
html[data-theme=gold] [class*=bg-purple-9]{background:#713f12!important}
html[data-theme=gold] [class*=text-purple]{color:#fde047!important}
html[data-theme=gold] [class*=border-purple]{border-color:#eab308!important}
html[data-theme=gold] [class*=from-purple]{--tw-gradient-from:#713f12!important}
html[data-theme=gold] [class*=glow-purple]{box-shadow:0 0 18px rgba(234,179,8,.4)!important}
</style>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  *{-webkit-tap-highlight-color:transparent;box-sizing:border-box}
  body{background:#08080f;color:#fff;font-family:'Segoe UI',system-ui,sans-serif;overscroll-behavior:none}
  ::-webkit-scrollbar{display:none}
  input::placeholder{color:#4b5563}
  input:focus{outline:none}

  .card-tap{transition:transform .12s ease,box-shadow .12s ease}
  .card-tap:active{transform:scale(.91)}

  @keyframes badgePop{0%{transform:scale(.2)}65%{transform:scale(1.35)}100%{transform:scale(1)}}
  .badge-pop{animation:badgePop .2s cubic-bezier(.175,.885,.32,1.275) forwards}

  @keyframes slideUp{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
  .slide-up{animation:slideUp .2s ease}

  @keyframes sheetIn{from{transform:translateY(100%)}to{transform:translateY(0)}}
  .sheet-in{animation:sheetIn .28s cubic-bezier(.32,0,.15,1)}

  @keyframes modalIn{from{opacity:0;transform:scale(.94) translateY(20px)}to{opacity:1;transform:scale(1) translateY(0)}}
  .modal-in{animation:modalIn .22s cubic-bezier(.175,.885,.32,1.275)}

  @keyframes shake{0%,100%{transform:translateX(0)}25%{transform:translateX(-6px)}75%{transform:translateX(6px)}}
  .shake{animation:shake .3s ease}

  .cat-pill{transition:all .15s ease}
  .glow-amber{box-shadow:0 0 18px rgba(217,119,6,.4)}
  .glow-emerald{box-shadow:0 0 18px rgba(16,185,129,.4)}
  .glow-purple{box-shadow:0 0 18px rgba(147,51,234,.4)}
  .glow-pink{box-shadow:0 0 18px rgba(236,72,153,.4)}
</style>
</head>
<body class="min-h-screen pb-28">

<!-- ═══ TOP STICKY ═══ -->
<div class="sticky top-0 z-40 bg-gray-950/97 backdrop-blur-md border-b border-gray-800/60 shadow-xl shadow-black/40">
  <!-- Logo + total -->
  <div class="flex items-center gap-2.5 px-4 pt-3 pb-1.5">
    <div class="w-8 h-8 rounded-xl bg-purple-600 flex items-center justify-center text-base shadow-lg shadow-purple-900/60 shrink-0">🍹</div>
    <span class="font-black tracking-widest text-purple-300 text-xs">DRUNKS POS</span>
    <div class="ml-auto flex items-center gap-2">
      <button onclick="toggleThemePicker()" title="Paleta de colores"
        class="w-8 h-8 rounded-xl bg-gray-800/80 border border-gray-700 text-sm flex items-center justify-center hover:bg-gray-700 transition-colors">🎨</button>
      <span class="text-gray-600 text-[10px]">TOTAL</span>
      <div id="totalDisplay" class="text-white font-extrabold text-xl leading-none tabular-nums">$0</div>
    </div>
  </div>
  <!-- Nombre del cliente — PROMINENTE -->
  <div class="px-4 pb-3">
    <div class="text-purple-400 font-black text-[10px] tracking-[0.22em] mb-1.5">PARA QUIEN ES EL PEDIDO</div>
    <input id="clientName" type="text" placeholder="Escribe el nombre del cliente..."
      autocomplete="off" autocorrect="off" spellcheck="false"
      class="w-full bg-gray-800/90 border-2 border-gray-700 focus:border-purple-400 rounded-2xl px-4 py-3 text-white text-[15px] font-bold transition-all placeholder-gray-600">
  </div>
</div>

<!-- ═══ CATEGORÍAS ═══ -->
<div class="sticky top-[115px] z-30 bg-gray-950/95 backdrop-blur-sm border-b border-gray-800/40">
  <div id="catBar" class="flex gap-2 px-4 py-2.5 overflow-x-auto" style="-webkit-overflow-scrolling:touch;scrollbar-width:none">
    <span class="text-gray-700 text-xs py-1 shrink-0">Cargando...</span>
  </div>
</div>

<!-- ═══ PRODUCTOS ═══ -->
<div class="px-3 pt-3 pb-1">
  <div id="prodsGrid" class="grid grid-cols-2 gap-2.5"></div>
</div>

<!-- ═══ CARRITO ═══ -->
<div class="px-3 pb-3 mt-1">
  <div class="flex items-center justify-between mb-1.5">
    <span class="text-[10px] text-gray-500 tracking-widest font-bold">CARRITO</span>
    <button id="clearCartBtn" onclick="clearCart()" class="hidden text-[10px] text-gray-600 hover:text-red-400 transition-colors">Limpiar todo</button>
  </div>
  <div id="cartList"></div>
</div>

<!-- ═══ NOTAS ═══ -->
<div class="px-3 pb-28">
  <div id="notesLabel" class="text-[10px] text-gray-500 tracking-widest font-bold mb-2">NOTAS RAPIDAS</div>
  <div id="notesBar" class="flex flex-wrap gap-1.5 mb-3"></div>
  <div class="flex gap-2">
    <input id="freeNote" type="text" placeholder="Nota manual para la bebida seleccionada..."
      class="flex-1 bg-gray-800 border border-gray-700 focus:border-purple-500 rounded-xl px-3 py-2.5 text-white text-sm transition-colors">
    <button onclick="applyFreeNote()"
      class="card-tap bg-purple-700 hover:bg-purple-600 text-white px-4 py-2.5 rounded-xl text-sm font-bold shrink-0">+</button>
  </div>
</div>

<!-- ═══ BARRA INFERIOR ═══ -->
<div class="fixed bottom-0 left-0 right-0 z-40 bg-gray-950/97 backdrop-blur-md border-t border-gray-800/60 px-3 py-3 shadow-2xl shadow-black">
  <div class="grid grid-cols-2 gap-3">
    <button onclick="openConfirm('Efectivo')"
      class="card-tap bg-gradient-to-br from-green-600 to-green-700 text-white font-black py-4 rounded-2xl text-sm shadow-lg shadow-green-950/60">
      <div class="text-xl mb-0.5">💵</div>EFECTIVO
    </button>
    <button onclick="openConfirm('Transferencia')"
      class="card-tap bg-gradient-to-br from-blue-600 to-blue-700 text-white font-black py-4 rounded-2xl text-sm shadow-lg shadow-blue-950/60">
      <div class="text-xl mb-0.5">📱</div>TRANSFERENCIA
    </button>
  </div>
</div>

<!-- ═══ THEME PICKER ═══ -->
<div id="themePicker" class="hidden fixed z-[90]" style="top:4.5rem;right:1rem">
  <div class="bg-gray-900 border border-gray-800 rounded-2xl p-4 shadow-2xl shadow-black/60 w-56">
    <div class="text-gray-500 text-[10px] tracking-[.2em] font-bold mb-3">PALETA DE COLORES</div>
    <div class="space-y-0.5">
      <button data-theme="dark"    onclick="applyTheme('dark')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#06060f"></span><span class="text-white text-sm font-semibold">Oscuro</span></button>
      <button data-theme="light"   onclick="applyTheme('light')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#f0ebff"></span><span class="text-white text-sm font-semibold">Claro</span></button>
      <button data-theme="ocean"   onclick="applyTheme('ocean')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#2563eb"></span><span class="text-white text-sm font-semibold">Océano</span></button>
      <button data-theme="emerald" onclick="applyTheme('emerald')" class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#0d9488"></span><span class="text-white text-sm font-semibold">Esmeralda</span></button>
      <button data-theme="amber"   onclick="applyTheme('amber')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#d97706"></span><span class="text-white text-sm font-semibold">Ámbar</span></button>
      <button data-theme="neon"    onclick="applyTheme('neon')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#16a34a"></span><span class="text-white text-sm font-semibold">Neón</span></button>
      <button data-theme="rose"    onclick="applyTheme('rose')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#e11d48"></span><span class="text-white text-sm font-semibold">Rosa</span></button>
      <button data-theme="violet"  onclick="applyTheme('violet')"  class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#6d28d9"></span><span class="text-white text-sm font-semibold">Violeta</span></button>
      <button data-theme="slate"   onclick="applyTheme('slate')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#475569"></span><span class="text-white text-sm font-semibold">Gris Pizarra</span></button>
      <button data-theme="gold"    onclick="applyTheme('gold')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#ca8a04"></span><span class="text-white text-sm font-semibold">Dorado</span></button>
    </div>
    <div class="mt-3 pt-3 border-t border-gray-800">
      <div class="text-gray-500 text-[10px] font-bold mb-2">COLOR PERSONALIZADO</div>
      <div class="flex items-center gap-2">
        <input type="color" id="customColorPicker" value="#9333ea"
          oninput="applyCustomTheme(this.value)"
          style="display:none;width:2rem;height:2rem;border-radius:.5rem;border:none;cursor:pointer;background:transparent;padding:0">
        <button onclick="document.getElementById('customColorPicker').style.display='inline-block';document.getElementById('customColorPicker').click()"
          class="flex items-center gap-2 px-3 py-1.5 bg-gray-800 hover:bg-gray-700 rounded-xl text-xs text-gray-300 transition-colors">
          <span>🎨</span> Elegir color
        </button>
      </div>
    </div>
  </div>
</div>

<!-- ═══ SELECTOR DE BASE (bottom sheet) ═══ -->
<div id="baseSheet" class="fixed inset-0 z-50 hidden items-end">
  <div class="absolute inset-0 bg-black/75" onclick="closeBase()"></div>
  <div class="sheet-in relative w-full bg-gray-900 rounded-t-3xl border-t border-gray-700 shadow-2xl pb-safe">
    <div class="flex justify-center pt-3 pb-1">
      <div class="w-12 h-1.5 bg-gray-700 rounded-full"></div>
    </div>
    <div class="px-5 pb-7">
      <div id="baseProdName" class="text-2xl font-extrabold text-center text-white mt-1 leading-tight"></div>
      <div class="text-gray-500 text-[10px] text-center tracking-[0.25em] mt-1 mb-5">SELECCIONA LA BASE</div>

      <div id="baseTypeSections"></div>

      <button onclick="closeBase()"
        class="mt-5 w-full bg-gray-800 hover:bg-gray-700 text-gray-400 font-semibold py-3 rounded-2xl text-sm transition-colors">
        Cancelar
      </button>
    </div>
  </div>
</div>

<!-- ═══ MODAL CONFIRMACIÓN ═══ -->
<div id="confirmModal" class="fixed inset-0 z-50 hidden items-end sm:items-center justify-center p-3 bg-black/85 backdrop-blur-sm">
  <div class="modal-in bg-gray-900 border border-gray-700 rounded-3xl w-full max-w-md shadow-2xl overflow-hidden">
    <div class="bg-gradient-to-r from-purple-900 via-purple-800 to-purple-900 px-6 py-5 text-center">
      <div class="text-3xl mb-1">📋</div>
      <h2 class="text-xl font-extrabold">Resumen del Pedido</h2>
      <p class="text-purple-300 text-xs mt-1 tracking-wide">Revisa antes de mandar a cocina</p>
    </div>
    <div class="px-5 py-4 max-h-64 overflow-y-auto space-y-3">
      <div class="flex items-start justify-between gap-3">
        <div>
          <div class="text-gray-500 text-[10px] tracking-widest">CLIENTE</div>
          <div id="smCliente" class="text-white font-extrabold text-2xl leading-tight"></div>
        </div>
        <div id="smMetodo" class="px-3 py-1.5 rounded-full text-sm font-bold shrink-0 mt-1"></div>
      </div>
      <div class="border-t border-gray-800 pt-2">
        <div id="smItems" class="space-y-2"></div>
      </div>
      <div class="border-t border-gray-700 pt-3 flex justify-between items-center">
        <span class="text-gray-400 font-semibold text-sm">TOTAL</span>
        <span id="smTotal" class="text-white font-extrabold text-2xl tabular-nums"></span>
      </div>
    </div>
    <div class="px-5 pb-5 pt-1 space-y-2">
      <button id="confirmBtn" onclick="sendOrder()"
        class="card-tap w-full bg-gradient-to-r from-green-500 to-green-600 text-white font-extrabold py-4 rounded-2xl text-base shadow-xl shadow-green-950/50">
        Confirmar y Mandar a Preparacion
      </button>
      <button onclick="closeConfirm()"
        class="w-full bg-gray-800 hover:bg-gray-700 text-gray-400 font-semibold py-3 rounded-2xl text-sm transition-colors">
        Cancelar - Seguir Editando
      </button>
    </div>
  </div>
</div>

<script>
// ── Constantes de UI ──
const CAT_CFG = {
  'Micheladas': { icon:'🍺', bg:'bg-amber-900/25 border-amber-700/20',    active:'bg-amber-900/60 border-amber-400/50 glow-amber',    pill:'bg-amber-700 text-amber-100' },
  'Mojitos':    { icon:'🍹', bg:'bg-emerald-900/25 border-emerald-700/20', active:'bg-emerald-900/60 border-emerald-400/50 glow-emerald', pill:'bg-emerald-700 text-emerald-100' },
  'Especiales': { icon:'✨', bg:'bg-pink-900/20 border-pink-700/20',       active:'bg-pink-900/50 border-pink-400/50 glow-pink',       pill:'bg-pink-700 text-pink-100' },
};
const DEF_CFG = { icon:'🍾', bg:'bg-gray-800/50 border-gray-700/20', active:'bg-gray-700/70 border-purple-400/60 glow-purple', pill:'bg-purple-700 text-purple-100' };

let availBases = [];  // cargadas dinámicamente desde /api/productos/{id}/bases
function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

// ── Estado ──
let products = [], categories = [], quickNotes = [];
let cart = {};          // key: "prodId|base" → { product, qty, base, obs }
let lastKey  = null;    // clave del item activo en carrito
let activeCat = null;   // categoría activa
let pendingProd = null; // producto esperando base
let pendingMethod = null;

// ── Init ──
async function init() {
  await Promise.all([loadCats(), loadProducts(), loadNotes()]);
}

async function loadCats() {
  try { const r = await fetch('/api/categorias'); categories = await r.json(); }
  catch(e) { categories = []; }
  renderCats();
}

async function loadProducts() {
  try { const r = await fetch('/api/productos'); products = await r.json(); }
  catch(e) { products = []; }
  renderProds();
}

async function loadNotes() {
  try { const r = await fetch('/api/notas_rapidas'); quickNotes = await r.json(); }
  catch(e) { quickNotes = []; }
  renderNotes();
}

// ── Categorías ──
function catCfg(nombre) { return CAT_CFG[nombre] || DEF_CFG; }

function renderCats() {
  const bar = document.getElementById('catBar');
  const all = `<button onclick="selectCat(null)"
    class="cat-pill shrink-0 px-4 py-1.5 rounded-full text-xs font-bold transition-all
      ${activeCat === null ? 'bg-purple-600 text-white shadow-lg glow-purple' : 'bg-gray-800 text-gray-400 border border-gray-700'}">
    Todas
  </button>`;
  bar.innerHTML = all + categories.map(c => {
    const cfg = catCfg(c.nombre);
    const isActive = activeCat === c.id;
    return `<button onclick="selectCat(${c.id})"
      class="cat-pill shrink-0 flex items-center gap-1.5 px-4 py-1.5 rounded-full text-xs font-bold transition-all
        ${isActive ? cfg.pill + ' shadow-lg' : 'bg-gray-800 text-gray-400 border border-gray-700'}">
      <span>${cfg.icon}</span>${c.nombre}
    </button>`;
  }).join('');
}

function selectCat(id) {
  activeCat = id;
  renderCats();
  renderProds();
}

// ── Productos ──
function totalQtyForProd(id) {
  return Object.entries(cart)
    .filter(([k]) => k.startsWith(id + '|'))
    .reduce((s, [, v]) => s + v.qty, 0);
}
function isActiveProd(id) {
  return lastKey && lastKey.startsWith(id + '|');
}

function renderProds() {
  const grid = document.getElementById('prodsGrid');
  const list = activeCat !== null ? products.filter(p => p.categoria_id === activeCat) : products;
  if (!list.length) {
    grid.innerHTML = '<div class="col-span-2 text-gray-700 text-center text-sm py-10">Sin productos en esta categoria</div>';
    return;
  }
  grid.innerHTML = list.map(p => {
    const qty    = totalQtyForProd(p.id);
    const avail  = !!p.disponible;
    const cfg    = catCfg(p.categoria_nombre || '');
    const active = isActiveProd(p.id);
    const isMich = !!p.tiene_base;
    return `<div
        class="card-tap relative rounded-2xl p-3.5 flex flex-col items-center justify-center min-h-[7rem] select-none border cursor-pointer
          ${avail ? (active ? cfg.active + ' shadow-lg' : cfg.bg) : 'opacity-30 cursor-not-allowed border-gray-800/20 bg-gray-800/10'}"
        ${avail ? 'onclick="tapProd(' + p.id + ')"' : ''}>
      ${qty > 0 ? '<div class="badge-pop absolute -top-2.5 -right-2.5 bg-purple-500 text-white text-xs font-extrabold w-7 h-7 rounded-full flex items-center justify-center shadow-xl z-10">' + qty + '</div>' : ''}
      ${isMich ? '<div class="absolute top-2 left-2 text-[8px] text-amber-400 font-bold tracking-wider bg-amber-900/40 px-1.5 py-0.5 rounded-full">BASE</div>' : ''}
      <div class="text-2xl mb-1.5">${cfg.icon}</div>
      <div class="text-white font-extrabold text-xs text-center leading-tight">${p.nombre}</div>
      <div class="font-extrabold text-sm mt-1.5 tabular-nums ${active ? 'text-white' : 'text-gray-300'}">$${Math.round(p.precio).toLocaleString('es-CO')}</div>
      ${!avail ? '<div class="absolute inset-0 flex items-center justify-center rounded-2xl bg-black/60"><span class="text-red-400 font-black text-[10px] tracking-widest">AGOTADO</span></div>' : ''}
    </div>`;
  }).join('');
}

function tapProd(id) {
  const p = products.find(x => x.id === id);
  if (!p || !p.disponible) return;
  if (p.tiene_base) {
    openBase(p);
  } else {
    const key = id + '|';
    lastKey = key;
    if (!cart[key]) { cart[key] = { product: p, qty: 1, base: '', obs: '' }; }
    else { cart[key].qty++; }
    renderProds(); renderCart(); renderNotes(); updateTotal();
  }
}

// ── Selector de Base ──
async function openBase(product) {
  pendingProd = product;
  document.getElementById('baseProdName').textContent = product.nombre;
  const container = document.getElementById('baseTypeSections');
  container.innerHTML = '<div class="text-gray-500 text-xs text-center py-4">Cargando...</div>';
  const s = document.getElementById('baseSheet');
  s.classList.remove('hidden'); s.classList.add('flex');
  try {
    const r = await fetch('/api/productos/'+product.id+'/bases');
    if (!r.ok) throw new Error('HTTP ' + r.status);
    const data = await r.json();
    availBases = Array.isArray(data) ? data : [];
  } catch(e) { availBases = []; }
  try {
    renderBaseSheet();
  } catch(e) {
    container.innerHTML = '<div class="text-red-400 text-xs text-center py-4">Error al cargar bases. Revisa la configuración.</div>';
  }
}

function renderBaseSheet() {
  // Agrupar por tipo usando tipo_nombre + tipo_icono del API
  const byTipo = {};
  availBases.filter(b => b.disponible).forEach(b => {
    const key = b.tipo_nombre || b.tipo || 'Otros';
    if (!byTipo[key]) byTipo[key] = { icono: b.tipo_icono || '🍹', items: [] };
    byTipo[key].items.push(b);
  });

  const container = document.getElementById('baseTypeSections');
  if (!Object.keys(byTipo).length) {
    container.innerHTML = '<div class="text-gray-600 text-xs text-center py-4">Sin bases disponibles</div>';
    return;
  }

  container.innerHTML = Object.entries(byTipo).map(([tipoNombre, { icono, items }]) => `
    <div class="mb-5">
      <div class="flex items-center gap-2 text-gray-400 text-[10px] tracking-widest font-bold mb-2.5">
        <span class="text-lg">${icono}</span> ${tipoNombre.toUpperCase()}
      </div>
      <div class="grid grid-cols-${Math.min(items.length,3)} gap-2">
        ${items.map(b => `
          <button data-base="${b.tipo_nombre||b.tipo}: ${b.nombre}" onclick="selectBase(this.dataset.base)"
            class="card-tap bg-gray-800/60 border border-gray-700 hover:bg-gray-700/60 hover:border-purple-500/40 rounded-2xl py-4 text-white font-bold text-sm transition-colors leading-tight">
            ${icono}<br>${esc(b.nombre)}
          </button>`).join('')}
      </div>
    </div>`).join('');
}

function closeBase() {
  const s = document.getElementById('baseSheet');
  s.classList.add('hidden'); s.classList.remove('flex');
  pendingProd = null;
}

function selectBase(base) {
  if (!pendingProd) return;
  const key = pendingProd.id + '|' + base;
  lastKey = key;
  if (!cart[key]) { cart[key] = { product: pendingProd, qty: 1, base, obs: '' }; }
  else { cart[key].qty++; }
  closeBase();
  renderProds(); renderCart(); renderNotes(); updateTotal();
}

// ── Carrito ──
function renderCart() {
  const el    = document.getElementById('cartList');
  const items = Object.entries(cart);
  const btn   = document.getElementById('clearCartBtn');
  if (!items.length) {
    el.innerHTML = '<div class="text-gray-700 text-center text-sm py-4 border border-gray-800/50 rounded-2xl">Toca un producto para agregarlo</div>';
    btn.classList.add('hidden');
    return;
  }
  btn.classList.remove('hidden');
  el.innerHTML = items.map(([key, item]) => {
    const isActive = key === lastKey;
    return `<div
        class="slide-up flex items-start gap-2.5 rounded-2xl px-3 py-2.5 mb-1.5 border cursor-pointer transition-all
          ${isActive ? 'bg-purple-900/40 border-purple-500/50 shadow-md' : 'bg-gray-800/50 border-gray-700/40'}"
        onclick="selectCartItem('${key.replace(/'/g,"\\'")}')">
      <div class="flex-1 min-w-0">
        <div class="text-white text-sm font-extrabold truncate">${item.product.nombre}</div>
        ${item.base ? '<div class="text-amber-400 text-xs font-bold mt-0.5">📌 ' + item.base + '</div>' : ''}
        ${item.obs  ? '<div class="text-yellow-300 text-xs mt-0.5">📝 ' + item.obs + '</div>' : ''}
      </div>
      <div class="flex items-center gap-2 shrink-0 mt-0.5">
        <div class="flex items-center gap-1">
          <button data-key="${key.replace(/"/g,'&quot;')}" onclick="event.stopPropagation();decreaseItem(this.dataset.key)"
            class="w-6 h-6 rounded-full bg-gray-700 hover:bg-gray-600 text-white text-xs font-bold flex items-center justify-center">-</button>
          <span class="text-white font-extrabold text-sm w-5 text-center tabular-nums">${item.qty}</span>
          <button data-key="${key.replace(/"/g,'&quot;')}" onclick="event.stopPropagation();increaseItem(this.dataset.key)"
            class="w-6 h-6 rounded-full bg-gray-700 hover:bg-gray-600 text-white text-xs font-bold flex items-center justify-center">+</button>
        </div>
        <div class="text-green-400 text-sm font-extrabold tabular-nums w-16 text-right">$${Math.round(item.product.precio * item.qty).toLocaleString('es-CO')}</div>
        <button data-key="${key.replace(/"/g,'&quot;')}" onclick="event.stopPropagation();removeItem(this.dataset.key)"
          class="text-gray-600 hover:text-red-400 text-lg leading-none ml-0.5 transition-colors">x</button>
      </div>
    </div>`;
  }).join('');
}

function selectCartItem(key) {
  lastKey = key;
  renderCart();
  renderNotes();
}

function increaseItem(key) {
  if (cart[key]) { cart[key].qty++; renderCart(); updateTotal(); }
}

function decreaseItem(key) {
  if (!cart[key]) return;
  cart[key].qty--;
  if (cart[key].qty <= 0) { delete cart[key]; if (lastKey === key) lastKey = null; }
  renderProds(); renderCart(); renderNotes(); updateTotal();
}

function removeItem(key) {
  delete cart[key];
  if (lastKey === key) lastKey = null;
  renderProds(); renderCart(); renderNotes(); updateTotal();
}

function clearCart() {
  if (!confirm('Limpiar todo el carrito?')) return;
  cart = {}; lastKey = null;
  renderProds(); renderCart(); renderNotes(); updateTotal();
}

function updateTotal() {
  const t = Object.values(cart).reduce((s, i) => s + i.product.precio * i.qty, 0);
  document.getElementById('totalDisplay').textContent = '$' + Math.round(t).toLocaleString('es-CO');
}

// ── Notas ──
function renderNotes() {
  const lbl  = document.getElementById('notesLabel');
  const bar  = document.getElementById('notesBar');
  let active = lastKey ? cart[lastKey] : null;

  if (active) {
    const prod = active.product;
    lbl.innerHTML = `NOTAS PARA: <span class="text-purple-400">${prod.nombre}${active.base ? ' · ' + active.base : ''}</span>`;
  } else {
    lbl.textContent = 'NOTAS RAPIDAS';
  }

  const catId = active ? active.product.categoria_id : null;
  const shown = quickNotes.filter(n => n.categoria_id === null || n.categoria_id === catId);

  if (!shown.length) {
    bar.innerHTML = '<span class="text-gray-700 text-xs">Sin notas configuradas para esta seccion.</span>';
    return;
  }

  const appliedNotes = (active && active.obs)
    ? active.obs.split(', ').map(x => x.trim()).filter(Boolean)
    : [];

  bar.innerHTML = shown.map(n => {
    const isCtx     = n.categoria_id !== null;
    const isApplied = appliedNotes.includes(n.texto);
    let cls;
    if (isApplied) {
      cls = 'bg-purple-700 border-purple-500 text-white shadow-md shadow-purple-900/40';
    } else if (isCtx) {
      cls = 'bg-amber-900/30 border-amber-700/40 text-amber-300 hover:bg-amber-800/50';
    } else {
      cls = 'bg-gray-800 border-gray-700 text-gray-300 hover:border-gray-500';
    }
    return `<button data-note="${n.texto.replace(/"/g,'&quot;')}" onclick="applyNote(this.dataset.note)"
      class="card-tap text-xs px-3 py-1.5 rounded-full border font-semibold transition-all ${cls}">
      ${isApplied ? '✓ ' : ''}${n.texto}
    </button>`;
  }).join('');
}

function applyNote(note) {
  if (!lastKey || !cart[lastKey]) { showToast('Toca una bebida del carrito primero', 'warn'); return; }
  const item  = cart[lastKey];
  const notes = item.obs ? item.obs.split(', ').map(n => n.trim()).filter(Boolean) : [];
  if (notes.includes(note)) {
    item.obs = notes.filter(n => n !== note).join(', ');
    renderCart(); renderNotes();
    showToast('Nota eliminada', 'warn');
  } else {
    item.obs = notes.length ? notes.join(', ') + ', ' + note : note;
    renderCart(); renderNotes();
    showToast(note, 'ok');
  }
}

function applyFreeNote() {
  const inp  = document.getElementById('freeNote');
  const note = inp.value.trim();
  if (!note) return;
  if (!lastKey || !cart[lastKey]) { showToast('Toca una bebida del carrito primero', 'warn'); return; }
  const item = cart[lastKey];
  item.obs = item.obs ? item.obs + ', ' + note : note;
  inp.value = '';
  renderCart();
  showToast('Nota agregada', 'ok');
}

// ── Flujo pedido ──
function openConfirm(method) {
  const nombre = document.getElementById('clientName').value.trim();
  const inp    = document.getElementById('clientName');
  if (!nombre) {
    inp.classList.add('shake', 'border-red-500');
    inp.focus();
    setTimeout(() => inp.classList.remove('shake', 'border-red-500'), 800);
    showToast('Escribe el nombre del cliente', 'error');
    return;
  }
  if (!Object.keys(cart).length) { showToast('El carrito esta vacio', 'error'); return; }

  pendingMethod = method;
  const items   = Object.values(cart);
  const total   = items.reduce((s, i) => s + i.product.precio * i.qty, 0);

  document.getElementById('smCliente').textContent = nombre;
  const mb = document.getElementById('smMetodo');
  mb.textContent = (method === 'Efectivo' ? '💵 ' : '📱 ') + method;
  mb.className   = 'px-3 py-1.5 rounded-full text-sm font-bold shrink-0 mt-1 ' +
    (method === 'Efectivo' ? 'bg-green-700 text-green-100' : 'bg-blue-700 text-blue-100');

  document.getElementById('smItems').innerHTML = items.map(i => `
    <div class="flex items-start gap-2 bg-gray-800/50 rounded-xl px-3 py-2">
      <span class="bg-purple-700 text-white text-xs font-extrabold px-2 py-0.5 rounded-md shrink-0">${i.qty}x</span>
      <div class="flex-1 min-w-0">
        <div class="text-white text-sm font-bold">${i.product.nombre}</div>
        ${i.base ? '<div class="text-amber-400 text-xs font-bold">📌 ' + i.base + '</div>' : ''}
        ${i.obs  ? '<div class="text-yellow-300 text-xs">📝 ' + i.obs + '</div>' : ''}
      </div>
      <div class="text-green-400 text-sm font-bold tabular-nums shrink-0">$${Math.round(i.product.precio * i.qty).toLocaleString('es-CO')}</div>
    </div>`).join('');

  document.getElementById('smTotal').textContent = '$' + Math.round(total).toLocaleString('es-CO');
  const m = document.getElementById('confirmModal');
  m.classList.remove('hidden'); m.classList.add('flex');
}

function closeConfirm() {
  const m = document.getElementById('confirmModal');
  m.classList.add('hidden'); m.classList.remove('flex');
  pendingMethod = null;
}

async function sendOrder() {
  const btn    = document.getElementById('confirmBtn');
  const cliente = document.getElementById('clientName').value.trim();
  const items   = Object.values(cart);
  const total   = items.reduce((s, i) => s + i.product.precio * i.qty, 0);
  btn.disabled  = true; btn.textContent = 'Enviando...';
  try {
    const r = await fetch('/api/pedidos', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        cliente, metodo_pago: pendingMethod, total,
        items: items.map(i => ({
          producto_id:   i.product.id,
          cantidad:      i.qty,
          observaciones: [i.base, i.obs].filter(Boolean).join(' | ')
        }))
      })
    });
    if (r.ok) {
      cart = {}; lastKey = null;
      document.getElementById('clientName').value = '';
      renderProds(); renderCart(); renderNotes(); updateTotal();
      closeConfirm();
      showToast('Pedido enviado a cocina!', 'ok');
    } else { showToast('Error al enviar', 'error'); }
  } catch(e) { showToast('Sin conexion', 'error'); }
  finally { btn.disabled = false; btn.textContent = 'Confirmar y Mandar a Preparacion'; }
}

function showToast(msg, type) {
  const c = { ok:'bg-green-700 border-green-600', error:'bg-red-800 border-red-700', warn:'bg-yellow-700 border-yellow-600' };
  const t = document.createElement('div');
  t.className = `fixed top-6 left-1/2 -translate-x-1/2 ${c[type]||'bg-gray-700 border-gray-600'} border text-white text-sm font-bold px-5 py-2.5 rounded-2xl z-[80] shadow-2xl pointer-events-none whitespace-nowrap`;
  t.textContent = msg;
  document.body.appendChild(t);
  setTimeout(() => t.remove(), 2200);
}

window.addEventListener('focus', loadNotes);
window.addEventListener('focus', loadProducts); // refresca disponibilidad y bases al volver al tab

// ── Settings persistidos en DB ──
function saveSetting(key, value) {
  localStorage.setItem(key, value);
  fetch('/api/settings/'+key, {
    method:'PUT', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({value})
  }).catch(()=>{});
}
async function loadSettings() {
  try {
    const s = await fetch('/api/settings').then(r=>r.json());
    const pal = s.palette || localStorage.getItem('drunksPalette') || 'dark';
    applyTheme(pal);
    Object.entries(s).forEach(([k,v]) => localStorage.setItem(k, v));
  } catch(e) {
    applyTheme(localStorage.getItem('drunksPalette') || 'dark');
  }
}

// ── Theme System ──
function applyTheme(t) {
  const theme = t||'dark';
  if (theme.startsWith('custom:')) {
    applyCustomTheme(theme.split(':')[1]); return;
  }
  document.documentElement.setAttribute('data-theme', theme);
  saveSetting('drunksPalette', theme);
  saveSetting('palette', theme);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b => {
    const on = b.dataset.theme === theme;
    b.style.background = on ? 'rgba(147,51,234,.2)' : '';
    b.style.outline    = on ? '1px solid rgba(147,51,234,.5)' : '';
  });
  const picker = document.getElementById('customColorPicker');
  if (picker) picker.style.display = 'none';
}
function applyCustomTheme(hex) {
  let style = document.getElementById('custom-theme-style');
  if (!style) { style = document.createElement('style'); style.id='custom-theme-style'; document.head.appendChild(style); }
  const h = hex||'#9333ea';
  const r = parseInt(h.slice(1,3),16), g = parseInt(h.slice(3,5),16), b = parseInt(h.slice(5,7),16);
  style.textContent = `
    html[data-theme="custom"] body{background:#06060f}
    html[data-theme="custom"] .bg-purple-600,.bg-purple-700{background-color:${h}!important}
    html[data-theme="custom"] .text-purple-400,.text-purple-300{color:${h}!important}
    html[data-theme="custom"] .border-purple-500{border-color:${h}!important}
    html[data-theme="custom"] .from-purple-600{--tw-gradient-from:${h}!important}
    html[data-theme="custom"] .glow{filter:drop-shadow(0 0 8px rgba(${r},${g},${b},.5))}`;
  document.documentElement.setAttribute('data-theme','custom');
  saveSetting('drunksPalette', 'custom:'+h);
  saveSetting('palette', 'custom:'+h);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b => {
    b.style.background=''; b.style.outline='';
  });
  const picker = document.getElementById('customColorPicker');
  if (picker) { picker.style.display='inline-block'; picker.value=h; }
}
function toggleThemePicker() {
  document.getElementById('themePicker').classList.toggle('hidden');
}
document.addEventListener('click', e => {
  if (!e.target.closest('#themePicker') && !e.target.closest('[onclick*="toggleThemePicker"]'))
    document.getElementById('themePicker')?.classList.add('hidden');
});

loadSettings().then(init);

// ── Auto-update ──
async function checkUpdate() {
  try {
    const d = await fetch('/api/update/check').then(r => r.json());
    if (d.has_update && d.latest) {
      document.getElementById('updateVersion').textContent = 'v'+d.current+' → v'+d.latest;
      document.getElementById('updateBanner').classList.remove('hidden');
    }
  } catch(e) {}
}
async function applyUpdate() {
  const btn = document.getElementById('updateBtn');
  btn.textContent = 'Descargando...'; btn.disabled = true;
  try {
    await fetch('/api/update/apply', {method:'POST'});
    btn.textContent = 'Reiniciando...';
  } catch(e) { btn.textContent = 'Error — intenta de nuevo'; btn.disabled = false; }
}
checkUpdate();
setInterval(checkUpdate, 30 * 60 * 1000);
</script>
<!-- ═══ BANNER DE ACTUALIZACIÓN ═══ -->
<div id="updateBanner" class="hidden fixed bottom-0 left-0 right-0 z-[100] bg-purple-950/97 border-t border-purple-500/50 backdrop-blur px-5 py-3 flex items-center justify-between gap-3 shadow-2xl">
  <div class="flex items-center gap-3">
    <span class="text-2xl">🔄</span>
    <div>
      <div class="text-white font-bold text-sm">Nueva versión disponible</div>
      <div id="updateVersion" class="text-purple-300 text-xs font-mono"></div>
    </div>
  </div>
  <div class="flex gap-2 shrink-0">
    <button id="updateBtn" onclick="applyUpdate()"
      class="bg-purple-600 hover:bg-purple-500 active:bg-purple-700 text-white text-sm font-extrabold px-4 py-2 rounded-xl transition-colors shadow-lg">
      Actualizar ahora
    </button>
    <button onclick="document.getElementById('updateBanner').classList.add('hidden')"
      class="text-purple-400 hover:text-white text-sm px-3 py-2 rounded-xl transition-colors">
      Después
    </button>
  </div>
</div>
</body>
</html>"""

# ─────────────────────────────────────────────
# FRONTEND — COCINA & ADMINISTRACIÓN  (v3 — priority layout + single-save + dashboard)
# ─────────────────────────────────────────────
COCINA_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Drunks · Cocina</title>
<script>document.documentElement.setAttribute('data-theme',localStorage.getItem('drunksPalette')||'dark')</script>
<style>
html[data-theme=light] body{background:#f0ebff!important;color:#1e1b4b!important}
html[data-theme=light] [class*=bg-gray-95]{background:#eee9ff!important}
html[data-theme=light] [class*=bg-gray-90]{background:#fff!important}
html[data-theme=light] [class*=bg-gray-8]{background:#ede9fe!important}
html[data-theme=light] [class*=bg-gray-7]{background:#ddd6fe!important}
html[data-theme=light] [class*=border-gray-8]{border-color:#c4b5fd!important}
html[data-theme=light] [class*=border-gray-7]{border-color:#ddd6fe!important}
html[data-theme=light] [class*=bg-black]{background:rgba(80,50,180,.55)!important}
html[data-theme=light] .text-white{color:#1e1b4b!important}
html[data-theme=light] [class*=text-gray-]{color:#5b21b6!important}
html[data-theme=ocean] body{background:#020b18!important}
html[data-theme=ocean] [class*=bg-purple-6]{background:#2563eb!important}
html[data-theme=ocean] [class*=bg-purple-7]{background:#1d4ed8!important}
html[data-theme=ocean] [class*=bg-purple-8]{background:#1e40af!important}
html[data-theme=ocean] [class*=bg-purple-9]{background:#1e3a8a!important}
html[data-theme=ocean] [class*=text-purple]{color:#60a5fa!important}
html[data-theme=ocean] [class*=border-purple]{border-color:#3b82f6!important}
html[data-theme=ocean] [class*=from-purple]{--tw-gradient-from:#1e3a8a!important}
html[data-theme=ocean] [class*=via-purple]{--tw-gradient-via:#1d4ed8!important}
html[data-theme=ocean] [class*=glow-purple]{box-shadow:0 0 18px rgba(37,99,235,.4)!important}
html[data-theme=ocean] .view-btn.active{background:rgba(37,99,235,.85)!important;box-shadow:0 0 14px rgba(37,99,235,.45)!important}
html[data-theme=emerald] body{background:#020f0b!important}
html[data-theme=emerald] [class*=bg-purple-6]{background:#0d9488!important}
html[data-theme=emerald] [class*=bg-purple-7]{background:#0f766e!important}
html[data-theme=emerald] [class*=bg-purple-8]{background:#115e59!important}
html[data-theme=emerald] [class*=bg-purple-9]{background:#134e4a!important}
html[data-theme=emerald] [class*=text-purple]{color:#2dd4bf!important}
html[data-theme=emerald] [class*=border-purple]{border-color:#14b8a6!important}
html[data-theme=emerald] [class*=from-purple]{--tw-gradient-from:#134e4a!important}
html[data-theme=emerald] [class*=via-purple]{--tw-gradient-via:#0f766e!important}
html[data-theme=emerald] [class*=glow-purple]{box-shadow:0 0 18px rgba(13,148,136,.4)!important}
html[data-theme=emerald] .view-btn.active{background:rgba(13,148,136,.85)!important;box-shadow:0 0 14px rgba(13,148,136,.45)!important}
html[data-theme=amber] body{background:#0d0800!important}
html[data-theme=amber] [class*=bg-purple-6]{background:#d97706!important}
html[data-theme=amber] [class*=bg-purple-7]{background:#b45309!important}
html[data-theme=amber] [class*=bg-purple-8]{background:#92400e!important}
html[data-theme=amber] [class*=bg-purple-9]{background:#78350f!important}
html[data-theme=amber] [class*=text-purple]{color:#fbbf24!important}
html[data-theme=amber] [class*=border-purple]{border-color:#f59e0b!important}
html[data-theme=amber] [class*=from-purple]{--tw-gradient-from:#78350f!important}
html[data-theme=amber] [class*=via-purple]{--tw-gradient-via:#b45309!important}
html[data-theme=amber] [class*=glow-purple]{box-shadow:0 0 18px rgba(217,119,6,.4)!important}
html[data-theme=amber] .view-btn.active{background:rgba(217,119,6,.85)!important;box-shadow:0 0 14px rgba(217,119,6,.45)!important}
html[data-theme=neon] body{background:#040804!important}
html[data-theme=neon] [class*=bg-purple-6]{background:#16a34a!important}html[data-theme=neon] [class*=bg-purple-7]{background:#15803d!important}html[data-theme=neon] [class*=bg-purple-8]{background:#166534!important}html[data-theme=neon] [class*=bg-purple-9]{background:#14532d!important}html[data-theme=neon] [class*=text-purple]{color:#4ade80!important}html[data-theme=neon] [class*=border-purple]{border-color:#22c55e!important}html[data-theme=neon] [class*=from-purple]{--tw-gradient-from:#14532d!important}html[data-theme=neon] [class*=glow-purple]{box-shadow:0 0 18px rgba(74,222,128,.4)!important}
html[data-theme=rose] body{background:#0f0208!important}
html[data-theme=rose] [class*=bg-purple-6]{background:#e11d48!important}html[data-theme=rose] [class*=bg-purple-7]{background:#be123c!important}html[data-theme=rose] [class*=bg-purple-8]{background:#9f1239!important}html[data-theme=rose] [class*=bg-purple-9]{background:#881337!important}html[data-theme=rose] [class*=text-purple]{color:#fb7185!important}html[data-theme=rose] [class*=border-purple]{border-color:#f43f5e!important}html[data-theme=rose] [class*=from-purple]{--tw-gradient-from:#881337!important}html[data-theme=rose] [class*=glow-purple]{box-shadow:0 0 18px rgba(244,63,94,.4)!important}
html[data-theme=violet] body{background:#06030f!important}
html[data-theme=violet] [class*=bg-purple-6]{background:#6d28d9!important}html[data-theme=violet] [class*=bg-purple-7]{background:#5b21b6!important}html[data-theme=violet] [class*=bg-purple-8]{background:#4c1d95!important}html[data-theme=violet] [class*=bg-purple-9]{background:#3b0764!important}html[data-theme=violet] [class*=text-purple]{color:#a78bfa!important}html[data-theme=violet] [class*=border-purple]{border-color:#7c3aed!important}html[data-theme=violet] [class*=from-purple]{--tw-gradient-from:#3b0764!important}html[data-theme=violet] [class*=glow-purple]{box-shadow:0 0 18px rgba(124,58,237,.5)!important}
html[data-theme=slate] body{background:#0a0c10!important}
html[data-theme=slate] [class*=bg-purple-6]{background:#475569!important}html[data-theme=slate] [class*=bg-purple-7]{background:#334155!important}html[data-theme=slate] [class*=bg-purple-8]{background:#1e293b!important}html[data-theme=slate] [class*=bg-purple-9]{background:#0f172a!important}html[data-theme=slate] [class*=text-purple]{color:#94a3b8!important}html[data-theme=slate] [class*=border-purple]{border-color:#64748b!important}html[data-theme=slate] [class*=from-purple]{--tw-gradient-from:#0f172a!important}html[data-theme=slate] [class*=glow-purple]{box-shadow:0 0 18px rgba(100,116,139,.3)!important}
html[data-theme=gold] body{background:#0d0900!important}
html[data-theme=gold] [class*=bg-purple-6]{background:#ca8a04!important}html[data-theme=gold] [class*=bg-purple-7]{background:#a16207!important}html[data-theme=gold] [class*=bg-purple-8]{background:#854d0e!important}html[data-theme=gold] [class*=bg-purple-9]{background:#713f12!important}html[data-theme=gold] [class*=text-purple]{color:#fde047!important}html[data-theme=gold] [class*=border-purple]{border-color:#eab308!important}html[data-theme=gold] [class*=from-purple]{--tw-gradient-from:#713f12!important}html[data-theme=gold] [class*=glow-purple]{box-shadow:0 0 18px rgba(234,179,8,.4)!important}
</style>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body{background:#06060f;font-family:'Segoe UI',system-ui,sans-serif}
  @keyframes heroIn{from{transform:translateY(-30px) scale(.95);opacity:0}to{transform:translateY(0) scale(1);opacity:1}}
  .hero-in{animation:heroIn .42s cubic-bezier(.175,.885,.32,1.275)}
  @keyframes cardIn{from{transform:translateY(-18px) scale(.93);opacity:0}to{transform:translateY(0) scale(1);opacity:1}}
  .card-in{animation:cardIn .32s cubic-bezier(.175,.885,.32,1.275)}
  @keyframes fadeOut{to{transform:scale(.85) translateY(12px);opacity:0;max-height:0;padding:0;margin:0;overflow:hidden}}
  .fade-out{animation:fadeOut .36s ease forwards}
  @keyframes pulseGreen{0%,100%{box-shadow:0 0 0 0 rgba(74,222,128,.45)}60%{box-shadow:0 0 0 10px rgba(74,222,128,0)}}
  .pulse-green{animation:pulseGreen 2s infinite}
  @keyframes badgePulse{0%,100%{transform:scale(1)}50%{transform:scale(1.08)}}
  .priority-badge{animation:badgePulse 2.5s ease infinite}
  @keyframes activeCardIn{from{opacity:0;transform:translateY(-10px) scale(.97)}to{opacity:1;transform:translateY(0) scale(1)}}
  .active-card-in{animation:activeCardIn .35s cubic-bezier(.175,.885,.32,1.275)}
  .queue-mini-card{min-width:148px;max-width:172px;flex-shrink:0}
  .station-col{display:flex;flex-direction:column;border-right:1px solid rgba(31,41,55,.6);min-width:300px;flex:1;height:100%}
  ::-webkit-scrollbar{width:5px}::-webkit-scrollbar-thumb{background:#374151;border-radius:4px}
  input:focus,select:focus{outline:none}
  .tab-active{border-bottom-color:#9333ea!important;color:#c084fc!important}
  .dirty-input{border-color:#eab308!important}
  .dirty-row{background:rgba(234,179,8,.04)!important}
  .view-btn{color:#4b5563;transition:all .15s ease;width:2.25rem;height:2.25rem;border-radius:.5rem;display:flex;align-items:center;justify-content:center}
  .view-btn.active{background:rgba(147,51,234,.85);color:#fff;box-shadow:0 0 14px rgba(147,51,234,.45)}
  .view-btn:not(.active):hover{background:rgba(75,85,99,.45);color:#d1d5db}
</style>
</head>
<body class="min-h-screen text-white">

<!-- ═══ HEADER ═══ -->
<div class="sticky top-0 z-30 bg-gray-950/96 backdrop-blur border-b border-gray-800 px-6 py-4 flex items-center justify-between">
  <div class="flex items-center gap-3">
    <div class="w-10 h-10 rounded-xl bg-purple-600 flex items-center justify-center text-xl shadow-lg shadow-purple-900/60 shrink-0">🍹</div>
    <div>
      <h1 class="font-extrabold text-lg tracking-wider leading-none">DRUNKS <span class="text-purple-400 font-light text-sm">· COCINA</span></h1>
      <div class="flex items-center gap-2 mt-0.5 flex-wrap">
        <div id="wsDot" class="w-2 h-2 rounded-full bg-red-500 shrink-0"></div>
        <span id="wsTxt" class="text-xs text-gray-500">Conectando...</span>
        <span class="text-gray-700 text-xs">·</span>
        <div id="sbDot" class="w-2 h-2 rounded-full bg-gray-600 shrink-0"></div>
        <span id="sbTxt" class="text-xs text-gray-600">Nube...</span>
      </div>
    </div>
  </div>
  <div class="flex items-center gap-3">
    <div class="text-center hidden sm:block">
      <div class="text-3xl font-extrabold text-purple-400 tabular-nums leading-none" id="orderCount">0</div>
      <div class="text-[10px] text-gray-600 tracking-widest mt-0.5">ACTIVOS</div>
    </div>
    <div id="viewModeBar" class="flex items-center bg-gray-800/60 border border-gray-700 rounded-xl p-1 gap-0.5">
      <button class="view-btn" id="vBtn-priority" onclick="setView('priority')" title="Prioridad — #1 destacado">
        <svg viewBox="0 0 16 16" width="15" height="15" fill="currentColor"><rect x="0" y="0" width="16" height="8" rx="1.5"/><rect x="0" y="9.5" width="7.5" height="6.5" rx="1.2"/><rect x="8.5" y="9.5" width="7.5" height="6.5" rx="1.2"/></svg>
      </button>
      <button class="view-btn" id="vBtn-duo" onclick="setView('duo')" title="Doble prioridad — #1 y #2">
        <svg viewBox="0 0 16 16" width="15" height="15" fill="currentColor"><rect x="0" y="0" width="7.5" height="8.5" rx="1.5"/><rect x="8.5" y="0" width="7.5" height="8.5" rx="1.5"/><rect x="0" y="10" width="4.5" height="6" rx="1"/><rect x="5.75" y="10" width="4.5" height="6" rx="1"/><rect x="11.5" y="10" width="4.5" height="6" rx="1"/></svg>
      </button>
      <button class="view-btn" id="vBtn-grid" onclick="setView('grid')" title="Parrilla — todos iguales">
        <svg viewBox="0 0 16 16" width="15" height="15" fill="currentColor"><rect x="0" y="0" width="7.5" height="7.5" rx="1.5"/><rect x="8.5" y="0" width="7.5" height="7.5" rx="1.5"/><rect x="0" y="8.5" width="7.5" height="7.5" rx="1.5"/><rect x="8.5" y="8.5" width="7.5" height="7.5" rx="1.5"/></svg>
      </button>
      <button class="view-btn" id="vBtn-compact" onclick="setView('compact')" title="Compacto — máxima densidad">
        <svg viewBox="0 0 16 16" width="15" height="15" fill="currentColor"><rect x="0" y="0.5" width="16" height="3" rx="1"/><rect x="0" y="4.5" width="16" height="3" rx="1"/><rect x="0" y="8.5" width="16" height="3" rx="1"/><rect x="0" y="12.5" width="16" height="3" rx="1"/></svg>
      </button>
    </div>
    <button onclick="toggleStationPanel()" title="Configurar Preparadores"
      class="bg-gray-800 hover:bg-gray-700 border border-gray-700 w-11 h-11 rounded-xl text-lg flex items-center justify-center transition-colors">👥</button>
    <button onclick="toggleThemePicker()" title="Paleta de colores"
      class="bg-gray-800 hover:bg-gray-700 border border-gray-700 w-11 h-11 rounded-xl text-lg flex items-center justify-center transition-colors">🎨</button>
    <button onclick="openAdmin()" title="Administracion"
      class="bg-gray-800 hover:bg-gray-700 border border-gray-700 w-11 h-11 rounded-xl text-lg flex items-center justify-center transition-colors">⚙️</button>
  </div>
</div>

<!-- ═══ KITCHEN VIEW ═══ -->
<!-- Modo una estacion (comportamiento actual) -->
<div id="singleView" class="p-5">
  <div id="heroSection" class="mb-5"></div>
  <div id="queueLabel" class="hidden flex items-center gap-3 mb-3">
    <div class="h-px flex-1 bg-gray-800"></div>
    <span id="queueTitle" class="text-gray-600 text-[10px] tracking-[0.2em] font-bold shrink-0">EN FILA</span>
    <div class="h-px flex-1 bg-gray-800"></div>
  </div>
  <div id="queueGrid" class="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 gap-3"></div>
</div>
<!-- Modo multi-estacion (columnas) -->
<div id="multiView" class="hidden flex flex-col" style="height:calc(100vh - 5rem)"></div>
<!-- Empty state (compartido) -->
<div id="emptyState" class="hidden flex flex-col items-center justify-center py-40 text-gray-800">
  <div class="text-8xl mb-6 opacity-15">🍹</div>
  <p class="text-2xl font-bold">Todo listo</p>
  <p class="text-sm mt-2 text-gray-700">Los pedidos aparecen aqui en tiempo real</p>
</div>

<!-- ═══ STATION PANEL ═══ -->
<div id="stationPanel" class="hidden fixed z-[91]" style="top:4.5rem;right:1rem">
  <div class="bg-gray-900 border border-gray-700 rounded-2xl p-5 shadow-2xl shadow-black/70 w-80">
    <div class="flex items-center justify-between mb-4">
      <div class="text-gray-400 text-[10px] tracking-[.2em] font-bold">PREPARADORES / ESTACIONES</div>
      <button onclick="toggleStationPanel()" class="text-gray-500 hover:text-white text-xl leading-none">x</button>
    </div>
    <div class="mb-4">
      <div class="text-gray-400 text-xs font-semibold mb-2">Numero de estaciones</div>
      <div class="flex gap-2">
        <button onclick="setStationCount(1)" id="sCnt-1" class="station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border border-gray-700 bg-gray-800 text-gray-400 transition-colors">1</button>
        <button onclick="setStationCount(2)" id="sCnt-2" class="station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border border-gray-700 bg-gray-800 text-gray-400 transition-colors">2</button>
        <button onclick="setStationCount(3)" id="sCnt-3" class="station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border border-gray-700 bg-gray-800 text-gray-400 transition-colors">3</button>
        <button onclick="setStationCount(4)" id="sCnt-4" class="station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border border-gray-700 bg-gray-800 text-gray-400 transition-colors">4</button>
      </div>
    </div>
    <div id="stationFormBlock" class="space-y-3 mb-4 max-h-64 overflow-y-auto pr-1"></div>
    <button onclick="saveStationConfig()"
      class="w-full bg-purple-600 hover:bg-purple-500 text-white font-bold py-2.5 rounded-xl text-sm transition-colors">
      Aplicar configuracion
    </button>
  </div>
</div>

<!-- ═══ THEME PICKER ═══ -->
<div id="themePicker" class="hidden fixed z-[90]" style="top:4.5rem;right:1rem">
  <div class="bg-gray-900 border border-gray-800 rounded-2xl p-4 shadow-2xl shadow-black/60 w-56 max-h-[85vh] overflow-y-auto">
    <div class="text-gray-500 text-[10px] tracking-[.2em] font-bold mb-3">PALETA DE COLORES</div>
    <div class="space-y-0.5">
      <button data-theme="dark"    onclick="applyTheme('dark')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#06060f"></span><span class="text-white text-sm font-semibold">Oscuro</span></button>
      <button data-theme="light"   onclick="applyTheme('light')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#f0ebff"></span><span class="text-white text-sm font-semibold">Claro</span></button>
      <button data-theme="ocean"   onclick="applyTheme('ocean')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#2563eb"></span><span class="text-white text-sm font-semibold">Océano</span></button>
      <button data-theme="emerald" onclick="applyTheme('emerald')" class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#0d9488"></span><span class="text-white text-sm font-semibold">Esmeralda</span></button>
      <button data-theme="amber"   onclick="applyTheme('amber')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#d97706"></span><span class="text-white text-sm font-semibold">Ámbar</span></button>
      <button data-theme="neon"    onclick="applyTheme('neon')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#16a34a"></span><span class="text-white text-sm font-semibold">Neón</span></button>
      <button data-theme="rose"    onclick="applyTheme('rose')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#e11d48"></span><span class="text-white text-sm font-semibold">Rosa</span></button>
      <button data-theme="violet"  onclick="applyTheme('violet')"  class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#6d28d9"></span><span class="text-white text-sm font-semibold">Violeta</span></button>
      <button data-theme="slate"   onclick="applyTheme('slate')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#475569"></span><span class="text-white text-sm font-semibold">Gris Pizarra</span></button>
      <button data-theme="gold"    onclick="applyTheme('gold')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors"><span class="w-5 h-5 rounded-full shrink-0" style="background:#ca8a04"></span><span class="text-white text-sm font-semibold">Dorado</span></button>
    </div>
    <div class="mt-3 pt-3 border-t border-gray-800">
      <div class="text-gray-500 text-[10px] font-bold mb-2">COLOR PERSONALIZADO</div>
      <div class="flex items-center gap-2">
        <input type="color" id="customColorPicker" value="#9333ea"
          oninput="applyCustomTheme(this.value)"
          style="display:none;width:2rem;height:2rem;border-radius:.5rem;border:none;cursor:pointer;background:transparent;padding:0">
        <button onclick="document.getElementById('customColorPicker').style.display='inline-block';document.getElementById('customColorPicker').click()"
          class="flex items-center gap-2 px-3 py-1.5 bg-gray-800 hover:bg-gray-700 rounded-xl text-xs text-gray-300 transition-colors">
          <span>🎨</span> Elegir color
        </button>
      </div>
    </div>
  </div>
</div>

<!-- ═══ ADMIN MODAL ═══ -->
<div id="adminModal" class="fixed inset-0 bg-black/82 backdrop-blur-md z-50 hidden items-center justify-center p-4">
  <div class="bg-gray-900 border border-gray-700 rounded-2xl w-full max-w-5xl max-h-[90vh] flex flex-col shadow-2xl">
    <!-- Header -->
    <div class="flex items-center justify-between px-6 py-5 border-b border-gray-700 shrink-0">
      <h2 class="text-xl font-extrabold">Administracion General</h2>
      <button onclick="closeAdmin()" class="text-gray-400 hover:text-white text-3xl leading-none transition-colors">x</button>
    </div>
    <!-- Tabs -->
    <div class="flex border-b border-gray-800 px-6 shrink-0">
      <button id="tab-prods" onclick="showTab('prods')" class="tab-active tab-btn px-5 py-3.5 text-sm font-semibold border-b-2 border-transparent text-gray-400 hover:text-white transition-colors mr-1">Productos</button>
      <button id="tab-cats"  onclick="showTab('cats')"  class="tab-btn px-5 py-3.5 text-sm font-semibold border-b-2 border-transparent text-gray-400 hover:text-white transition-colors mr-1">Categorias</button>
      <button id="tab-notes" onclick="showTab('notes')" class="tab-btn px-5 py-3.5 text-sm font-semibold border-b-2 border-transparent text-gray-400 hover:text-white transition-colors">Notas Rapidas</button>
      <button id="tab-bases" onclick="showTab('bases')" class="tab-btn px-5 py-3.5 text-sm font-semibold border-b-2 border-transparent text-gray-400 hover:text-white transition-colors ml-1">Bases</button>
    </div>
    <!-- Toast -->
    <div id="adminMsg" class="hidden mx-6 mt-4 text-sm px-4 py-2 rounded-xl shrink-0"></div>
    <!-- Body -->
    <div class="flex-1 overflow-y-auto px-6 pt-4 pb-2">

      <!-- ── Productos ── -->
      <div id="pane-prods">
        <div class="flex justify-end mb-4">
          <button onclick="addProd()" class="bg-purple-600 hover:bg-purple-500 text-white px-4 py-2 rounded-xl text-sm font-bold transition-colors">+ Nuevo Producto</button>
        </div>
        <div class="overflow-x-auto">
          <table class="w-full text-sm min-w-[580px]">
            <thead>
              <tr class="text-gray-500 text-[10px] tracking-widest border-b border-gray-800">
                <th class="text-left py-2 pr-3">NOMBRE</th>
                <th class="text-left py-2 pr-3">CATEGORIA</th>
                <th class="text-left py-2 pr-3 w-28">PRECIO</th>
                <th class="text-center py-2 pr-3 w-24">ESTADO</th>
                <th class="text-center py-2 pr-3 w-24">BASE</th>
                <th class="text-center py-2 w-12">DEL</th>
              </tr>
            </thead>
            <tbody id="prodsTable"></tbody>
          </table>
        </div>
      </div>

      <!-- ── Categorias ── -->
      <div id="pane-cats" class="hidden">
        <div class="flex gap-2 mb-5">
          <input id="newCatName" type="text" placeholder="Nueva categoria..."
            class="flex-1 bg-gray-800 border border-gray-700 focus:border-purple-500 rounded-xl px-3 py-2.5 text-white text-sm transition-colors">
          <button onclick="addCat()" class="bg-purple-600 hover:bg-purple-500 text-white px-5 py-2.5 rounded-xl text-sm font-bold shrink-0">Agregar</button>
        </div>
        <div id="catsList" class="space-y-2"></div>
      </div>

      <!-- ── Notas ── -->
      <div id="pane-notes" class="hidden">
        <p class="text-gray-500 text-sm mb-4">Las notas con categoria aparecen dinamicamente al vendedor segun el producto seleccionado.</p>
        <div class="flex gap-2 mb-5 flex-wrap sm:flex-nowrap">
          <input id="newNoteText" type="text" placeholder="Texto de la nota..."
            class="flex-1 bg-gray-800 border border-gray-700 focus:border-purple-500 rounded-xl px-3 py-2.5 text-white text-sm min-w-0 transition-colors">
          <select id="newNoteCat" class="bg-gray-800 border border-gray-700 focus:border-purple-500 rounded-xl px-3 py-2.5 text-white text-sm shrink-0 transition-colors">
            <option value="">Global</option>
          </select>
          <button onclick="addNote()" class="bg-purple-600 hover:bg-purple-500 text-white px-5 py-2.5 rounded-xl text-sm font-bold shrink-0">Agregar</button>
        </div>
        <div id="notesList" class="flex flex-wrap gap-2.5"></div>
      </div>

      <!-- ── Bases ── -->
      <div id="pane-bases" class="hidden">
        <!-- Sección 1: Tipos de Base -->
        <div class="mb-6">
          <div class="flex items-center justify-between mb-3">
            <span class="text-white font-bold text-sm">Tipos de Base</span>
            <button onclick="toggleNewTipoForm()" class="text-xs bg-gray-800 hover:bg-purple-700 text-gray-300 hover:text-white px-3 py-1.5 rounded-lg transition-colors">+ Nuevo tipo</button>
          </div>
          <!-- Form nuevo tipo (oculto por defecto) -->
          <div id="newTipoForm" class="hidden flex gap-2 mb-3 p-3 bg-gray-800/50 rounded-xl border border-gray-700">
            <input id="newTipoIcono" type="text" value="🍹" maxlength="4"
              class="bg-gray-900 border border-gray-700 focus:border-purple-500 rounded-lg px-2.5 py-2 text-white text-sm w-14 text-center transition-colors" placeholder="🍹">
            <input id="newTipoNombre" type="text" placeholder="Nombre del tipo (ej: Ron)"
              class="flex-1 bg-gray-900 border border-gray-700 focus:border-purple-500 rounded-lg px-3 py-2 text-white text-sm transition-colors"
              onkeydown="if(event.key==='Enter')addTipoBase()">
            <button onclick="addTipoBase()" class="bg-purple-600 hover:bg-purple-500 text-white px-4 py-2 rounded-lg text-sm font-bold transition-colors">Crear</button>
            <button onclick="toggleNewTipoForm()" class="bg-gray-700 hover:bg-gray-600 text-gray-300 px-3 py-2 rounded-lg text-sm transition-colors">✕</button>
          </div>
          <!-- Chips de tipos -->
          <div id="tipoChips" class="flex flex-wrap gap-2"></div>
        </div>
        <!-- Sección 2: Items del tipo seleccionado -->
        <div id="baseItemsSection" class="hidden">
          <div class="flex items-center justify-between mb-3 border-t border-gray-800 pt-5">
            <div class="flex items-center gap-2">
              <span id="baseItemsTipoLabel" class="text-white font-bold text-sm"></span>
              <span class="text-gray-500 text-xs">— items de base</span>
            </div>
            <div class="flex gap-2">
              <input id="newBaseNombre" type="text" placeholder="Nombre del item..."
                class="bg-gray-800 border border-gray-700 focus:border-purple-500 rounded-lg px-3 py-1.5 text-white text-sm w-40 transition-colors"
                onkeydown="if(event.key==='Enter')addBaseItem()">
              <button onclick="addBaseItem()" class="bg-purple-600 hover:bg-purple-500 text-white px-3 py-1.5 rounded-lg text-sm font-bold transition-colors">+ Agregar</button>
            </div>
          </div>
          <div id="basesList" class="mb-2"></div>
        </div>
        <!-- Estado vacío -->
        <div id="basesEmptyState" class="hidden text-gray-600 text-sm text-center py-8">
          Crea un tipo de base para empezar (ej: Gaseosa, Cerveza, Ron...)
        </div>
      </div>

    </div>
    <!-- ── SAVE BAR (single-save) ── -->
    <div id="saveBar" class="hidden shrink-0 border-t border-yellow-700/30 bg-gray-900/90 rounded-b-2xl px-6 py-3 flex items-center justify-between">
      <div class="flex items-center gap-2.5">
        <div class="w-2 h-2 rounded-full bg-yellow-400 animate-pulse shrink-0"></div>
        <span id="saveCount" class="text-yellow-400 text-sm font-semibold"></span>
      </div>
      <div class="flex gap-2">
        <button onclick="discardAll()" class="bg-gray-700 hover:bg-gray-600 text-gray-300 px-4 py-2 rounded-xl text-sm font-medium transition-colors">Descartar</button>
        <button id="saveAllBtn" onclick="saveAll()" class="bg-green-600 hover:bg-green-500 text-white px-5 py-2 rounded-xl text-sm font-bold transition-colors shadow-lg">Guardar Todo</button>
      </div>
    </div>
  </div>
</div>

<script>
// ─── STATE ───
let ws = null, wsTimer = null;
let orders = {};
let viewMode = localStorage.getItem('kitchenView') || 'priority';
const dirty = { prods: new Set(), cats: new Set() };
let adminCats = [];

// ─── WEBSOCKET ───
function connectWS() {
  if (ws && ws.readyState < 2) return;
  const _wsProto = location.protocol === 'https:' ? 'wss:' : 'ws:';
  ws = new WebSocket(_wsProto + '//' + location.host + '/ws');
  ws.onopen = () => {
    document.getElementById('wsDot').className = 'w-2 h-2 rounded-full bg-green-500 shrink-0 pulse-green';
    document.getElementById('wsTxt').textContent = 'Conectado en vivo';
    if (wsTimer) { clearTimeout(wsTimer); wsTimer = null; }
  };
  ws.onmessage = e => { try { const d = JSON.parse(e.data); if (d.type === 'new_order') addOrder(d.order); } catch(_){} };
  ws.onclose = ws.onerror = () => {
    document.getElementById('wsDot').className = 'w-2 h-2 rounded-full bg-red-500 shrink-0';
    document.getElementById('wsTxt').textContent = 'Reconectando...';
    wsTimer = setTimeout(connectWS, 3000);
  };
}

// ─── KITCHEN RENDER ───
function addOrder(order) {
  if (orders[order.id]) return;
  orders[order.id] = order;
  renderKitchen();
}

function renderKitchen() {

  if (stationCfg.count >= 2) { renderMultiStation(); return; }

  const list    = Object.values(orders).sort((a, b) => a.id - b.id); // FIFO
  const heroSec = document.getElementById('heroSection');
  const qLabel  = document.getElementById('queueLabel');
  const qTitle  = document.getElementById('queueTitle');
  const qGrid   = document.getElementById('queueGrid');
  const empty   = document.getElementById('emptyState');

  updateCount(list.length);

  if (!list.length) {
    heroSec.innerHTML = ''; qGrid.innerHTML = '';
    qLabel.classList.add('hidden');
    empty.classList.remove('hidden');
    return;
  }
  empty.classList.add('hidden');

  const hide = () => { qLabel.classList.add('hidden'); };
  const showLabel = (n) => {
    qLabel.classList.remove('hidden');
    qTitle.textContent = 'EN FILA — ' + n + ' pedido' + (n !== 1 ? 's' : '');
  };

  if (viewMode === 'priority') {
    // ── Modo Prioridad: #1 hero full-width, resto grilla ──
    heroSec.innerHTML = '';
    heroSec.appendChild(buildHeroCard(list[0]));
    qGrid.className = 'grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 gap-3';
    qGrid.innerHTML = '';
    if (list.length > 1) {
      showLabel(list.length - 1);
      list.slice(1).forEach(o => qGrid.appendChild(buildQueueCard(o)));
    } else { hide(); }

  } else if (viewMode === 'duo') {
    // ── Modo Doble: top-2 medianos lado a lado, resto grilla compacta ──
    heroSec.innerHTML = '';
    if (list.length === 1) {
      heroSec.appendChild(buildHeroCard(list[0]));
      hide(); qGrid.innerHTML = '';
    } else {
      const duoWrap = document.createElement('div');
      duoWrap.className = 'grid grid-cols-1 sm:grid-cols-2 gap-4 mb-4';
      duoWrap.appendChild(buildMediumCard(list[0], 1));
      duoWrap.appendChild(buildMediumCard(list[1], 2));
      heroSec.appendChild(duoWrap);
      qGrid.className = 'grid grid-cols-2 sm:grid-cols-3 lg:grid-cols-4 xl:grid-cols-5 gap-2';
      qGrid.innerHTML = '';
      if (list.length > 2) {
        showLabel(list.length - 2);
        list.slice(2).forEach(o => qGrid.appendChild(buildQueueCard(o)));
      } else { hide(); }
    }

  } else if (viewMode === 'grid') {
    // ── Modo Parrilla: todos iguales, sin jerarquía ──
    heroSec.innerHTML = ''; hide();
    qGrid.className = 'grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 gap-3';
    qGrid.innerHTML = '';
    list.forEach(o => qGrid.appendChild(buildQueueCard(o)));

  } else if (viewMode === 'compact') {
    // ── Modo Compacto: máxima densidad ──
    heroSec.innerHTML = ''; hide();
    qGrid.className = 'grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 xl:grid-cols-4 2xl:grid-cols-5 gap-2';
    qGrid.innerHTML = '';
    list.forEach(o => qGrid.appendChild(buildCompactCard(o)));
  }
}

function buildItemRow(it, big) {
  const parts = (it.observaciones||'').split(' | ');
  const isBase = s => s && /^[^:]+: /.test(s);
  const base  = isBase(parts[0]) ? parts[0] : null;
  const notas = base ? parts.slice(1).join(' | ') : (it.observaciones||'');
  return `<div class="flex items-start gap-3 py-${big?3:2} border-b border-gray-800/70 last:border-0">
    <span class="bg-purple-700 text-white ${big?'text-sm':'text-xs'} font-extrabold px-2.5 py-0.5 rounded-lg shrink-0 mt-0.5">${it.cantidad}x</span>
    <div class="flex-1">
      <div class="text-white font-bold ${big?'text-base':'text-sm'}">${it.nombre}</div>
      ${base  ? '<div class="text-amber-400 font-bold text-xs mt-0.5">📌 ' + base + '</div>' : ''}
      ${notas ? '<div class="text-yellow-300 font-extrabold text-' + (big?'sm':'xs') + ' mt-0.5">⚠ ' + notas + '</div>' : ''}
    </div>
  </div>`;
}

function buildHeroCard(o) {
  const div = document.createElement('div');
  div.id = 'order-' + o.id;
  div.className = 'hero-in';
  const efect = o.metodo_pago === 'Efectivo';
  const hBg   = efect ? 'from-green-700 to-green-800' : 'from-blue-700 to-blue-800';
  const icon  = efect ? '💵' : '📱';
  const tStr  = new Date(o.fecha||Date.now()).toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  const rows  = (o.items||[]).map(it => buildItemRow(it, true)).join('');
  div.innerHTML = `
    <div class="flex items-center gap-3 mb-3">
      <div class="priority-badge bg-red-500 text-white text-xs font-extrabold px-3 py-1 rounded-full tracking-wider shadow-lg shadow-red-900/50">#1 PRIORIDAD</div>
      <div class="h-px flex-1 bg-gray-800"></div>
    </div>
    <div class="bg-gray-900 border-2 border-purple-500/40 rounded-3xl overflow-hidden shadow-2xl shadow-purple-950/30">
      <div class="bg-gradient-to-r ${hBg} px-6 py-5 flex items-start justify-between">
        <div>
          <div class="flex items-center gap-2 mb-1.5">
            <span class="bg-white/20 text-white text-[10px] font-extrabold px-2 py-0.5 rounded-full tracking-wider">${o.numero_factura||('#'+o.id)}</span>
            <span class="text-white/60 text-xs">${icon} ${o.metodo_pago} &middot; ${tStr}</span>
          </div>
          <div class="text-white font-extrabold text-4xl leading-none">${o.cliente}</div>
        </div>
        <div class="text-white font-extrabold text-4xl tabular-nums leading-none">$${Math.round(o.total).toLocaleString('es-CO')}</div>
      </div>
      <div class="px-6 py-5">${rows||'<div class="text-gray-600 text-sm text-center py-4">Sin detalle</div>'}</div>
      <div class="px-6 pb-6">
        <button data-id="${o.id}" onclick="completeOrder(parseInt(this.dataset.id))"
          class="w-full bg-green-600 hover:bg-green-500 active:bg-green-700 text-white font-extrabold py-5 rounded-2xl text-xl transition-colors shadow-xl shadow-green-950/40">
          ✓ COMPLETADO
        </button>
      </div>
    </div>`;
  return div;
}

function buildQueueCard(o) {
  const div = document.createElement('div');
  div.id = 'order-' + o.id;
  div.className = 'card-in bg-gray-900 border border-gray-700 rounded-2xl overflow-hidden flex flex-col';
  const efect = o.metodo_pago === 'Efectivo';
  const hBg   = efect ? 'from-green-800/60 to-green-900/60' : 'from-blue-800/60 to-blue-900/60';
  const icon  = efect ? '💵' : '📱';
  const tStr  = new Date(o.fecha||Date.now()).toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  const rows  = (o.items||[]).map(it => buildItemRow(it, false)).join('');
  div.innerHTML = `
    <div class="bg-gradient-to-r ${hBg} px-4 py-3 flex justify-between items-start">
      <div>
        <div class="text-white font-extrabold text-xl leading-tight">${o.cliente}</div>
        <div class="flex items-center gap-1.5 mt-0.5">
          <span class="bg-white/15 text-white/80 text-[9px] font-bold px-1.5 py-0.5 rounded-full">${o.numero_factura||('#'+o.id)}</span>
          <span class="text-white/50 text-xs">${icon} ${o.metodo_pago} &middot; ${tStr}</span>
        </div>
      </div>
      <div class="text-white font-extrabold text-lg tabular-nums">$${Math.round(o.total).toLocaleString('es-CO')}</div>
    </div>
    <div class="flex-1 px-4 py-3">${rows||'<div class="text-gray-600 text-xs text-center py-3">Sin detalle</div>'}</div>
    <div class="px-4 pb-4">
      <button data-id="${o.id}" onclick="completeOrder(parseInt(this.dataset.id))"
        class="w-full bg-green-700 hover:bg-green-600 active:bg-green-800 text-white font-bold py-3 rounded-xl text-sm transition-colors">
        ✓ Completado
      </button>
    </div>`;
  return div;
}

function buildMediumCard(o, priority) {
  const div   = document.createElement('div');
  div.id      = 'order-' + o.id;
  div.className = 'hero-in';
  const efect = o.metodo_pago === 'Efectivo';
  const hBg   = efect ? 'from-green-700 to-green-800' : 'from-blue-700 to-blue-800';
  const icon  = efect ? '💵' : '📱';
  const tStr  = new Date(o.fecha||Date.now()).toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  const rows  = (o.items||[]).map(it => buildItemRow(it, false)).join('');
  const isFirst = priority === 1;
  const badgeCls  = isFirst ? 'priority-badge bg-red-500 shadow-red-900/50' : 'bg-orange-600 shadow-orange-900/50';
  const borderCls = isFirst ? 'border-purple-500/40' : 'border-orange-500/30';
  const badgeTxt  = isFirst ? '#1 PRIORIDAD' : '#2 EN CURSO';
  div.innerHTML = `
    <div class="flex items-center gap-2 mb-2">
      <div class="${badgeCls} text-white text-xs font-extrabold px-2.5 py-0.5 rounded-full tracking-wider shadow-lg">${badgeTxt}</div>
      <div class="h-px flex-1 bg-gray-800"></div>
    </div>
    <div class="bg-gray-900 border-2 ${borderCls} rounded-2xl overflow-hidden shadow-xl h-full flex flex-col">
      <div class="bg-gradient-to-r ${hBg} px-4 py-3 flex items-start justify-between">
        <div class="min-w-0">
          <div class="flex items-center gap-2 mb-1">
            <span class="bg-white/20 text-white text-[10px] font-extrabold px-2 py-0.5 rounded-full tracking-wider">${o.numero_factura||('#'+o.id)}</span>
            <span class="text-white/60 text-xs">${icon} ${o.metodo_pago} &middot; ${tStr}</span>
          </div>
          <div class="text-white font-extrabold text-2xl leading-tight truncate">${o.cliente}</div>
        </div>
        <div class="text-white font-extrabold text-xl tabular-nums ml-3 shrink-0">$${Math.round(o.total).toLocaleString('es-CO')}</div>
      </div>
      <div class="flex-1 px-4 py-3">${rows||'<div class="text-gray-600 text-sm text-center py-3">Sin detalle</div>'}</div>
      <div class="px-4 pb-4">
        <button data-id="${o.id}" onclick="completeOrder(parseInt(this.dataset.id))"
          class="w-full bg-green-600 hover:bg-green-500 active:bg-green-700 text-white font-extrabold py-3 rounded-xl text-base transition-colors shadow-lg shadow-green-950/40">
          ✓ COMPLETADO
        </button>
      </div>
    </div>`;
  return div;
}

function buildCompactCard(o) {
  const div   = document.createElement('div');
  div.id      = 'order-' + o.id;
  div.className = 'card-in bg-gray-900 border border-gray-700/60 rounded-xl overflow-hidden';
  const efect = o.metodo_pago === 'Efectivo';
  const hBg   = efect ? 'from-green-800/55 to-green-900/55' : 'from-blue-800/55 to-blue-900/55';
  const icon  = efect ? '💵' : '📱';
  const tStr  = new Date(o.fecha||Date.now()).toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  const itemRows = (o.items||[]).map(it => {
    const parts = (it.observaciones||'').split(' | ');
    const isBase = s => s && /^[^:]+: /.test(s);
    const extra = isBase(parts[0]) ? parts[0] : (it.observaciones||'');
    return `<div class="flex items-center gap-1 py-0.5 min-w-0">
      <span class="text-purple-400 text-[10px] font-extrabold shrink-0">${it.cantidad}x</span>
      <span class="text-white text-xs font-semibold truncate">${it.nombre}</span>
      ${extra ? `<span class="text-amber-400 text-[10px] shrink-0 truncate">· ${extra}</span>` : ''}
    </div>`;
  }).join('');
  div.innerHTML = `
    <div class="bg-gradient-to-r ${hBg} px-3 py-2 flex items-center justify-between gap-2">
      <div class="min-w-0">
        <div class="text-white font-extrabold text-base leading-tight truncate">${o.cliente}</div>
        <div class="text-white/50 text-[10px] leading-none mt-0.5">${o.numero_factura||('#'+o.id)} &middot; ${icon} ${tStr}</div>
      </div>
      <div class="text-white font-extrabold text-sm tabular-nums shrink-0">$${Math.round(o.total).toLocaleString('es-CO')}</div>
    </div>
    <div class="px-3 py-2 space-y-0.5">${itemRows||'<div class="text-gray-700 text-xs py-1">Sin detalle</div>'}</div>
    <div class="px-3 pb-3">
      <button data-id="${o.id}" onclick="completeOrder(parseInt(this.dataset.id))"
        class="w-full bg-green-700 hover:bg-green-600 active:bg-green-800 text-white font-bold py-1.5 rounded-lg text-xs transition-colors">
        ✓ Listo
      </button>
    </div>`;
  return div;
}

function setView(mode) {
  if (stationCfg && stationCfg.count >= 2) return; // no-op en modo multi-estacion
  viewMode = mode;
  saveSetting('kitchenView', mode);
  saveSetting('kitchen_view', mode);
  document.querySelectorAll('.view-btn').forEach(b =>
    b.classList.toggle('active', b.id === 'vBtn-' + mode));
  renderKitchen();
}

async function completeOrder(id) {
  try {
    const r = await fetch('/api/pedidos/'+id+'/entregar', {method:'PUT'});
    if (r.ok) {
      const card = document.getElementById('order-'+id);
      if (card) {
        card.classList.add('fade-out');
        setTimeout(() => { delete orders[id]; renderKitchen(); }, 370);
      } else {
        delete orders[id]; renderKitchen();
      }
    }
  } catch(e) {}
}

function updateCount(n) {
  if (n === undefined) n = Object.keys(orders).length;
  document.getElementById('orderCount').textContent = n;
}

async function loadPending() {
  try {
    const r    = await fetch('/api/pedidos/pendientes');
    const list = await r.json();
    list.forEach(o => { orders[o.id] = o; });
    renderKitchen();
  } catch(e) {}
}

// ─── ADMIN MODAL ───
function openAdmin() {
  document.getElementById('adminModal').classList.remove('hidden');
  document.getElementById('adminModal').classList.add('flex');
  showTab('prods');
}
function closeAdmin() {
  if (dirty.prods.size + dirty.cats.size > 0) {
    if (!confirm('Tienes cambios sin guardar. Cerrar de todas formas?')) return;
    dirty.prods.clear(); dirty.cats.clear(); updateSaveBar();
  }
  document.getElementById('adminModal').classList.add('hidden');
  document.getElementById('adminModal').classList.remove('flex');
}
function showTab(tab) {
  if (typeof closeBasePop === 'function') closeBasePop();
  ['prods','cats','notes','bases'].forEach(t => {
    document.getElementById('pane-'+t).classList.add('hidden');
    const b = document.getElementById('tab-'+t);
    b.classList.remove('tab-active'); b.style.borderBottomColor='transparent'; b.style.color='';
  });
  document.getElementById('pane-'+tab).classList.remove('hidden');
  document.getElementById('tab-'+tab).classList.add('tab-active');
  if (tab === 'prods')  loadProds();
  if (tab === 'cats')   loadCats();
  if (tab === 'notes')  loadNotes();
  if (tab === 'bases')  loadBases();
}
function adminToast(msg, ok=true) {
  const el = document.getElementById('adminMsg');
  el.textContent = msg;
  el.className = 'mx-6 mt-4 shrink-0 text-sm px-4 py-2 rounded-xl ' +
    (ok ? 'bg-green-800/50 border border-green-700 text-green-300' : 'bg-red-900/50 border border-red-700 text-red-300');
  el.classList.remove('hidden');
  setTimeout(() => el.classList.add('hidden'), 2800);
}
function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

// ─── DIRTY TRACKING (single-save) ───
function markDirty(type, id) {
  dirty[type].add(id);
  // Visual indicator on the row
  const row = document.getElementById('drow-'+id);
  if (row) { row.classList.add('dirty-row'); }
  const inputs = document.querySelectorAll('[data-dirty-id="'+id+'"]');
  inputs.forEach(el => el.classList.add('dirty-input'));
  updateSaveBar();
}

function updateSaveBar() {
  const total = dirty.prods.size + dirty.cats.size;
  const bar = document.getElementById('saveBar');
  if (total > 0) {
    bar.classList.remove('hidden');
    document.getElementById('saveCount').textContent =
      total + ' cambio' + (total !== 1 ? 's' : '') + ' sin guardar';
  } else {
    bar.classList.add('hidden');
  }
}

async function saveAll() {
  const btn = document.getElementById('saveAllBtn');
  btn.disabled = true; btn.textContent = 'Guardando...';
  let ok = 0, errors = 0;

  for (const id of [...dirty.prods]) {
    const n = document.getElementById('pn-'+id)?.value.trim();
    const c = parseInt(document.getElementById('pc-'+id)?.value);
    const p = parseFloat(document.getElementById('pp-'+id)?.value);
    if (!n) { errors++; continue; }
    try {
      const r = await fetch('/api/productos/'+id, {
        method:'PUT', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({nombre:n, categoria_id:c, precio:p})
      });
      if (r.ok) { dirty.prods.delete(id); ok++; }
      else errors++;
    } catch(e) { errors++; }
  }

  for (const id of [...dirty.cats]) {
    const n = document.getElementById('cn-'+id)?.value.trim();
    if (!n) { errors++; continue; }
    try {
      const r = await fetch('/api/categorias/'+id, {
        method:'PUT', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({nombre:n})
      });
      if (r.ok) { dirty.cats.delete(id); ok++; }
      else errors++;
    } catch(e) { errors++; }
  }

  updateSaveBar();
  if (errors === 0) adminToast('Todo guardado correctamente (' + ok + ' cambios)');
  else adminToast(errors + ' errores. ' + ok + ' guardados.', false);
  btn.disabled = false; btn.textContent = 'Guardar Todo';
}

function discardAll() {
  dirty.prods.clear(); dirty.cats.clear();
  updateSaveBar();
  const pane = document.querySelector('[id^="pane-"]:not(.hidden)');
  if (pane?.id === 'pane-prods') loadProds();
  if (pane?.id === 'pane-cats')  loadCats();
}

// ─── PRODUCTOS TAB ───
let adminBases = [];   // cache de bases para el selector
let prodBaseMap = {};  // { productoId: [baseId,...] }
let openBasePopover = null; // pid del popover abierto

function buildBasesPopover(pid, assignedIds) {
  // Agrupa bases por tipo para mostrar checkboxes organizados
  const byTipo = {};
  adminBases.forEach(b => {
    const t = b.tipo_nombre || b.tipo || 'Otros';
    if (!byTipo[t]) byTipo[t] = { icono: b.tipo_icono || '🍹', items: [] };
    byTipo[t].items.push(b);
  });
  const groups = Object.entries(byTipo).map(([tipo, { icono, items }]) => `
    <div class="mb-2">
      <div class="text-[10px] font-bold text-gray-500 mb-1">${icono} ${tipo}</div>
      ${items.map(b => `
        <label class="flex items-center gap-1.5 px-1 py-0.5 rounded hover:bg-gray-700 cursor-pointer">
          <input type="checkbox" value="${b.id}" ${assignedIds.includes(b.id)?'checked':''}
            onchange="onBaseCheckChange(${pid},this)"
            class="accent-purple-500 w-3 h-3">
          <span class="text-white text-xs">${esc(b.nombre)}</span>
        </label>`).join('')}
    </div>`).join('');
  return `<div id="basepop-${pid}" class="absolute z-50 right-0 top-full mt-1 bg-gray-900 border border-gray-700 rounded-xl p-3 shadow-2xl min-w-[160px] max-h-64 overflow-y-auto">
    ${groups || '<div class="text-gray-600 text-xs">Sin bases. Crea en tab Bases.</div>'}
    <button onclick="closeBasePop()" class="mt-2 w-full text-[10px] text-gray-500 hover:text-gray-300 transition-colors">Cerrar</button>
  </div>`;
}

function openBasePop(pid) {
  // Cierra cualquier popover abierto
  closeBasePop();
  openBasePopover = pid;
  const td = document.getElementById('basetd-'+pid);
  if (!td) return;
  td.style.position = 'relative';
  const assignedIds = prodBaseMap[pid] || [];
  td.insertAdjacentHTML('beforeend', buildBasesPopover(pid, assignedIds));
  // Cierra al clic fuera
  setTimeout(() => document.addEventListener('click', outsideBasePop), 0);
}

function outsideBasePop(e) {
  if (!e.target.closest('[id^="basepop-"]') && !e.target.closest('[onclick^="openBasePop"]')) {
    closeBasePop();
  }
}

function closeBasePop() {
  if (openBasePopover !== null) {
    document.getElementById('basepop-'+openBasePopover)?.remove();
    openBasePopover = null;
    document.removeEventListener('click', outsideBasePop);
  }
}

async function onBaseCheckChange(pid, cb) {
  const td = document.getElementById('basetd-'+pid);
  const cbs = td?.querySelectorAll('input[type="checkbox"]');
  const selectedIds = [...(cbs||[])].filter(c=>c.checked).map(c=>parseInt(c.value));
  prodBaseMap[pid] = selectedIds;
  // Actualizar etiqueta
  const lbl = document.getElementById('baselbl-'+pid);
  if (lbl) {
    lbl.className = selectedIds.length ? 'text-amber-400 text-[10px] font-bold' : 'text-gray-600 text-[10px]';
    lbl.textContent = selectedIds.length ? `🍹 ${selectedIds.length} base${selectedIds.length>1?'s':''}` : 'Sin base';
  }
  try {
    await fetch('/api/productos/'+pid+'/bases', {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({base_ids: selectedIds})
    });
    await fetch('/api/productos/'+pid, {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({tiene_base: selectedIds.length > 0 ? 1 : 0})
    });
  } catch(e) {}
}

async function loadProds() {
  try {
    const [pr, cr, br, bm] = await Promise.all([
      fetch('/api/productos').then(r=>r.json()),
      fetch('/api/categorias').then(r=>r.json()),
      fetch('/api/bases').then(r=>r.json()),
      fetch('/api/productos/bases-bulk').then(r=>r.json()),
    ]);
    adminCats = cr; adminBases = br; prodBaseMap = bm;
    document.getElementById('prodsTable').innerHTML = pr.map(p => {
      const assignedIds = prodBaseMap[p.id] || [];
      const baseLabel = assignedIds.length
        ? `<span id="baselbl-${p.id}" class="text-amber-400 text-[10px] font-bold">🍹 ${assignedIds.length} base${assignedIds.length>1?'s':''}</span>`
        : `<span id="baselbl-${p.id}" class="text-gray-600 text-[10px]">Sin base</span>`;
      return `
      <tr id="drow-${p.id}" class="border-b border-gray-800/50 hover:bg-gray-800/10 transition-colors ${dirty.prods.has(p.id)?'dirty-row':''}">
        <td class="py-2.5 pr-3">
          <input id="pn-${p.id}" data-dirty-id="${p.id}" type="text" value="${esc(p.nombre)}"
            oninput="markDirty('prods',${p.id})"
            class="bg-gray-800 border ${dirty.prods.has(p.id)?'border-yellow-500 dirty-input':'border-gray-700'} focus:border-purple-500 rounded-lg px-2.5 py-1.5 text-white text-sm w-full transition-colors">
        </td>
        <td class="py-2.5 pr-3">
          <select id="pc-${p.id}" data-dirty-id="${p.id}" onchange="markDirty('prods',${p.id})"
            class="bg-gray-800 border ${dirty.prods.has(p.id)?'border-yellow-500':'border-gray-700'} focus:border-purple-500 rounded-lg px-2 py-1.5 text-white text-sm w-36 transition-colors">
            ${adminCats.map(c=>`<option value="${c.id}"${c.id===p.categoria_id?' selected':''}>${esc(c.nombre)}</option>`).join('')}
          </select>
        </td>
        <td class="py-2.5 pr-3">
          <input id="pp-${p.id}" data-dirty-id="${p.id}" type="number" value="${p.precio}" step="500" min="0"
            oninput="markDirty('prods',${p.id})"
            class="bg-gray-800 border ${dirty.prods.has(p.id)?'border-yellow-500 dirty-input':'border-gray-700'} focus:border-purple-500 rounded-lg px-2.5 py-1.5 text-white text-sm w-28 transition-colors">
        </td>
        <td class="py-2.5 pr-3 text-center">
          <button data-id="${p.id}" onclick="toggleProd(parseInt(this.dataset.id))"
            class="${p.disponible?'bg-green-700 hover:bg-green-600':'bg-red-800 hover:bg-red-700'} text-white text-xs px-3 py-1.5 rounded-lg font-bold transition-colors w-24">
            ${p.disponible?'Activo':'Agotado'}
          </button>
        </td>
        <td id="basetd-${p.id}" class="py-2.5 pr-3 text-center">
          <button onclick="openBasePop(${p.id})"
            class="flex flex-col items-center gap-0.5 w-full hover:bg-gray-800 rounded-lg px-2 py-1.5 transition-colors">
            ${baseLabel}
            <span class="text-gray-600 text-[9px]">▼ elegir</span>
          </button>
        </td>
        <td class="py-2.5 text-center">
          <button data-id="${p.id}" onclick="deleteProd(parseInt(this.dataset.id))"
            class="bg-gray-700 hover:bg-red-800 text-white text-xs px-2.5 py-1.5 rounded-lg transition-colors">x</button>
        </td>
      </tr>`;
    }).join('');
  } catch(e) { adminToast('Error cargando', false); }
}

async function toggleProd(id) {
  try { await fetch('/api/productos/'+id+'/toggle', {method:'PATCH'}); loadProds(); }
  catch(e) {}
}
async function deleteProd(id) {
  if (!confirm('Eliminar este producto?')) return;
  dirty.prods.delete(id); updateSaveBar();
  try { await fetch('/api/productos/'+id, {method:'DELETE'}); adminToast('Eliminado'); loadProds(); }
  catch(e) {}
}
async function addProd() {
  if (!adminCats.length) { adminToast('Crea una categoria primero', false); return; }
  try {
    const r = await fetch('/api/productos', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({nombre:'Nuevo Producto', categoria_id:adminCats[0].id, precio:0, disponible:true})
    });
    if (r.ok) { adminToast('Producto creado, edita y guarda'); loadProds(); }
  } catch(e) {}
}

// ─── CATEGORIAS TAB ───
async function loadCats() {
  try { const r = await fetch('/api/categorias'); renderCats(await r.json()); }
  catch(e) {}
}
function renderCats(cats) {
  document.getElementById('catsList').innerHTML = cats.length ? cats.map(c => `
    <div id="drow-${c.id}" class="flex items-center gap-2 bg-gray-800/50 border border-gray-700 rounded-xl px-4 py-2.5 ${dirty.cats.has(c.id)?'dirty-row border-yellow-700/40':''}">
      <input id="cn-${c.id}" data-dirty-id="${c.id}" type="text" value="${esc(c.nombre)}"
        oninput="markDirty('cats',${c.id})"
        class="flex-1 bg-transparent text-white font-semibold text-sm border-b ${dirty.cats.has(c.id)?'border-yellow-500':'border-transparent'} focus:border-purple-500 transition-colors">
      <button data-id="${c.id}" onclick="deleteCat(parseInt(this.dataset.id))"
        class="bg-gray-700 hover:bg-red-800 text-white text-xs px-2 py-1.5 rounded-lg transition-colors shrink-0">x</button>
    </div>`).join('') : '<div class="text-gray-600 text-sm text-center py-6">Sin categorias.</div>';
}
async function deleteCat(id) {
  if (!confirm('Eliminar esta categoria?')) return;
  dirty.cats.delete(id); updateSaveBar();
  try { await fetch('/api/categorias/'+id, {method:'DELETE'}); adminToast('Eliminada'); loadCats(); loadProds(); }
  catch(e) {}
}
async function addCat() {
  const inp = document.getElementById('newCatName'); const n = inp.value.trim(); if (!n) return;
  try {
    const r = await fetch('/api/categorias', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({nombre:n})});
    if (r.ok) { inp.value = ''; adminToast('Creada'); loadCats(); }
  } catch(e) {}
}

// ─── NOTAS TAB ───
async function loadNotes() {
  try {
    const [nr, cr] = await Promise.all([fetch('/api/notas_rapidas'), fetch('/api/categorias')]);
    const notes = await nr.json(); const cats = await cr.json();
    const sel = document.getElementById('newNoteCat');
    sel.innerHTML = '<option value="">Global</option>' + cats.map(c => `<option value="${c.id}">${esc(c.nombre)}</option>`).join('');
    const catMap = {}; cats.forEach(c => catMap[c.id] = c.nombre);
    document.getElementById('notesList').innerHTML = notes.length ? notes.map(n => `
      <div class="flex items-center gap-2 rounded-full pl-3 pr-2 py-1.5 border
        ${n.categoria_id ? 'bg-amber-900/30 border-amber-700/40' : 'bg-gray-800 border-gray-700'}">
        ${n.categoria_id ? '<span class="text-amber-400 text-[10px] font-bold shrink-0">'+esc(catMap[n.categoria_id]||'')+'</span>' : ''}
        <span class="text-white text-sm font-medium">${esc(n.texto)}</span>
        <button data-id="${n.id}" onclick="deleteNote(parseInt(this.dataset.id))"
          class="bg-gray-700 hover:bg-red-800 text-white text-xs w-6 h-6 rounded-full flex items-center justify-center transition-colors ml-0.5 shrink-0">x</button>
      </div>`).join('') : '<div class="text-gray-600 text-sm">Sin notas.</div>';
  } catch(e) {}
}
async function addNote() {
  const inp = document.getElementById('newNoteText'); const sel = document.getElementById('newNoteCat');
  const texto = inp.value.trim(); if (!texto) return;
  const cat_id = sel.value ? parseInt(sel.value) : null;
  try {
    const r = await fetch('/api/notas_rapidas', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({texto, categoria_id:cat_id})});
    if (r.ok) { inp.value = ''; adminToast('Nota agregada'); loadNotes(); }
  } catch(e) {}
}
async function deleteNote(id) {
  try { await fetch('/api/notas_rapidas/'+id, {method:'DELETE'}); loadNotes(); }
  catch(e) {}
}

// ─── BASES TAB ───
let allTipos = [];   // cache de tipos_base
let allBases2 = [];  // cache de bases con tipo_nombre
let selectedTipoId = null;

async function loadBases() {
  try {
    [allTipos, allBases2] = await Promise.all([
      fetch('/api/tipos-base').then(r=>r.json()),
      fetch('/api/bases').then(r=>r.json()),
    ]);
    renderTipoChips();
    // Seleccionar primer tipo si ninguno está seleccionado o el seleccionado ya no existe
    if (!selectedTipoId || !allTipos.find(t => t.id === selectedTipoId)) {
      selectedTipoId = allTipos[0]?.id || null;
    }
    renderBaseItems();
  } catch(e) { adminToast('Error cargando bases', false); }
}

function renderTipoChips() {
  const el = document.getElementById('tipoChips');
  const empty = document.getElementById('basesEmptyState');
  if (!allTipos.length) {
    el.innerHTML = '';
    empty?.classList.remove('hidden');
    document.getElementById('baseItemsSection')?.classList.add('hidden');
    return;
  }
  empty?.classList.add('hidden');
  el.innerHTML = allTipos.map(t => {
    const isActive = t.id === selectedTipoId;
    const itemCount = allBases2.filter(b => b.tipo_id === t.id).length;
    return `<button onclick="selectTipo(${t.id})"
      class="flex items-center gap-1.5 px-3 py-2 rounded-xl text-sm font-bold transition-colors border ${isActive
        ? 'bg-purple-700 border-purple-500 text-white'
        : 'bg-gray-800 border-gray-700 text-gray-300 hover:bg-gray-700 hover:border-gray-600'}">
      <span class="tipo-icono-${t.id}">${esc(t.icono)}</span>
      <span class="tipo-nombre-${t.id}">${esc(t.nombre)}</span>
      <span class="text-[10px] opacity-60">(${itemCount})</span>
      <span onclick="event.stopPropagation();deleteTipoBase(${t.id})"
        class="ml-1 text-gray-400 hover:text-red-400 leading-none" title="Eliminar tipo">×</span>
    </button>`;
  }).join('');
}

function selectTipo(id) {
  selectedTipoId = id;
  renderTipoChips();
  renderBaseItems();
}

function renderBaseItems() {
  const section = document.getElementById('baseItemsSection');
  if (!selectedTipoId || !allTipos.length) { section?.classList.add('hidden'); return; }
  section?.classList.remove('hidden');

  const tipo = allTipos.find(t => t.id === selectedTipoId);
  const label = document.getElementById('baseItemsTipoLabel');
  if (label && tipo) label.textContent = tipo.icono + ' ' + tipo.nombre;

  const items = allBases2.filter(b => b.tipo_id === selectedTipoId);
  const el = document.getElementById('basesList');
  if (!items.length) {
    el.innerHTML = '<div class="text-gray-700 text-xs text-center py-6">Sin items. Agrega uno arriba.</div>';
    return;
  }
  el.innerHTML = `<div class="grid grid-cols-3 gap-2 mb-2">${items.map(b => `
    <div class="bg-gray-800/60 border ${b.disponible?'border-gray-700':'border-red-900/50'} rounded-2xl p-3 flex flex-col gap-2 ${b.disponible?'':'opacity-70'}">
      <input id="bn-${b.id}" type="text" value="${esc(b.nombre)}"
        onblur="saveBaseItem(${b.id})"
        onkeydown="if(event.key==='Enter')this.blur()"
        class="bg-transparent text-white font-bold text-sm text-center border-b border-transparent hover:border-gray-600 focus:border-purple-500 outline-none transition-colors w-full">
      <button data-id="${b.id}" onclick="toggleBaseItem(parseInt(this.dataset.id))"
        class="w-full font-extrabold text-xs py-2 rounded-xl transition-colors ${b.disponible
          ? 'bg-red-900/60 hover:bg-red-700 text-red-300 hover:text-white'
          : 'bg-green-900/60 hover:bg-green-700 text-green-300 hover:text-white'}">
        ${b.disponible ? '⚠ AGOTADO' : '✓ DISPONIBLE'}
      </button>
      <button data-id="${b.id}" onclick="deleteBaseItem(parseInt(this.dataset.id))"
        class="text-gray-600 hover:text-red-400 text-[10px] transition-colors text-center">Eliminar</button>
    </div>`).join('')}</div>`;
}

function toggleNewTipoForm() {
  const form = document.getElementById('newTipoForm');
  form?.classList.toggle('hidden');
  if (!form?.classList.contains('hidden')) document.getElementById('newTipoNombre')?.focus();
}

async function addTipoBase() {
  const nombre = document.getElementById('newTipoNombre')?.value.trim();
  const icono  = document.getElementById('newTipoIcono')?.value.trim() || '🍹';
  if (!nombre) { adminToast('Escribe un nombre para el tipo', false); return; }
  try {
    const r = await fetch('/api/tipos-base', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({nombre, icono})
    });
    if (r.ok) {
      document.getElementById('newTipoNombre').value = '';
      document.getElementById('newTipoIcono').value = '🍹';
      document.getElementById('newTipoForm')?.classList.add('hidden');
      adminToast('Tipo creado');
      await loadBases();
      const nuevo = allTipos.find(t => t.nombre === nombre);
      if (nuevo) selectTipo(nuevo.id);
    } else {
      const err = await r.json().catch(()=>({detail:'Error'}));
      adminToast(err.detail || 'Error al crear', false);
    }
  } catch(e) {}
}

async function deleteTipoBase(id) {
  if (!confirm('Eliminar este tipo? Solo se puede si no tiene items.')) return;
  try {
    const r = await fetch('/api/tipos-base/'+id, {method:'DELETE'});
    if (r.ok) {
      if (selectedTipoId === id) selectedTipoId = null;
      adminToast('Tipo eliminado'); loadBases();
    } else {
      const err = await r.json().catch(()=>({detail:'Error'}));
      adminToast(err.detail || 'No se pudo eliminar', false);
    }
  } catch(e) {}
}

async function addBaseItem() {
  const nombre = document.getElementById('newBaseNombre')?.value.trim();
  if (!nombre) { adminToast('Escribe un nombre para el item', false); return; }
  if (!selectedTipoId) { adminToast('Selecciona un tipo primero', false); return; }
  try {
    const r = await fetch('/api/bases', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({nombre, tipo_id: selectedTipoId})
    });
    if (r.ok) {
      document.getElementById('newBaseNombre').value = '';
      adminToast('Item agregado'); loadBases();
    } else {
      const err = await r.json().catch(()=>({detail:'Error'}));
      adminToast(err.detail || 'Error al agregar', false);
    }
  } catch(e) {}
}

async function saveBaseItem(id) {
  const inp = document.getElementById('bn-'+id);
  const nombre = inp?.value.trim();
  if (!nombre) { inp && (inp.value = inp.defaultValue); return; }
  const base = allBases2.find(b => b.id === id);
  if (!base) return;
  try {
    const r = await fetch('/api/bases/'+id, {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({nombre, tipo_id: base.tipo_id})
    });
    if (r.ok) { adminToast('Guardado'); loadBases(); }
    else adminToast('Error', false);
  } catch(e) {}
}

async function toggleBaseItem(id) {
  try { await fetch('/api/bases/'+id+'/toggle', {method:'PATCH'}); loadBases(); }
  catch(e) {}
}

async function deleteBaseItem(id) {
  if (!confirm('Eliminar este item? Se quitará de todos los productos.')) return;
  try { await fetch('/api/bases/'+id, {method:'DELETE'}); adminToast('Eliminado'); loadBases(); }
  catch(e) {}
}

// ─── DASHBOARD ───  (ahora en ruta /dashboard)


// ─── SETTINGS (DB como fuente de verdad) ───
function saveSetting(key, value) {
  localStorage.setItem(key, value);
  fetch('/api/settings/'+key, {
    method:'PUT', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({value})
  }).catch(()=>{});
}
async function loadSettings() {
  try {
    const s = await fetch('/api/settings').then(r=>r.json());
    if (s.palette || s.drunksPalette) applyTheme(s.palette || s.drunksPalette);
    if (s.kitchen_view) { viewMode = s.kitchen_view; localStorage.setItem('kitchenView', s.kitchen_view); }
    if (s.station_cfg)  { try { stationCfg   = JSON.parse(s.station_cfg); } catch(e){} }
    if (s.station_active){ try { stationActive= JSON.parse(s.station_active); } catch(e){} }
    Object.entries(s).forEach(([k,v]) => localStorage.setItem(k, v));
  } catch(e) {
    applyTheme(localStorage.getItem('drunksPalette') || 'dark');
  }
}

// ─── THEME SYSTEM ───
function applyTheme(t) {
  const theme = t||'dark';
  if (theme.startsWith('custom:')) { applyCustomTheme(theme.split(':')[1]); return; }
  document.documentElement.setAttribute('data-theme', theme);
  saveSetting('drunksPalette', theme);
  saveSetting('palette', theme);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b => {
    const on = b.dataset.theme === theme;
    b.style.background = on ? 'rgba(147,51,234,.2)' : '';
    b.style.outline    = on ? '1px solid rgba(147,51,234,.5)' : '';
  });
  const picker = document.getElementById('customColorPicker');
  if (picker) picker.style.display = 'none';
}
function applyCustomTheme(hex) {
  let style = document.getElementById('custom-theme-style');
  if (!style) { style = document.createElement('style'); style.id='custom-theme-style'; document.head.appendChild(style); }
  const h = hex||'#9333ea';
  const ri = parseInt(h.slice(1,3),16), gi = parseInt(h.slice(3,5),16), bi = parseInt(h.slice(5,7),16);
  style.textContent = `
    html[data-theme="custom"] body{background:#06060f}
    html[data-theme="custom"] .bg-purple-600,.bg-purple-700{background-color:${h}!important}
    html[data-theme="custom"] .text-purple-400,.text-purple-300{color:${h}!important}
    html[data-theme="custom"] .border-purple-500{border-color:${h}!important}
    html[data-theme="custom"] .from-purple-600{--tw-gradient-from:${h}!important}`;
  document.documentElement.setAttribute('data-theme','custom');
  saveSetting('drunksPalette','custom:'+h); saveSetting('palette','custom:'+h);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b=>{b.style.background='';b.style.outline='';});
  const picker = document.getElementById('customColorPicker');
  if (picker) { picker.style.display='inline-block'; picker.value=h; }
}
function toggleThemePicker() {
  document.getElementById('themePicker').classList.toggle('hidden');
}
document.addEventListener('click', e => {
  if (!e.target.closest('#themePicker') && !e.target.closest('[onclick*="toggleThemePicker"]'))
    document.getElementById('themePicker')?.classList.add('hidden');
});

// ─── ESTADO SUPABASE ───
async function checkSupabase() {
  const dot = document.getElementById('sbDot');
  const txt = document.getElementById('sbTxt');
  try {
    const r = await fetch('/api/sync/status');
    const d = await r.json();
    if (!d.configured) {
      dot.className = 'w-2 h-2 rounded-full bg-gray-600 shrink-0';
      txt.textContent = 'Nube no configurada';
      txt.style.color = '#4b5563';
    } else if (!d.reachable) {
      dot.className = 'w-2 h-2 rounded-full bg-red-500 shrink-0';
      txt.textContent = d.pending > 0 ? `Sin nube · ${d.pending} sin subir` : 'Sin conexión a la nube';
      txt.style.color = '#f87171';
    } else if (d.pending > 0) {
      dot.className = 'w-2 h-2 rounded-full bg-yellow-400 shrink-0';
      txt.textContent = `Subiendo ${d.pending} pedido${d.pending !== 1 ? 's' : ''}...`;
      txt.style.color = '#fbbf24';
    } else {
      dot.className = 'w-2 h-2 rounded-full bg-green-500 shrink-0';
      txt.textContent = 'Nube sincronizada';
      txt.style.color = '#4ade80';
    }
  } catch(_) {}
}

// ─── STATION SYSTEM ───
let stationCfg = JSON.parse(localStorage.getItem('stationCfg') || 'null') || {
  count: 1,
  stations: [{ name: 'Preparador 1', catIds: [] }]
};
let stationActive = JSON.parse(localStorage.getItem('stationActive') || '{}');
// { "0": orderId, "1": orderId } — pedido activo por estación
let stationCats  = [];   // cached from /api/categorias

function toggleStationPanel() {
  const p = document.getElementById('stationPanel');
  p.classList.toggle('hidden');
  if (!p.classList.contains('hidden')) {
    // Load cats and render form
    if (!stationCats.length) {
      fetch('/api/categorias').then(r=>r.json()).then(c=>{
        stationCats = c; renderStationForm();
      }).catch(()=>{ renderStationForm(); });
    } else {
      renderStationForm();
    }
    // Highlight active count button
    for (let i=1;i<=4;i++) {
      const btn = document.getElementById('sCnt-'+i);
      if (btn) btn.className = 'station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border transition-colors '
        + (i===stationCfg.count
          ? 'border-purple-500 bg-purple-700 text-white'
          : 'border-gray-700 bg-gray-800 text-gray-400');
    }
  }
}

function setStationCount(n) {
  // Adjust stations array
  while (stationCfg.stations.length < n)
    stationCfg.stations.push({ name: 'Preparador '+(stationCfg.stations.length+1), catIds: [] });
  stationCfg.stations = stationCfg.stations.slice(0, n);
  stationCfg.count = n;
  for (let i=1;i<=4;i++) {
    const btn = document.getElementById('sCnt-'+i);
    if (btn) btn.className = 'station-cnt-btn flex-1 py-2 rounded-xl text-sm font-bold border transition-colors '
      + (i===n
        ? 'border-purple-500 bg-purple-700 text-white'
        : 'border-gray-700 bg-gray-800 text-gray-400');
  }
  renderStationForm();
}

function renderStationForm() {
  const block = document.getElementById('stationFormBlock');
  block.innerHTML = stationCfg.stations.map((s, si) => {
    const catChecks = stationCats.map(c =>
      `<label class="flex items-center gap-1.5 text-xs text-gray-300 cursor-pointer">
        <input type="checkbox" data-si="${si}" data-cat="${c.id}"
          ${s.catIds.includes(c.id)?'checked':''} class="scat-chk rounded accent-purple-500">
        <span>${c.nombre}</span>
      </label>`).join('');
    return `<div class="bg-gray-800/60 border border-gray-700 rounded-xl p-3">
      <input data-si="${si}" type="text" value="${s.name.replace(/"/g,'&quot;')}"
        placeholder="Nombre de la estacion"
        class="sname-inp w-full bg-transparent text-white text-sm font-semibold border-b border-gray-700 focus:border-purple-500 outline-none mb-2 pb-1">
      ${stationCats.length
        ? '<div class="text-gray-500 text-[10px] tracking-widest font-bold mb-1.5">CATEGORIAS</div><div class="flex flex-wrap gap-2">'+catChecks+'</div>'
        : '<div class="text-gray-600 text-xs">Sin categorias (atiende todos los pedidos)</div>'}
    </div>`;
  }).join('');
}

function saveStationConfig() {
  // Read names and categories from form
  document.querySelectorAll('.sname-inp').forEach(inp => {
    const si = parseInt(inp.dataset.si);
    if (stationCfg.stations[si]) stationCfg.stations[si].name = inp.value.trim() || ('Preparador '+(si+1));
  });
  document.querySelectorAll('.scat-chk').forEach(cb => {
    const si = parseInt(cb.dataset.si), catId = parseInt(cb.dataset.cat);
    if (!stationCfg.stations[si]) return;
    const ids = stationCfg.stations[si].catIds;
    if (cb.checked && !ids.includes(catId)) ids.push(catId);
    else if (!cb.checked) stationCfg.stations[si].catIds = ids.filter(id=>id!==catId);
  });
  saveSetting('stationCfg', JSON.stringify(stationCfg));
  saveSetting('station_cfg', JSON.stringify(stationCfg));
  toggleStationPanel();
  applyStationLayout();
}

function applyStationLayout() {
  const multi = stationCfg.count >= 2;
  const viewBar = document.getElementById('viewModeBar');
  if (viewBar) viewBar.style.display = multi ? 'none' : '';
  document.getElementById('singleView').classList.toggle('hidden', multi);
  document.getElementById('multiView').classList.toggle('hidden', !multi);
  renderKitchen();
}

// ─── COLA COMPARTIDA — asignación automática FIFO ────────────────────────────

// Asigna pedidos de la cola a estaciones libres. Se llama en cada render.
function autoAssignFreeStations(allOrders) {
  // Limpiar asignaciones de pedidos que ya no existen
  Object.keys(stationActive).forEach(si => {
    if (!orders[Number(stationActive[si])]) delete stationActive[si];
  });

  const activeIds = new Set(Object.values(stationActive).map(Number));
  // Cola: pedidos no activos en ninguna estación, orden FIFO
  const queue = allOrders.filter(o => !activeIds.has(o.id));

  for (let si = 0; si < stationCfg.count; si++) {
    if (stationActive[String(si)] === undefined && queue.length) {
      const next = queue.shift();
      stationActive[String(si)] = next.id;
      activeIds.add(next.id);
    }
  }
  saveSetting('stationActive', JSON.stringify(stationActive));
  saveSetting('station_active', JSON.stringify(stationActive));
}

// Save station name edited inline in the column header
function saveStationName(si, value) {
  if (!stationCfg.stations[si]) return;
  stationCfg.stations[si].name = value.trim() || ('Preparador ' + (si + 1));
  saveSetting('stationCfg', JSON.stringify(stationCfg));
  saveSetting('station_cfg', JSON.stringify(stationCfg));
  // Also sync the config panel input if open
  const inp = document.querySelector('.sname-inp[data-si="'+si+'"]');
  if (inp) inp.value = stationCfg.stations[si].name;
}

// ─── STATION RENDERING ────────────────────────────────────────────────────────

// Hero card — el pedido activo que está preparando ahora el preparador
function buildActiveCard(order, stationItems, si) {
  const div = document.createElement('div');
  div.id = 'active-order-'+order.id+'-'+si;
  const efect = order.metodo_pago === 'Efectivo';
  const hBg   = efect ? 'from-green-700 to-green-800' : 'from-blue-700 to-blue-800';
  const icon  = efect ? '💵' : '📱';
  const tStr  = new Date(order.fecha||Date.now()).toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  const rows  = stationItems.map(it => buildItemRow(it, true)).join('');
  div.className = 'active-card-in bg-gray-900 border border-purple-500/50 rounded-2xl overflow-hidden flex-shrink-0 shadow-lg shadow-purple-900/20';
  div.innerHTML = `
    <div class="flex items-center gap-2 px-4 pt-3 pb-1">
      <div class="priority-badge bg-purple-600 text-white text-[10px] font-extrabold px-2.5 py-0.5 rounded-full tracking-wider">PREPARANDO</div>
      <div class="h-px flex-1 bg-gray-800"></div>
    </div>
    <div class="bg-gradient-to-r ${hBg} px-4 py-3 flex items-start justify-between">
      <div class="min-w-0">
        <div class="flex items-center gap-2 mb-1">
          <span class="bg-white/20 text-white text-[10px] font-extrabold px-2 py-0.5 rounded-full">${order.numero_factura||('#'+order.id)}</span>
          <span class="text-white/60 text-xs">${icon} ${order.metodo_pago} · ${tStr}</span>
        </div>
        <div class="text-white font-extrabold text-2xl leading-tight">${order.cliente}</div>
      </div>
      <div class="text-white/80 font-bold text-base tabular-nums ml-2 shrink-0">$${Math.round(order.total).toLocaleString('es-CO')}</div>
    </div>
    <div class="px-4 py-3 space-y-1">${rows||'<div class="text-gray-600 text-xs text-center py-2">Sin items</div>'}</div>
    <div class="px-4 pb-4 pt-1">
      <button data-oid="${order.id}" data-si="${si}"
        onclick="completeStationOrder(parseInt(this.dataset.oid),parseInt(this.dataset.si))"
        class="w-full bg-green-500 hover:bg-green-400 active:bg-green-600 text-white font-extrabold py-3.5 rounded-xl text-base transition-colors shadow-lg shadow-green-900/30 tracking-wide">
        ✓ LISTO — COMPLETAR PEDIDO
      </button>
    </div>`;
  return div;
}

// Mini card — pedidos esperando en la cola de este preparador
function buildQueueMiniCard(order, position) {
  const div = document.createElement('div');
  div.className = 'queue-mini-card bg-gray-800/70 border border-gray-700 rounded-xl overflow-hidden';
  const itemCount = (order.items||[]).length;
  const efect = order.metodo_pago === 'Efectivo';
  const dot = efect ? 'bg-green-500' : 'bg-blue-500';
  div.innerHTML = `
    <div class="px-3 pt-2.5 pb-1 flex items-center gap-1.5">
      <span class="text-gray-500 text-[10px] font-bold">#${position+1}</span>
      <span class="text-purple-400 text-[10px] font-extrabold">${order.numero_factura||('#'+order.id)}</span>
    </div>
    <div class="px-3 pb-1">
      <div class="text-white text-xs font-bold truncate">${order.cliente}</div>
    </div>
    <div class="px-3 pb-2.5 flex items-center gap-1.5">
      <span class="w-1.5 h-1.5 rounded-full ${dot} shrink-0"></span>
      <span class="text-gray-400 text-[10px]">${itemCount} ${itemCount===1?'bebida':'bebidas'}</span>
    </div>`;
  return div;
}

// Panel de un preparador: header editable + tarjeta activa o placeholder
function renderActivePanel(si, activeOrder) {
  const cfg = stationCfg.stations[si];
  const panel = document.createElement('div');
  panel.className = 'flex flex-col flex-1 border-r border-gray-800/60 overflow-y-auto min-w-0';

  // ── Header editable ────────────────────────────
  const hdr = document.createElement('div');
  hdr.className = 'sticky top-0 z-10 bg-gray-950/95 backdrop-blur border-b border-gray-800 px-4 py-3 flex items-center gap-2 shrink-0';
  hdr.innerHTML = `
    <input class="station-col-name bg-transparent text-white font-extrabold text-sm tracking-wide uppercase border-b-2 border-transparent hover:border-gray-600 focus:border-purple-500 focus:outline-none min-w-0 flex-1 transition-colors cursor-text"
      data-si="${si}" value="${cfg.name}"
      onblur="saveStationName(${si},this.value)"
      onkeydown="if(event.key==='Enter')this.blur()"
      title="Clic para editar nombre">`;
  panel.appendChild(hdr);

  // ── Pedido activo o placeholder ────────────────
  if (activeOrder) {
    const catchAll = !cfg.catIds || cfg.catIds.length === 0;
    const items = (activeOrder.items||[]).filter(it =>
      catchAll || cfg.catIds.includes(it.categoria_id));
    const wrap = document.createElement('div');
    wrap.className = 'p-3';
    wrap.appendChild(buildActiveCard(activeOrder, items, si));
    panel.appendChild(wrap);
  } else {
    const empty = document.createElement('div');
    empty.className = 'flex flex-col items-center justify-center flex-1 text-gray-700 py-16';
    empty.innerHTML = '<div class="text-4xl mb-3 opacity-20">🍹</div><p class="text-xs">Libre</p>';
    panel.appendChild(empty);
  }

  return panel;
}

// Franja inferior: cola compartida de todos los pedidos en espera
function renderSharedQueueStrip(queueOrders) {
  const strip = document.createElement('div');
  strip.className = 'border-t border-gray-800 bg-gray-950/80 px-4 py-3 shrink-0';
  if (!queueOrders.length) {
    strip.innerHTML = '<p class="text-gray-700 text-xs text-center py-0.5">Cola vacía — todos los pedidos están siendo preparados</p>';
    return strip;
  }
  const label = document.createElement('div');
  label.className = 'text-gray-500 text-[10px] font-extrabold uppercase tracking-widest mb-2';
  label.textContent = `En cola (${queueOrders.length})`;
  strip.appendChild(label);
  const row = document.createElement('div');
  row.className = 'flex gap-2 overflow-x-auto pb-1';
  row.style.scrollbarWidth = 'none';
  queueOrders.forEach((o, i) => row.appendChild(buildQueueMiniCard(o, i)));
  strip.appendChild(row);
  return strip;
}

function renderMultiStation() {
  const list = Object.values(orders).sort((a,b) => a.id - b.id);
  const container = document.getElementById('multiView');
  container.innerHTML = '';
  const empty = document.getElementById('emptyState');
  if (!list.length) { empty.classList.remove('hidden'); return; }
  empty.classList.add('hidden');

  autoAssignFreeStations(list);

  const activeIds = new Set(Object.values(stationActive).map(Number));
  const queueOrders = list.filter(o => !activeIds.has(o.id));

  // ── Fila superior: paneles de cada preparador ──
  const topRow = document.createElement('div');
  topRow.className = 'flex flex-1 min-h-0 overflow-hidden';
  for (let si = 0; si < stationCfg.count; si++) {
    const oid = stationActive[String(si)];
    const activeOrder = (oid !== undefined && orders[oid]) ? orders[oid] : null;
    topRow.appendChild(renderActivePanel(si, activeOrder));
  }
  container.appendChild(topRow);

  // ── Franja inferior: cola compartida ───────────
  container.appendChild(renderSharedQueueStrip(queueOrders));
}

// Un clic = pedido completado, estación libre, siguiente de la cola sube solo
async function completeStationOrder(orderId, si) {
  const order = orders[orderId];
  if (!order) return;

  // Liberar la estación
  delete stationActive[String(si)];
  saveSetting('stationActive', JSON.stringify(stationActive));
  saveSetting('station_active', JSON.stringify(stationActive));

  // Animar salida
  const card = document.getElementById('active-order-'+orderId+'-'+si);
  if (card) card.classList.add('fade-out');

  try {
    const r = await fetch('/api/pedidos/'+orderId+'/entregar', {method:'PUT'});
    if (r.ok) {
      setTimeout(() => { delete orders[orderId]; renderKitchen(); }, 350);
    }
  } catch(e) { renderKitchen(); }
}

// ─── INIT ───
loadSettings().then(() => {
  applyStationLayout();
  loadPending();
  connectWS();
  checkSupabase();
  setInterval(checkSupabase, 30000);
});

// ── Auto-update ──
async function checkUpdate() {
  try {
    const d = await fetch('/api/update/check').then(r => r.json());
    if (d.has_update && d.latest) {
      document.getElementById('updateVersion').textContent = 'v'+d.current+' → v'+d.latest;
      document.getElementById('updateBanner').classList.remove('hidden');
    }
  } catch(e) {}
}
async function applyUpdate() {
  const btn = document.getElementById('updateBtn');
  btn.textContent = 'Descargando...'; btn.disabled = true;
  try {
    await fetch('/api/update/apply', {method:'POST'});
    btn.textContent = 'Reiniciando...';
  } catch(e) { btn.textContent = 'Error — intenta de nuevo'; btn.disabled = false; }
}
checkUpdate();
setInterval(checkUpdate, 30 * 60 * 1000);
</script>
<!-- ═══ BANNER DE ACTUALIZACIÓN ═══ -->
<div id="updateBanner" class="hidden fixed bottom-0 left-0 right-0 z-[100] bg-purple-950/97 border-t border-purple-500/50 backdrop-blur px-5 py-3 flex items-center justify-between gap-3 shadow-2xl">
  <div class="flex items-center gap-3">
    <span class="text-2xl">🔄</span>
    <div>
      <div class="text-white font-bold text-sm">Nueva versión disponible</div>
      <div id="updateVersion" class="text-purple-300 text-xs font-mono"></div>
    </div>
  </div>
  <div class="flex gap-2 shrink-0">
    <button id="updateBtn" onclick="applyUpdate()"
      class="bg-purple-600 hover:bg-purple-500 active:bg-purple-700 text-white text-sm font-extrabold px-4 py-2 rounded-xl transition-colors shadow-lg">
      Actualizar ahora
    </button>
    <button onclick="document.getElementById('updateBanner').classList.add('hidden')"
      class="text-purple-400 hover:text-white text-sm px-3 py-2 rounded-xl transition-colors">
      Después
    </button>
  </div>
</div>
</body>
</html>"""

# ─────────────────────────────────────────────
# BASE DE DATOS
# ─────────────────────────────────────────────
def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS categorias (
                id     INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS productos (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                categoria_id INTEGER REFERENCES categorias(id) ON DELETE SET NULL,
                nombre       TEXT    NOT NULL,
                precio       REAL    NOT NULL DEFAULT 0.0,
                disponible   INTEGER NOT NULL DEFAULT 1,
                tiene_base   INTEGER NOT NULL DEFAULT 0
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS notas_rapidas (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                texto        TEXT    NOT NULL,
                categoria_id INTEGER REFERENCES categorias(id) ON DELETE SET NULL
            )""")
        # Migraciones de columnas (seguras si ya existen)
        for migration in [
            "ALTER TABLE notas_rapidas ADD COLUMN categoria_id INTEGER REFERENCES categorias(id) ON DELETE SET NULL",
            "ALTER TABLE productos ADD COLUMN tiene_base INTEGER NOT NULL DEFAULT 0",
        ]:
            try:
                conn.execute(migration)
                conn.commit()
            except Exception:
                pass
        # Migración: numero_factura en pedidos
        try:
            conn.execute("ALTER TABLE pedidos ADD COLUMN numero_factura TEXT NOT NULL DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        # Migración: sync_id (UUID único por pedido para identificación en Supabase)
        try:
            conn.execute("ALTER TABLE pedidos ADD COLUMN sync_id TEXT NOT NULL DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        try:
            rows = conn.execute("SELECT id FROM pedidos WHERE sync_id='' OR sync_id IS NULL").fetchall()
            for row in rows:
                conn.execute("UPDATE pedidos SET sync_id=? WHERE id=?", (str(uuid.uuid4()), row["id"]))
            if rows:
                conn.commit()
        except Exception:
            pass
        try:
            rows = conn.execute("SELECT id FROM pedidos WHERE numero_factura='' OR numero_factura IS NULL").fetchall()
            for row in rows:
                conn.execute("UPDATE pedidos SET numero_factura=? WHERE id=?",
                             (f"DRK-{row['id']:05d}", row["id"]))
            if rows:
                conn.commit()
        except Exception:
            pass

        # Post-migración: si existe categoria Micheladas, marcar sus productos con tiene_base=1
        try:
            mich_row = conn.execute("SELECT id FROM categorias WHERE nombre='Micheladas'").fetchone()
            if mich_row:
                conn.execute("UPDATE productos SET tiene_base=1 WHERE categoria_id=? AND tiene_base=0", (mich_row["id"],))
                conn.commit()
        except Exception:
            pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pedidos (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente        TEXT    NOT NULL,
                metodo_pago    TEXT    NOT NULL,
                total          REAL    NOT NULL,
                estado         TEXT    NOT NULL DEFAULT 'pendiente',
                fecha          TEXT    NOT NULL,
                sincronizado   INTEGER NOT NULL DEFAULT 0,
                numero_factura TEXT    NOT NULL DEFAULT '',
                sync_id        TEXT    NOT NULL DEFAULT ''
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS detalle_pedidos (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                pedido_id     INTEGER NOT NULL REFERENCES pedidos(id),
                producto_id   INTEGER NOT NULL REFERENCES productos(id),
                cantidad      INTEGER NOT NULL,
                observaciones TEXT    DEFAULT ''
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tipos_base (
                id     INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT    NOT NULL UNIQUE,
                icono  TEXT    NOT NULL DEFAULT '🍹'
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bases (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre     TEXT    NOT NULL,
                tipo       TEXT    NOT NULL DEFAULT 'Gaseosa',
                tipo_id    INTEGER REFERENCES tipos_base(id) ON DELETE SET NULL,
                disponible INTEGER NOT NULL DEFAULT 1
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS producto_bases (
                producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE CASCADE,
                base_id     INTEGER NOT NULL REFERENCES bases(id)     ON DELETE CASCADE,
                PRIMARY KEY (producto_id, base_id)
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )""")
        conn.commit()

        # Seed tipos_base (incluyendo Soda como tercer tipo predeterminado)
        if conn.execute("SELECT COUNT(*) FROM tipos_base").fetchone()[0] == 0:
            conn.executemany("INSERT INTO tipos_base (nombre, icono) VALUES (?,?)", [
                ("Cerveza", "🍺"), ("Gaseosa", "🧃"), ("Soda", "🫧"),
            ])
            conn.commit()
        else:
            # Migración: asegurar que "Soda" existe como tipo
            if not conn.execute("SELECT id FROM tipos_base WHERE nombre='Soda'").fetchone():
                conn.execute("INSERT INTO tipos_base (nombre, icono) VALUES ('Soda','🫧')")
                conn.commit()

        # Seed bases predeterminadas
        if conn.execute("SELECT COUNT(*) FROM bases").fetchone()[0] == 0:
            cer_id = conn.execute("SELECT id FROM tipos_base WHERE nombre='Cerveza'").fetchone()["id"]
            gas_id = conn.execute("SELECT id FROM tipos_base WHERE nombre='Gaseosa'").fetchone()["id"]
            sod_id = conn.execute("SELECT id FROM tipos_base WHERE nombre='Soda'").fetchone()["id"]
            conn.executemany("INSERT INTO bases (nombre, tipo, tipo_id) VALUES (?,?,?)", [
                ("Aguila",  "Cerveza", cer_id), ("Corona",  "Cerveza", cer_id),
                ("Costena", "Cerveza", cer_id),
                ("Quatro",  "Gaseosa", gas_id), ("Sprite",  "Gaseosa", gas_id),
                ("Bretana", "Soda",    sod_id),  ("Ginger",  "Soda",    sod_id),
            ])
            conn.commit()
        else:
            # Migración: reclasificar Bretana/Ginger a Soda si aún son Gaseosa
            try:
                sod_id = conn.execute("SELECT id FROM tipos_base WHERE nombre='Soda'").fetchone()["id"]
                conn.execute(
                    "UPDATE bases SET tipo='Soda', tipo_id=? WHERE LOWER(nombre) IN ('bretana','ginger') AND (tipo='Gaseosa' OR tipo_id IS NULL OR tipo_id!=(?))",
                    (sod_id, sod_id)
                )
                conn.commit()
            except Exception:
                pass

        # Migración: añadir columna tipo_id si no existe (instalaciones antiguas)
        # DEBE ir ANTES de poblarla
        try:
            conn.execute("ALTER TABLE bases ADD COLUMN tipo_id INTEGER REFERENCES tipos_base(id) ON DELETE SET NULL")
            conn.commit()
        except Exception:
            pass  # Ya existe — ignorar

        # Migración: poblar tipo_id en bases existentes que aún no lo tengan
        try:
            rows = conn.execute("SELECT id, tipo FROM bases WHERE tipo_id IS NULL").fetchall()
            if rows:
                tipos = {r["nombre"]: r["id"] for r in conn.execute("SELECT id, nombre FROM tipos_base").fetchall()}
                for row in rows:
                    tid = tipos.get(row["tipo"])
                    if tid:
                        conn.execute("UPDATE bases SET tipo_id=? WHERE id=?", (tid, row["id"]))
                conn.commit()
        except Exception:
            pass

        # ── Seeding ──
        if conn.execute("SELECT COUNT(*) FROM categorias").fetchone()[0] == 0:
            conn.executemany("INSERT INTO categorias (nombre) VALUES (?)", [
                ("Micheladas",), ("Mojitos",), ("Especiales",)
            ])
            conn.commit()

        if conn.execute("SELECT COUNT(*) FROM productos").fetchone()[0] == 0:
            ids = [r["id"] for r in conn.execute("SELECT id FROM categorias ORDER BY id").fetchall()]
            c_mich, c_moji, c_esp = ids[0], ids[1], ids[2]
            conn.executemany(
                "INSERT INTO productos (categoria_id, nombre, precio, disponible, tiene_base) VALUES (?,?,?,1,?)",
                [
                    (c_mich, "Michelada Maracuya",      14000, 1),
                    (c_mich, "Michelada Tamarindo",     14000, 1),
                    (c_mich, "Michelada Frutos Rojos",  14000, 1),
                    (c_mich, "Michelada Clasica",       10000, 1),
                    (c_moji, "Mojito Maracuya",         16000, 0),
                    (c_moji, "Mojito Tamarindo",        16000, 0),
                    (c_moji, "Mojito Frutos Rojos",     16000, 0),
                    (c_esp,  "Pati Chamoy",             12000, 0),
                    (c_esp,  "Chamoy",                   5000, 0),
                ]
            )
            conn.commit()

        if conn.execute("SELECT COUNT(*) FROM notas_rapidas").fetchone()[0] == 0:
            ids = {r["nombre"]: r["id"] for r in conn.execute("SELECT id, nombre FROM categorias").fetchall()}
            mich = ids.get("Micheladas")
            moji = ids.get("Mojitos")
            conn.executemany("INSERT INTO notas_rapidas (texto, categoria_id) VALUES (?,?)", [
                # Globales
                ("Sin Hielo",      None),
                ("Para Llevar",    None),
                # Específicas de Micheladas
                ("Poco Chamoy",    mich),
                ("Poca Sal",       mich),
                ("Extra Limon",    mich),
                ("Sin Chamoy",     mich),
                ("Extra Chamoy",   mich),
                ("Poca Crema",     mich),
                ("Sin Sal",        mich),
                ("Extra Crema",    mich),
                # Específicas de Mojitos
                ("Extra Limon",    moji),
                ("Poco Azucar",    moji),
                ("Con Gas",        moji),
                ("Sin Hierbabuena", moji),
            ])
            conn.commit()

# ─────────────────────────────────────────────
# WEBSOCKET MANAGER
# ─────────────────────────────────────────────
class ConnectionManager:
    def __init__(self):
        self._clients: list[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self._clients.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self._clients:
            self._clients.remove(ws)

    async def broadcast(self, data: dict):
        dead = []
        for ws in self._clients:
            try:
                await ws.send_json(data)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self._clients.remove(ws)

manager = ConnectionManager()

# ─────────────────────────────────────────────
# MODELOS
# ─────────────────────────────────────────────
class CategoriaBody(BaseModel):
    nombre: str

class ProductoCreate(BaseModel):
    categoria_id: Optional[int] = None
    nombre:       str
    precio:       float = 0.0
    disponible:   bool  = True
    tiene_base:   bool  = False

class ProductoUpdate(BaseModel):
    categoria_id: Optional[int]   = None
    nombre:       Optional[str]   = None
    precio:       Optional[float] = None
    disponible:   Optional[bool]  = None
    tiene_base:   Optional[bool]  = None

class NotaBody(BaseModel):
    texto:        str
    categoria_id: Optional[int] = None

class OrderItem(BaseModel):
    producto_id:   int
    cantidad:      int
    observaciones: str = ""

class TipoBaseCreate(BaseModel):
    nombre: str
    icono:  str = "🍹"

class TipoBaseUpdate(BaseModel):
    nombre: str
    icono:  str = "🍹"

class BaseCreate(BaseModel):
    nombre:  str
    tipo_id: int

class BaseUpdate(BaseModel):
    nombre:  str
    tipo_id: int

class ProductoBasesUpdate(BaseModel):
    base_ids: List[int]

class PedidoCreate(BaseModel):
    cliente:     str
    metodo_pago: str
    total:       float
    items:       List[OrderItem]

# ─────────────────────────────────────────────
# SUPABASE SYNC
# ─────────────────────────────────────────────
async def sync_to_supabase(pedido_id: int, payload: dict):
    if not SUPABASE_URL or not SUPABASE_KEY or not HTTPX_AVAILABLE:
        return
    try:
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }
        # Filtrar solo los campos que Supabase conoce
        campos = {"sync_id", "numero_factura", "cliente", "metodo_pago",
                  "total", "estado", "fecha", "items_json"}
        data = {k: v for k, v in payload.items() if k in campos}
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.post(
                f"{SUPABASE_URL}/rest/v1/pedidos", headers=headers, json=data
            )
            if r.status_code not in (200, 201):
                # Fallback: intentar sin sync_id/items_json (columnas opcionales)
                fallback = {k: v for k, v in data.items()
                            if k not in ("sync_id", "items_json")}
                r2 = await client.post(
                    f"{SUPABASE_URL}/rest/v1/pedidos", headers=headers, json=fallback
                )
                if r2.status_code not in (200, 201):
                    return  # No marcar como sincronizado si falló
        with get_conn() as conn:
            conn.execute("UPDATE pedidos SET sincronizado=1 WHERE id=?", (pedido_id,))
            conn.commit()
    except Exception:
        pass

async def sync_deliver_to_supabase(pedido_id: int):
    if not SUPABASE_URL or not SUPABASE_KEY or not HTTPX_AVAILABLE:
        return
    try:
        with get_conn() as conn:
            row = conn.execute("SELECT sync_id FROM pedidos WHERE id=?", (pedido_id,)).fetchone()
        if not row or not row["sync_id"]:
            return
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }
        async with httpx.AsyncClient(timeout=3.0) as client:
            await client.patch(
                f"{SUPABASE_URL}/rest/v1/pedidos?sync_id=eq.{row['sync_id']}",
                headers=headers,
                json={"estado": "entregado"},
            )
    except Exception:
        pass

async def sync_pending_loop():
    while True:
        await asyncio.sleep(300)
        if not SUPABASE_URL or not SUPABASE_KEY or not HTTPX_AVAILABLE:
            continue
        try:
            with get_conn() as conn:
                rows = conn.execute("SELECT * FROM pedidos WHERE sincronizado=0").fetchall()
            for row in rows:
                await sync_to_supabase(row["id"], dict(row))
        except Exception:
            pass

async def download_from_supabase():
    """Al arrancar: descarga pedidos de Supabase y detecta los que fallaron al subir."""
    if not SUPABASE_URL or not SUPABASE_KEY or not HTTPX_AVAILABLE:
        return
    try:
        headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
        async with httpx.AsyncClient(timeout=15.0) as client:
            r = await client.get(
                f"{SUPABASE_URL}/rest/v1/pedidos?select=*&order=fecha.asc&limit=5000",
                headers=headers,
            )
            if r.status_code != 200:
                return
            cloud_pedidos = r.json()

        cloud_sync_ids = {p.get("sync_id") for p in cloud_pedidos if p.get("sync_id")}

        with get_conn() as conn:
            # ── Resetea sincronizado=0 para pedidos marcados como subidos pero que no están en la nube ──
            local_synced = conn.execute(
                "SELECT id, sync_id FROM pedidos WHERE sincronizado=1 AND sync_id != '' AND sync_id IS NOT NULL"
            ).fetchall()
            for row in local_synced:
                if row["sync_id"] not in cloud_sync_ids:
                    conn.execute("UPDATE pedidos SET sincronizado=0 WHERE id=?", (row["id"],))

            # ── Descarga pedidos que no existen localmente ──
            for p in cloud_pedidos:
                sid = p.get("sync_id") or ""
                nf  = p.get("numero_factura") or ""

                # Buscar por sync_id primero, luego por numero_factura como fallback
                existing = None
                if sid:
                    existing = conn.execute(
                        "SELECT id, estado FROM pedidos WHERE sync_id=?", (sid,)
                    ).fetchone()
                if not existing and nf:
                    existing = conn.execute(
                        "SELECT id, estado FROM pedidos WHERE numero_factura=?", (nf,)
                    ).fetchone()

                if existing:
                    if p.get("estado") == "entregado" and existing["estado"] != "entregado":
                        conn.execute(
                            "UPDATE pedidos SET estado='entregado', sincronizado=1 WHERE id=?",
                            (existing["id"],)
                        )
                    elif sid and not conn.execute(
                        "SELECT id FROM pedidos WHERE sync_id=?", (sid,)
                    ).fetchone():
                        conn.execute("UPDATE pedidos SET sync_id=? WHERE id=?", (sid, existing["id"]))
                    continue

                # Insertar pedido nuevo descargado de la nube
                cur = conn.execute(
                    "INSERT INTO pedidos "
                    "(cliente,metodo_pago,total,estado,fecha,sincronizado,numero_factura,sync_id) "
                    "VALUES (?,?,?,?,?,1,?,?)",
                    (p["cliente"], p["metodo_pago"], float(p.get("total") or 0),
                     p.get("estado", "pendiente"), p["fecha"], nf, sid),
                )
                pedido_id = cur.lastrowid
                items = []
                try:
                    items = json.loads(p.get("items_json") or "[]")
                except Exception:
                    pass
                for it in items:
                    prod = conn.execute(
                        "SELECT id FROM productos WHERE nombre=?", (it.get("nombre", ""),)
                    ).fetchone()
                    if prod:
                        conn.execute(
                            "INSERT INTO detalle_pedidos "
                            "(pedido_id,producto_id,cantidad,observaciones) VALUES (?,?,?,?)",
                            (pedido_id, prod["id"],
                             it.get("cantidad", 1), it.get("observaciones", "")),
                        )
            conn.commit()
    except Exception:
        pass

# ─────────────────────────────────────────────
# APP
# ─────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    asyncio.create_task(download_from_supabase())
    asyncio.create_task(sync_pending_loop())
    yield

app = FastAPI(title="Drunks POS", lifespan=lifespan)

def to_dict(row) -> dict:
    d = dict(row)
    for k in ("disponible", "sincronizado", "tiene_base"):
        if k in d:
            d[k] = bool(d[k])
    return d

# ── Settings ──
@app.get("/api/settings")
def get_settings():
    with get_conn() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    return {r["key"]: r["value"] for r in rows}

@app.put("/api/settings/{key}")
async def set_setting(key: str, req: Request):
    data = await req.json()
    value = str(data.get("value", ""))
    with get_conn() as conn:
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
        conn.commit()
    return {"ok": True}

# ── Auto-Update ──
@app.get("/api/update/check")
def api_check_update():
    if _updater is None:
        return {"has_update": False, "current": APP_VERSION, "latest": None, "url": None}
    return _updater._update_info

@app.post("/api/update/apply")
async def api_apply_update(background_tasks: BackgroundTasks):
    if _updater is None:
        raise HTTPException(503, "Módulo de actualización no disponible")
    info = _updater._update_info
    if not info.get("has_update") or not info.get("url"):
        raise HTTPException(400, "No hay actualización disponible o URL no encontrada")
    background_tasks.add_task(_updater.download_and_apply, info["url"])
    return {"ok": True, "message": "Descargando actualización, la app se reiniciará en breve..."}

# ── Categorías ──
@app.get("/api/categorias")
def get_categorias():
    with get_conn() as conn:
        return [to_dict(r) for r in conn.execute("SELECT * FROM categorias ORDER BY nombre").fetchall()]

@app.post("/api/categorias", status_code=201)
def create_categoria(data: CategoriaBody):
    with get_conn() as conn:
        cur = conn.execute("INSERT INTO categorias (nombre) VALUES (?)", (data.nombre,))
        conn.commit()
        return to_dict(conn.execute("SELECT * FROM categorias WHERE id=?", (cur.lastrowid,)).fetchone())

@app.put("/api/categorias/{id}")
def update_categoria(id: int, data: CategoriaBody):
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM categorias WHERE id=?", (id,)).fetchone():
            raise HTTPException(404, "No encontrada")
        conn.execute("UPDATE categorias SET nombre=? WHERE id=?", (data.nombre, id))
        conn.commit()
        return to_dict(conn.execute("SELECT * FROM categorias WHERE id=?", (id,)).fetchone())

@app.delete("/api/categorias/{id}")
def delete_categoria(id: int):
    with get_conn() as conn:
        conn.execute("UPDATE productos SET categoria_id=NULL WHERE categoria_id=?", (id,))
        conn.execute("UPDATE notas_rapidas SET categoria_id=NULL WHERE categoria_id=?", (id,))
        conn.execute("DELETE FROM categorias WHERE id=?", (id,))
        conn.commit()
    return {"ok": True}

# ── Productos ──
PROD_Q = """
    SELECT p.*, COALESCE(c.nombre,'Sin Categoria') AS categoria_nombre
    FROM productos p LEFT JOIN categorias c ON p.categoria_id=c.id
"""

@app.get("/api/productos")
def get_productos():
    with get_conn() as conn:
        productos = [to_dict(r) for r in conn.execute(PROD_Q + " ORDER BY c.nombre, p.nombre").fetchall()]
        # Corregir inconsistencia: tiene_base=True pero sin entradas en producto_bases
        ids_con_bases = {r[0] for r in conn.execute(
            "SELECT DISTINCT producto_id FROM producto_bases").fetchall()}
        for p in productos:
            if p.get("tiene_base") and p["id"] not in ids_con_bases:
                p["tiene_base"] = False
                conn.execute("UPDATE productos SET tiene_base=0 WHERE id=?", (p["id"],))
        conn.commit()
        return productos

@app.post("/api/productos", status_code=201)
def create_producto(data: ProductoCreate):
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO productos (categoria_id, nombre, precio, disponible, tiene_base) VALUES (?,?,?,?,?)",
            (data.categoria_id, data.nombre, data.precio, 1 if data.disponible else 0, 1 if data.tiene_base else 0))
        conn.commit()
        return to_dict(conn.execute(PROD_Q + " WHERE p.id=?", (cur.lastrowid,)).fetchone())

@app.put("/api/productos/{id}")
def update_producto(id: int, data: ProductoUpdate):
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM productos WHERE id=?", (id,)).fetchone():
            raise HTTPException(404, "No encontrado")
        u: dict = {}
        if data.categoria_id is not None: u["categoria_id"] = data.categoria_id
        if data.nombre       is not None: u["nombre"]       = data.nombre
        if data.precio       is not None: u["precio"]       = data.precio
        if data.disponible   is not None: u["disponible"]   = 1 if data.disponible else 0
        if data.tiene_base   is not None: u["tiene_base"]   = 1 if data.tiene_base else 0
        if u:
            conn.execute(f"UPDATE productos SET {', '.join(k+'=?' for k in u)} WHERE id=?", [*u.values(), id])
            conn.commit()
        return to_dict(conn.execute(PROD_Q + " WHERE p.id=?", (id,)).fetchone())

@app.delete("/api/productos/{id}")
def delete_producto(id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM productos WHERE id=?", (id,))
        conn.commit()
    return {"ok": True}

@app.patch("/api/productos/{id}/toggle")
def toggle_producto(id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT disponible FROM productos WHERE id=?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "No encontrado")
        conn.execute("UPDATE productos SET disponible=? WHERE id=?", (0 if row["disponible"] else 1, id))
        conn.commit()
        return to_dict(conn.execute(PROD_Q + " WHERE p.id=?", (id,)).fetchone())

@app.patch("/api/productos/{id}/toggle-base")
def toggle_base_producto(id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT tiene_base FROM productos WHERE id=?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "No encontrado")
        conn.execute("UPDATE productos SET tiene_base=? WHERE id=?", (0 if row["tiene_base"] else 1, id))
        conn.commit()
        return to_dict(conn.execute(PROD_Q + " WHERE p.id=?", (id,)).fetchone())

# ── Tipos de Base ──
BASE_JOIN = """
    SELECT b.id, b.nombre, b.disponible, b.tipo_id,
           t.nombre AS tipo_nombre, t.icono AS tipo_icono
    FROM bases b
    LEFT JOIN tipos_base t ON b.tipo_id = t.id
"""

@app.get("/api/tipos-base")
def get_tipos_base():
    with get_conn() as conn:
        return [to_dict(r) for r in conn.execute(
            "SELECT * FROM tipos_base ORDER BY nombre").fetchall()]

@app.post("/api/tipos-base", status_code=201)
def create_tipo_base(data: TipoBaseCreate):
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(400, "Nombre requerido")
    with get_conn() as conn:
        existing = conn.execute("SELECT id FROM tipos_base WHERE LOWER(nombre)=LOWER(?)", (nombre,)).fetchone()
        if existing:
            raise HTTPException(409, "Ya existe un tipo con ese nombre")
        cur = conn.execute("INSERT INTO tipos_base (nombre, icono) VALUES (?,?)",
                           (nombre, data.icono.strip() or "🍹"))
        conn.commit()
        return to_dict(conn.execute("SELECT * FROM tipos_base WHERE id=?", (cur.lastrowid,)).fetchone())

@app.put("/api/tipos-base/{id}")
def update_tipo_base(id: int, data: TipoBaseUpdate):
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(400, "Nombre requerido")
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM tipos_base WHERE id=?", (id,)).fetchone():
            raise HTTPException(404, "Tipo no encontrado")
        dup = conn.execute("SELECT id FROM tipos_base WHERE LOWER(nombre)=LOWER(?) AND id!=?", (nombre, id)).fetchone()
        if dup:
            raise HTTPException(409, "Ya existe un tipo con ese nombre")
        conn.execute("UPDATE tipos_base SET nombre=?, icono=? WHERE id=?",
                     (nombre, data.icono.strip() or "🍹", id))
        conn.commit()
        return to_dict(conn.execute("SELECT * FROM tipos_base WHERE id=?", (id,)).fetchone())

@app.delete("/api/tipos-base/{id}")
def delete_tipo_base(id: int):
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM bases WHERE tipo_id=?", (id,)).fetchone()[0]
        if count:
            raise HTTPException(409, f"Este tipo tiene {count} base(s). Elimínalas primero.")
        conn.execute("DELETE FROM tipos_base WHERE id=?", (id,))
        conn.commit()
    return {"ok": True}

# ── Bases ──
@app.get("/api/bases")
def get_bases():
    with get_conn() as conn:
        return [to_dict(r) for r in conn.execute(
            BASE_JOIN + " ORDER BY t.nombre, b.nombre").fetchall()]

@app.post("/api/bases", status_code=201)
def create_base(data: BaseCreate):
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(400, "Nombre requerido")
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM tipos_base WHERE id=?", (data.tipo_id,)).fetchone():
            raise HTTPException(400, "Tipo de base no válido")
        tipo_text = conn.execute("SELECT nombre FROM tipos_base WHERE id=?", (data.tipo_id,)).fetchone()["nombre"]
        cur = conn.execute("INSERT INTO bases (nombre, tipo, tipo_id) VALUES (?,?,?)",
                           (nombre, tipo_text, data.tipo_id))
        conn.commit()
        return to_dict(conn.execute(BASE_JOIN + " WHERE b.id=?", (cur.lastrowid,)).fetchone())

@app.put("/api/bases/{id}")
def update_base(id: int, data: BaseUpdate):
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(400, "Nombre requerido")
    with get_conn() as conn:
        if not conn.execute("SELECT id FROM bases WHERE id=?", (id,)).fetchone():
            raise HTTPException(404, "Base no encontrada")
        if not conn.execute("SELECT id FROM tipos_base WHERE id=?", (data.tipo_id,)).fetchone():
            raise HTTPException(400, "Tipo de base no válido")
        tipo_text = conn.execute("SELECT nombre FROM tipos_base WHERE id=?", (data.tipo_id,)).fetchone()["nombre"]
        conn.execute("UPDATE bases SET nombre=?, tipo=?, tipo_id=? WHERE id=?",
                     (nombre, tipo_text, data.tipo_id, id))
        conn.commit()
        return to_dict(conn.execute(BASE_JOIN + " WHERE b.id=?", (id,)).fetchone())

@app.delete("/api/bases/{id}")
def delete_base(id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM bases WHERE id=?", (id,))
        conn.commit()
    return {"ok": True}

@app.patch("/api/bases/{id}/toggle")
def toggle_base_disponible(id: int):
    with get_conn() as conn:
        row = conn.execute("SELECT disponible FROM bases WHERE id=?", (id,)).fetchone()
        if not row:
            raise HTTPException(404, "Base no encontrada")
        conn.execute("UPDATE bases SET disponible=? WHERE id=?",
                     (0 if row["disponible"] else 1, id))
        conn.commit()
        return to_dict(conn.execute(BASE_JOIN + " WHERE b.id=?", (id,)).fetchone())

@app.get("/api/productos/{id}/bases")
def get_producto_bases(id: int):
    with get_conn() as conn:
        assigned_ids = {r["base_id"] for r in conn.execute(
            "SELECT base_id FROM producto_bases WHERE producto_id=?", (id,)).fetchall()}
        all_bases = [to_dict(r) for r in conn.execute(
            BASE_JOIN + " ORDER BY t.nombre, b.nombre").fetchall()]
        if assigned_ids:
            for b in all_bases:
                b["assigned"] = b["id"] in assigned_ids
        else:
            all_bases = [dict(**b, assigned=False) for b in all_bases if b["disponible"]]
        return all_bases

@app.put("/api/productos/{id}/bases")
def set_producto_bases(id: int, data: ProductoBasesUpdate):
    with get_conn() as conn:
        conn.execute("DELETE FROM producto_bases WHERE producto_id=?", (id,))
        if data.base_ids:
            conn.executemany(
                "INSERT INTO producto_bases (producto_id, base_id) VALUES (?,?)",
                [(id, bid) for bid in data.base_ids])
        conn.commit()
    return {"ok": True, "count": len(data.base_ids)}

@app.get("/api/productos/bases-bulk")
def get_all_producto_bases():
    with get_conn() as conn:
        rows = conn.execute("SELECT producto_id, base_id FROM producto_bases").fetchall()
    result: dict[int, list[int]] = {}
    for r in rows:
        result.setdefault(r["producto_id"], []).append(r["base_id"])
    return result

# ── Notas Rápidas ──
@app.get("/api/notas_rapidas")
def get_notas():
    with get_conn() as conn:
        return [to_dict(r) for r in conn.execute("SELECT * FROM notas_rapidas ORDER BY categoria_id NULLS LAST, id").fetchall()]

@app.post("/api/notas_rapidas", status_code=201)
def create_nota(data: NotaBody):
    with get_conn() as conn:
        cur = conn.execute("INSERT INTO notas_rapidas (texto, categoria_id) VALUES (?,?)", (data.texto, data.categoria_id))
        conn.commit()
        return to_dict(conn.execute("SELECT * FROM notas_rapidas WHERE id=?", (cur.lastrowid,)).fetchone())

@app.delete("/api/notas_rapidas/{id}")
def delete_nota(id: int):
    with get_conn() as conn:
        conn.execute("DELETE FROM notas_rapidas WHERE id=?", (id,))
        conn.commit()
    return {"ok": True}

# ── Pedidos ──
@app.post("/api/pedidos", status_code=201)
async def create_pedido(data: PedidoCreate, background_tasks: BackgroundTasks):
    fecha   = datetime.now().isoformat()
    sync_id = str(uuid.uuid4())
    items_detail: list[dict] = []
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO pedidos (cliente,metodo_pago,total,estado,fecha,sincronizado,numero_factura,sync_id) VALUES (?,?,?,?,?,?,'',?)",
            (data.cliente, data.metodo_pago, data.total, "pendiente", fecha, 0, sync_id))
        pedido_id = cur.lastrowid
        numero_factura = f"DRK-{pedido_id:05d}"
        conn.execute("UPDATE pedidos SET numero_factura=? WHERE id=?", (numero_factura, pedido_id))
        for item in data.items:
            conn.execute(
                "INSERT INTO detalle_pedidos (pedido_id,producto_id,cantidad,observaciones) VALUES (?,?,?,?)",
                (pedido_id, item.producto_id, item.cantidad, item.observaciones))
            prod = conn.execute("""
                SELECT p.nombre, p.precio, p.categoria_id, COALESCE(c.nombre,'Sin Categoria') AS categoria
                FROM productos p LEFT JOIN categorias c ON p.categoria_id=c.id
                WHERE p.id=?
            """, (item.producto_id,)).fetchone()
            items_detail.append({
                "producto_id":   item.producto_id,
                "nombre":        prod["nombre"]       if prod else "Desconocido",
                "cantidad":      item.cantidad,
                "precio":        prod["precio"]       if prod else 0,
                "categoria_id":  prod["categoria_id"] if prod else None,
                "categoria":     prod["categoria"]    if prod else "Sin Categoria",
                "observaciones": item.observaciones,
            })
        conn.commit()
    payload = {
        "id": pedido_id, "numero_factura": numero_factura,
        "cliente": data.cliente, "metodo_pago": data.metodo_pago,
        "total": data.total, "estado": "pendiente", "fecha": fecha, "items": items_detail,
    }
    await manager.broadcast({"type": "new_order", "order": payload})
    background_tasks.add_task(sync_to_supabase, pedido_id, {
        "sync_id": sync_id, "numero_factura": numero_factura,
        "cliente": data.cliente, "metodo_pago": data.metodo_pago,
        "total": data.total, "estado": "pendiente", "fecha": fecha,
        "items_json": json.dumps(items_detail),
    })
    return payload

@app.get("/api/pedidos/pendientes")
def get_pendientes():
    with get_conn() as conn:
        pedidos = conn.execute("SELECT * FROM pedidos WHERE estado='pendiente' ORDER BY fecha DESC").fetchall()
        result  = []
        for p in pedidos:
            pd = dict(p)
            pd["items"] = [dict(d) for d in conn.execute(
                """SELECT dp.producto_id, dp.cantidad, dp.observaciones,
                          pr.nombre, pr.categoria_id
                   FROM detalle_pedidos dp JOIN productos pr ON dp.producto_id=pr.id
                   WHERE dp.pedido_id=?""", (p["id"],)).fetchall()]
            result.append(pd)
    return result

@app.get("/api/pedidos/{id}")
def get_pedido(id: int):
    with get_conn() as conn:
        p = conn.execute("SELECT * FROM pedidos WHERE id=?", (id,)).fetchone()
        if not p:
            raise HTTPException(404, "Pedido no encontrado")
        pd = dict(p)
        pd["items"] = [dict(d) for d in conn.execute("""
            SELECT dp.cantidad, dp.observaciones,
                   pr.nombre, pr.precio, dp.cantidad * pr.precio AS subtotal,
                   COALESCE(c.nombre,'Sin Categoria') AS categoria
            FROM detalle_pedidos dp
            JOIN  productos   pr ON dp.producto_id  = pr.id
            LEFT JOIN categorias c  ON pr.categoria_id = c.id
            WHERE dp.pedido_id = ?
        """, (id,)).fetchall()]
    return pd

@app.put("/api/pedidos/{id}/entregar")
async def entregar_pedido(id: int, background_tasks: BackgroundTasks):
    with get_conn() as conn:
        conn.execute("UPDATE pedidos SET estado='entregado' WHERE id=?", (id,))
        conn.commit()
    background_tasks.add_task(sync_deliver_to_supabase, id)
    return {"ok": True}

@app.get("/api/sync/status")
async def sync_status():
    with get_conn() as conn:
        pending = conn.execute("SELECT COUNT(*) FROM pedidos WHERE sincronizado=0").fetchone()[0]
    if not SUPABASE_URL or not SUPABASE_KEY:
        return {"configured": False, "reachable": False, "pending": pending}
    reachable = False
    if HTTPX_AVAILABLE:
        try:
            headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
            async with httpx.AsyncClient(timeout=3.0) as client:
                r = await client.get(f"{SUPABASE_URL}/rest/v1/pedidos?select=id&limit=1", headers=headers)
                reachable = r.status_code == 200
        except Exception:
            reachable = False
    return {"configured": True, "reachable": reachable, "pending": pending}

@app.get("/api/dashboard")
def get_dashboard(range: str = "week"):
    with get_conn() as conn:
        today = datetime.now().strftime("%Y-%m-%d")
        if   range == "today": df = f"AND DATE(fecha)='{today}'"
        elif range == "month": df = "AND DATE(fecha)>=DATE('now','-29 days')"
        elif range == "all":   df = ""
        else:                  df = "AND DATE(fecha)>=DATE('now','-6 days')"

        def q(sql, *a): return conn.execute(sql, a).fetchone()[0]
        def qa(sql):    return [dict(r) for r in conn.execute(sql).fetchall()]

        ventas_hoy     = q(f"SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado' AND DATE(fecha)='{today}'")
        pedidos_hoy    = q(f"SELECT COUNT(*) FROM pedidos WHERE DATE(fecha)='{today}'")
        pendientes_now = q("SELECT COUNT(*) FROM pedidos WHERE estado='pendiente'")
        ventas_periodo = q(f"SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado' {df}")
        pedidos_periodo= q(f"SELECT COUNT(*) FROM pedidos WHERE estado='entregado' {df}")
        ticket_prom    = q("SELECT COALESCE(AVG(total),0) FROM pedidos WHERE estado='entregado'")
        total_todo     = q("SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado'")

        if range == "all":
            daily = qa("""
                SELECT DATE(fecha) as dia, COALESCE(SUM(total),0) as total, COUNT(*) as cnt
                FROM pedidos WHERE estado='entregado'
                GROUP BY DATE(fecha) ORDER BY dia DESC LIMIT 30""")
            daily.reverse()
        else:
            daily = qa(f"""
                SELECT DATE(fecha) as dia, COALESCE(SUM(total),0) as total, COUNT(*) as cnt
                FROM pedidos WHERE estado='entregado' {df}
                GROUP BY DATE(fecha) ORDER BY dia""")

        hourly = qa("""
            SELECT CAST(strftime('%H',fecha) AS INTEGER) as hora,
                   COUNT(*) as cnt, COALESCE(SUM(total),0) as total
            FROM pedidos WHERE estado='entregado'
            GROUP BY hora ORDER BY hora""")

        top_prods = qa(f"""
            SELECT pr.nombre, SUM(dp.cantidad) as vendidos, SUM(dp.cantidad*pr.precio) as revenue
            FROM detalle_pedidos dp
            JOIN productos pr ON dp.producto_id=pr.id
            JOIN pedidos p ON dp.pedido_id=p.id
            WHERE p.estado='entregado' {df}
            GROUP BY pr.id, pr.nombre ORDER BY vendidos DESC LIMIT 8""")

        metodos = qa(f"""
            SELECT metodo_pago, COUNT(*) as cnt, COALESCE(SUM(total),0) as total
            FROM pedidos WHERE estado='entregado' {df}
            GROUP BY metodo_pago""")

        by_cat = qa(f"""
            SELECT COALESCE(c.nombre,'Sin Categoria') as nombre,
                   SUM(dp.cantidad) as vendidos, SUM(dp.cantidad*pr.precio) as revenue
            FROM detalle_pedidos dp
            JOIN productos pr ON dp.producto_id=pr.id
            JOIN pedidos p ON dp.pedido_id=p.id
            LEFT JOIN categorias c ON pr.categoria_id=c.id
            WHERE p.estado='entregado' {df}
            GROUP BY c.id, c.nombre ORDER BY revenue DESC""")

        recientes = qa(f"""
            SELECT id, cliente, metodo_pago, total, estado, fecha
            FROM pedidos WHERE 1=1 {df.replace("estado='entregado' AND","").replace("AND estado='entregado'","")}
            ORDER BY id DESC LIMIT 30""")

    return {
        "range":           range,
        "ventas_hoy":      float(ventas_hoy),
        "pedidos_hoy":     int(pedidos_hoy),
        "pendientes_now":  int(pendientes_now),
        "ventas_periodo":  float(ventas_periodo),
        "pedidos_periodo": int(pedidos_periodo),
        "ticket_prom":     float(ticket_prom),
        "total_todo":      float(total_todo),
        "daily":           daily,
        "hourly":          hourly,
        "top_prods":       top_prods,
        "metodos":         metodos,
        "by_cat":          by_cat,
        "recientes":       recientes,
    }

# ── Exportar Excel ──
@app.get("/api/export/excel")
def export_excel():
    if not XLSX_AVAILABLE:
        raise HTTPException(503, "openpyxl no está instalado")

    now = datetime.now()

    # ── Helpers de estilo ──
    C_PURPLE   = "4C1D95"
    C_PURP_LT  = "F0EBFF"
    C_PURP_MED = "7C3AED"
    C_GREEN    = "166534"
    C_RED      = "991B1B"
    C_AMBER    = "92400E"

    def h_font():  return Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    def h_fill():  return PatternFill("solid", fgColor=C_PURPLE)
    def h_align(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def row_fill(i): return PatternFill("solid", fgColor=C_PURP_LT if i % 2 == 0 else "FFFFFF")
    def thin(color="D8B4FE"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def apply_header(ws, cols, row=1):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = h_font(); cell.fill = h_fill(); cell.alignment = h_align()
        ws.row_dimensions[row].height = 22

    def auto_col(ws, mn=10, mx=55):
        for col in ws.columns:
            w = mn
            for cell in col:
                if cell.value is not None:
                    w = max(w, min(len(str(cell.value)) + 4, mx))
            ws.column_dimensions[get_column_letter(col[0].column)].width = w

    # ── Queries ──
    with get_conn() as conn:
        pedidos = conn.execute(
            "SELECT id, cliente, metodo_pago, total, estado, fecha FROM pedidos ORDER BY fecha DESC"
        ).fetchall()

        detalle = conn.execute("""
            SELECT p.id AS pedido_id, p.cliente, p.metodo_pago, p.fecha,
                   pr.nombre AS producto, COALESCE(c.nombre,'Sin Categoria') AS categoria,
                   dp.cantidad, pr.precio, dp.cantidad*pr.precio AS subtotal, dp.observaciones
            FROM detalle_pedidos dp
            JOIN pedidos  p  ON dp.pedido_id   = p.id
            JOIN productos pr ON dp.producto_id = pr.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            ORDER BY p.fecha DESC, p.id, pr.nombre
        """).fetchall()

        productos = conn.execute("""
            SELECT p.id, p.nombre, COALESCE(c.nombre,'Sin Categoria') AS categoria,
                   p.precio, p.disponible, p.tiene_base
            FROM productos p LEFT JOIN categorias c ON p.categoria_id=c.id
            ORDER BY c.nombre, p.nombre
        """).fetchall()

        kpi_ingresos  = conn.execute("SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado'").fetchone()[0]
        kpi_pedidos   = conn.execute("SELECT COUNT(*) FROM pedidos").fetchone()[0]
        kpi_ticket    = conn.execute("SELECT COALESCE(AVG(total),0) FROM pedidos WHERE estado='entregado'").fetchone()[0]
        kpi_pendiente = conn.execute("SELECT COUNT(*) FROM pedidos WHERE estado='pendiente'").fetchone()[0]

        metodos_sum = conn.execute("""
            SELECT metodo_pago, COUNT(*) AS cnt, COALESCE(SUM(total),0) AS total
            FROM pedidos WHERE estado='entregado' GROUP BY metodo_pago ORDER BY total DESC
        """).fetchall()

        top5 = conn.execute("""
            SELECT pr.nombre, COALESCE(c.nombre,'Sin Categoria') AS cat,
                   SUM(dp.cantidad) AS uds, SUM(dp.cantidad*pr.precio) AS rev
            FROM detalle_pedidos dp
            JOIN productos  pr ON dp.producto_id = pr.id
            JOIN pedidos    p  ON dp.pedido_id   = p.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            WHERE p.estado='entregado'
            GROUP BY pr.id ORDER BY rev DESC LIMIT 5
        """).fetchall()

        by_cat = conn.execute("""
            SELECT COALESCE(c.nombre,'Sin Categoria') AS cat,
                   COUNT(DISTINCT p.id) AS pedidos, SUM(dp.cantidad) AS uds,
                   SUM(dp.cantidad*pr.precio) AS rev
            FROM detalle_pedidos dp
            JOIN productos  pr ON dp.producto_id = pr.id
            JOIN pedidos    p  ON dp.pedido_id   = p.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            WHERE p.estado='entregado'
            GROUP BY c.id ORDER BY rev DESC
        """).fetchall()

    wb = Workbook()

    # ═══════════════════════════════════════
    # HOJA 1 — RESUMEN
    # ═══════════════════════════════════════
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "DRUNKS POS  —  REPORTE DE VENTAS"
    t.font      = Font(bold=True, size=20, color=C_PURP_MED, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill      = PatternFill("solid", fgColor="FAF5FF")
    ws.row_dimensions[1].height = 46

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"Generado el {now.strftime('%d/%m/%Y a las %H:%M:%S')}"
    s.font      = Font(italic=True, size=10, color="9CA3AF", name="Calibri")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # KPIs
    r = 4
    ws.cell(r, 1).value = "ESTADÍSTICAS GENERALES"
    ws.cell(r, 1).font  = Font(bold=True, size=13, color=C_PURP_MED, name="Calibri")
    ws.row_dimensions[r].height = 22
    r += 1

    kpis = [
        ("Total de Pedidos Registrados", int(kpi_pedidos),   False, C_PURPLE),
        ("Ingresos Totales (entregados)", float(kpi_ingresos), True,  C_GREEN),
        ("Ticket Promedio",               float(kpi_ticket),   True,  C_PURP_MED),
        ("Pedidos Pendientes Actualmente",int(kpi_pendiente), False, C_AMBER),
    ]
    for label, value, money, color in kpis:
        cl = ws.cell(r, 1, label)
        cl.font = Font(size=11, color="374151", name="Calibri")
        cl.fill = PatternFill("solid", fgColor="F9FAFB")
        cv = ws.cell(r, 2, round(value) if money else value)
        cv.font          = Font(bold=True, size=13, color=color, name="Calibri")
        cv.number_format = "#,##0" if money else "0"
        cv.alignment     = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 22
        r += 1

    # Ventas por método
    r += 1
    ws.cell(r, 1).value = "VENTAS POR MÉTODO DE PAGO"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    ws.row_dimensions[r].height = 20
    r += 1
    apply_header(ws, ["Método de Pago", "# Pedidos", "Total Ingresos"], row=r)
    r += 1
    for i, m in enumerate(metodos_sum):
        ws.cell(r+i, 1, m["metodo_pago"]).fill = row_fill(i)
        ws.cell(r+i, 2, m["cnt"]).fill          = row_fill(i)
        tc = ws.cell(r+i, 3, round(m["total"]))
        tc.number_format = "#,##0"; tc.font = Font(bold=True, color=C_GREEN); tc.fill = row_fill(i)
    r += len(metodos_sum) + 2

    # Por categoría
    ws.cell(r, 1).value = "INGRESOS POR CATEGORÍA"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    r += 1
    apply_header(ws, ["Categoría", "Pedidos", "Unidades", "Ingresos"], row=r)
    r += 1
    for i, cat in enumerate(by_cat):
        ws.cell(r+i, 1, cat["cat"]).fill       = row_fill(i)
        ws.cell(r+i, 2, cat["pedidos"]).fill    = row_fill(i)
        ws.cell(r+i, 3, cat["uds"]).fill        = row_fill(i)
        rc = ws.cell(r+i, 4, round(cat["rev"]))
        rc.number_format = "#,##0"; rc.font = Font(bold=True, color=C_GREEN); rc.fill = row_fill(i)
    r += len(by_cat) + 2

    # Top 5 productos
    ws.cell(r, 1).value = "TOP 5 PRODUCTOS MÁS VENDIDOS"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    r += 1
    apply_header(ws, ["Producto", "Categoría", "Unidades Vendidas", "Ingresos Generados"], row=r)
    r += 1
    for i, p in enumerate(top5):
        ws.cell(r+i, 1, p["nombre"]).fill = row_fill(i)
        ws.cell(r+i, 2, p["cat"]).fill    = row_fill(i)
        ws.cell(r+i, 3, p["uds"]).fill    = row_fill(i)
        ri = ws.cell(r+i, 4, round(p["rev"]))
        ri.number_format = "#,##0"; ri.font = Font(bold=True, color=C_GREEN); ri.fill = row_fill(i)

    auto_col(ws)

    # ═══════════════════════════════════════
    # HOJA 2 — PEDIDOS
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("Pedidos")
    ws2.sheet_view.showGridLines = False
    cols2 = ["#", "Cliente", "Método de Pago", "Total (COP)", "Estado", "Fecha", "Hora"]
    apply_header(ws2, cols2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
    ws2.row_dimensions[1].height = 22

    for i, p in enumerate(pedidos):
        row_n = i + 2
        try:    dt = datetime.fromisoformat(str(p["fecha"]))
        except: dt = now

        bg = row_fill(i)
        entregado = p["estado"] == "entregado"

        ws2.cell(row_n, 1, p["id"]).fill  = bg
        ws2.cell(row_n, 2, p["cliente"]).fill = bg
        ws2.cell(row_n, 3, p["metodo_pago"]).fill = bg

        tc = ws2.cell(row_n, 4, round(p["total"]))
        tc.number_format = "#,##0"; tc.font = Font(bold=True, color=C_GREEN); tc.fill = bg

        ec = ws2.cell(row_n, 5, "Entregado" if entregado else "Pendiente")
        ec.font = Font(bold=True, color=C_GREEN if entregado else C_AMBER)
        ec.fill = PatternFill("solid", fgColor="D1FAE5" if entregado else "FEF3C7")
        ec.alignment = Alignment(horizontal="center")

        ws2.cell(row_n, 6, dt.strftime("%d/%m/%Y")).fill = bg
        ws2.cell(row_n, 7, dt.strftime("%H:%M")).fill    = bg
        ws2.row_dimensions[row_n].height = 18

    auto_col(ws2)

    # ═══════════════════════════════════════
    # HOJA 3 — DETALLE DE PEDIDOS
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("Detalle de Pedidos")
    ws3.sheet_view.showGridLines = False
    cols3 = ["Pedido #", "Cliente", "Método", "Fecha y Hora", "Producto",
             "Categoría", "Cantidad", "Precio Unit.", "Subtotal", "Base / Notas"]
    apply_header(ws3, cols3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"
    ws3.row_dimensions[1].height = 22

    for i, d in enumerate(detalle):
        row_n = i + 2
        try:    dt = datetime.fromisoformat(str(d["fecha"]))
        except: dt = now
        bg = row_fill(i)

        ws3.cell(row_n, 1, d["pedido_id"]).fill = bg
        ws3.cell(row_n, 2, d["cliente"]).fill    = bg
        ws3.cell(row_n, 3, d["metodo_pago"]).fill= bg
        ws3.cell(row_n, 4, dt.strftime("%d/%m/%Y %H:%M")).fill = bg
        ws3.cell(row_n, 5, d["producto"]).fill   = bg
        ws3.cell(row_n, 6, d["categoria"]).fill  = bg
        ws3.cell(row_n, 7, d["cantidad"]).fill   = bg

        pu = ws3.cell(row_n, 8, round(d["precio"]))
        pu.number_format = "#,##0"; pu.fill = bg

        sb = ws3.cell(row_n, 9, round(d["subtotal"]))
        sb.number_format = "#,##0"; sb.font = Font(bold=True, color=C_GREEN); sb.fill = bg

        ws3.cell(row_n, 10, d["observaciones"] or "").fill = bg
        ws3.row_dimensions[row_n].height = 18

    auto_col(ws3)

    # ═══════════════════════════════════════
    # HOJA 4 — PRODUCTOS
    # ═══════════════════════════════════════
    ws4 = wb.create_sheet("Productos")
    ws4.sheet_view.showGridLines = False
    cols4 = ["#", "Nombre del Producto", "Categoría", "Precio (COP)", "Estado", "Requiere Base"]
    apply_header(ws4, cols4)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"

    for i, p in enumerate(productos):
        row_n = i + 2
        bg    = row_fill(i)
        activo = bool(p["disponible"])
        base   = bool(p["tiene_base"])

        ws4.cell(row_n, 1, p["id"]).fill     = bg
        ws4.cell(row_n, 2, p["nombre"]).fill  = bg
        ws4.cell(row_n, 3, p["categoria"]).fill = bg

        pc = ws4.cell(row_n, 4, round(p["precio"]))
        pc.number_format = "#,##0"; pc.font = Font(bold=True, color=C_GREEN); pc.fill = bg

        dc = ws4.cell(row_n, 5, "Activo" if activo else "Agotado")
        dc.font = Font(bold=True, color=C_GREEN if activo else C_RED)
        dc.fill = PatternFill("solid", fgColor="D1FAE5" if activo else "FEE2E2")
        dc.alignment = Alignment(horizontal="center")

        bc = ws4.cell(row_n, 6, "Sí" if base else "No")
        bc.font = Font(bold=True, color="D97706" if base else "6B7280")
        bc.fill = PatternFill("solid", fgColor="FEF3C7" if base else "F9FAFB")
        bc.alignment = Alignment(horizontal="center")

        ws4.row_dimensions[row_n].height = 18

    auto_col(ws4)

    # ── Generar archivo ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"Drunks_Reporte_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

# ── WebSocket ──
@app.websocket("/ws")
async def ws_endpoint(ws: WebSocket):
    await manager.connect(ws)
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(ws)

# ─────────────────────────────────────────────
# FRONTEND — DASHBOARD  (ruta propia /dashboard)
# ─────────────────────────────────────────────
DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Drunks · Dashboard</title>
<script>document.documentElement.setAttribute('data-theme',localStorage.getItem('drunksPalette')||'dark')</script>
<style>
html[data-theme=light] body{background:#f0ebff!important;color:#1e1b4b!important}
html[data-theme=light] [class*=bg-gray-95]{background:#eee9ff!important}
html[data-theme=light] [class*=bg-gray-90]{background:#fff!important}
html[data-theme=light] [class*=bg-gray-8]{background:#ede9fe!important}
html[data-theme=light] [class*=bg-gray-7]{background:#ddd6fe!important}
html[data-theme=light] [class*=border-gray-8]{border-color:#c4b5fd!important}
html[data-theme=light] [class*=border-gray-7]{border-color:#ddd6fe!important}
html[data-theme=light] [class*=bg-black]{background:rgba(80,50,180,.55)!important}
html[data-theme=light] .text-white{color:#1e1b4b!important}
html[data-theme=light] [class*=text-gray-]{color:#5b21b6!important}
html[data-theme=ocean] body{background:#020b18!important}
html[data-theme=ocean] [class*=bg-purple-6]{background:#2563eb!important}
html[data-theme=ocean] [class*=bg-purple-7]{background:#1d4ed8!important}
html[data-theme=ocean] [class*=bg-purple-8]{background:#1e40af!important}
html[data-theme=ocean] [class*=bg-purple-9]{background:#1e3a8a!important}
html[data-theme=ocean] [class*=text-purple]{color:#60a5fa!important}
html[data-theme=ocean] [class*=border-purple]{border-color:#3b82f6!important}
html[data-theme=ocean] [class*=from-purple]{--tw-gradient-from:#1e3a8a!important}
html[data-theme=ocean] [class*=via-purple]{--tw-gradient-via:#1d4ed8!important}
html[data-theme=ocean] [class*=glow-purple]{box-shadow:0 0 18px rgba(37,99,235,.4)!important}
html[data-theme=emerald] body{background:#020f0b!important}
html[data-theme=emerald] [class*=bg-purple-6]{background:#0d9488!important}
html[data-theme=emerald] [class*=bg-purple-7]{background:#0f766e!important}
html[data-theme=emerald] [class*=bg-purple-8]{background:#115e59!important}
html[data-theme=emerald] [class*=bg-purple-9]{background:#134e4a!important}
html[data-theme=emerald] [class*=text-purple]{color:#2dd4bf!important}
html[data-theme=emerald] [class*=border-purple]{border-color:#14b8a6!important}
html[data-theme=emerald] [class*=from-purple]{--tw-gradient-from:#134e4a!important}
html[data-theme=emerald] [class*=via-purple]{--tw-gradient-via:#0f766e!important}
html[data-theme=emerald] [class*=glow-purple]{box-shadow:0 0 18px rgba(13,148,136,.4)!important}
html[data-theme=amber] body{background:#0d0800!important}
html[data-theme=amber] [class*=bg-purple-6]{background:#d97706!important}
html[data-theme=amber] [class*=bg-purple-7]{background:#b45309!important}
html[data-theme=amber] [class*=bg-purple-8]{background:#92400e!important}
html[data-theme=amber] [class*=bg-purple-9]{background:#78350f!important}
html[data-theme=amber] [class*=text-purple]{color:#fbbf24!important}
html[data-theme=amber] [class*=border-purple]{border-color:#f59e0b!important}
html[data-theme=amber] [class*=from-purple]{--tw-gradient-from:#78350f!important}
html[data-theme=amber] [class*=via-purple]{--tw-gradient-via:#b45309!important}
html[data-theme=amber] [class*=glow-purple]{box-shadow:0 0 18px rgba(217,119,6,.4)!important}
html[data-theme=neon] body{background:#040804!important}html[data-theme=neon] [class*=bg-purple-6]{background:#16a34a!important}html[data-theme=neon] [class*=bg-purple-7]{background:#15803d!important}html[data-theme=neon] [class*=text-purple]{color:#4ade80!important}html[data-theme=neon] [class*=border-purple]{border-color:#22c55e!important}html[data-theme=neon] [class*=from-purple]{--tw-gradient-from:#14532d!important}
html[data-theme=rose] body{background:#0f0208!important}html[data-theme=rose] [class*=bg-purple-6]{background:#e11d48!important}html[data-theme=rose] [class*=bg-purple-7]{background:#be123c!important}html[data-theme=rose] [class*=text-purple]{color:#fb7185!important}html[data-theme=rose] [class*=border-purple]{border-color:#f43f5e!important}html[data-theme=rose] [class*=from-purple]{--tw-gradient-from:#881337!important}
html[data-theme=violet] body{background:#06030f!important}html[data-theme=violet] [class*=bg-purple-6]{background:#6d28d9!important}html[data-theme=violet] [class*=bg-purple-7]{background:#5b21b6!important}html[data-theme=violet] [class*=text-purple]{color:#a78bfa!important}html[data-theme=violet] [class*=border-purple]{border-color:#7c3aed!important}html[data-theme=violet] [class*=from-purple]{--tw-gradient-from:#3b0764!important}
html[data-theme=slate] body{background:#0a0c10!important}html[data-theme=slate] [class*=bg-purple-6]{background:#475569!important}html[data-theme=slate] [class*=bg-purple-7]{background:#334155!important}html[data-theme=slate] [class*=text-purple]{color:#94a3b8!important}html[data-theme=slate] [class*=border-purple]{border-color:#64748b!important}html[data-theme=slate] [class*=from-purple]{--tw-gradient-from:#0f172a!important}
html[data-theme=gold] body{background:#0d0900!important}html[data-theme=gold] [class*=bg-purple-6]{background:#ca8a04!important}html[data-theme=gold] [class*=bg-purple-7]{background:#a16207!important}html[data-theme=gold] [class*=text-purple]{color:#fde047!important}html[data-theme=gold] [class*=border-purple]{border-color:#eab308!important}html[data-theme=gold] [class*=from-purple]{--tw-gradient-from:#713f12!important}
</style>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
<style>
  body{background:#06060f;font-family:'Segoe UI',system-ui,sans-serif}
  ::-webkit-scrollbar{width:5px}::-webkit-scrollbar-thumb{background:#374151;border-radius:4px}
  .kpi-card{transition:transform .15s ease,box-shadow .15s ease}
  .kpi-card:hover{transform:translateY(-2px)}
  @keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
  .fade-in{animation:fadeIn .35s ease forwards}
  .range-btn{transition:all .15s ease}
</style>
</head>
<body class="min-h-screen text-white">

<!-- ═══ HEADER ═══ -->
<div class="sticky top-0 z-20 bg-gray-950/97 backdrop-blur border-b border-gray-800 px-6 py-3.5">
  <div class="max-w-screen-2xl mx-auto flex items-center justify-between">
    <div class="flex items-center gap-3">
      <div class="w-9 h-9 rounded-xl bg-purple-600 flex items-center justify-center text-lg shadow-lg shadow-purple-900/60 shrink-0">🍹</div>
      <div>
        <h1 class="font-extrabold text-base leading-none">DRUNKS <span class="text-purple-400 font-light text-sm">· Dashboard</span></h1>
        <p id="lastUpdate" class="text-gray-600 text-xs mt-0.5">—</p>
      </div>
    </div>
    <div class="flex items-center gap-2">
      <button onclick="toggleThemePicker()" title="Paleta de colores"
        class="bg-gray-800 hover:bg-gray-700 border border-gray-700 text-white px-3 py-2 rounded-xl text-sm transition-colors">🎨</button>
      <button id="exportBtn" onclick="exportExcel()"
        class="bg-green-700 hover:bg-green-600 text-white px-4 py-2 rounded-xl text-xs font-bold transition-colors shadow-lg shadow-green-950/50 flex items-center gap-1.5">
        <span>📥</span><span>Exportar Excel</span>
      </button>
      <button id="refreshBtn" onclick="refresh()"
        class="bg-purple-700 hover:bg-purple-600 text-white px-4 py-2 rounded-xl text-xs font-bold transition-colors shadow-lg shadow-purple-950/50">
        ↻
      </button>
    </div>
  </div>
</div>

<!-- ═══ FILTRO DE RANGO ═══ -->
<div class="bg-gray-950/80 border-b border-gray-800/50 px-6 py-2.5">
  <div class="max-w-screen-2xl mx-auto flex items-center gap-2 flex-wrap">
    <span class="text-gray-600 text-[10px] tracking-[0.2em] font-bold mr-1">PERIODO</span>
    <button data-r="today" onclick="setRange('today')" class="range-btn px-4 py-1.5 rounded-full text-xs font-bold bg-gray-800 border border-gray-700 text-gray-400">Hoy</button>
    <button data-r="week"  onclick="setRange('week')"  class="range-btn px-4 py-1.5 rounded-full text-xs font-bold bg-purple-600 text-white shadow-lg">Semana</button>
    <button data-r="month" onclick="setRange('month')" class="range-btn px-4 py-1.5 rounded-full text-xs font-bold bg-gray-800 border border-gray-700 text-gray-400">Mes</button>
    <button data-r="all"   onclick="setRange('all')"   class="range-btn px-4 py-1.5 rounded-full text-xs font-bold bg-gray-800 border border-gray-700 text-gray-400">Todo</button>
    <div id="spinner" class="hidden ml-2 flex items-center gap-1.5 text-gray-500 text-xs">
      <div class="w-2.5 h-2.5 rounded-full bg-purple-500 animate-ping"></div>cargando
    </div>
  </div>
</div>

<!-- ═══ CONTENIDO ═══ -->
<div class="max-w-screen-2xl mx-auto px-6 py-5 space-y-5">

  <!-- KPIs -->
  <div id="kpiRow" class="grid grid-cols-2 sm:grid-cols-3 lg:grid-cols-5 gap-4"></div>

  <!-- Ingresos + Métodos de pago -->
  <div class="grid grid-cols-1 xl:grid-cols-3 gap-4">
    <div class="xl:col-span-2 bg-gray-900 border border-gray-800 rounded-2xl p-5">
      <div class="flex items-center justify-between mb-4">
        <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold">INGRESOS POR DIA</h3>
        <span id="chartSubtitle" class="text-gray-600 text-xs"></span>
      </div>
      <div class="relative" style="height:240px">
        <canvas id="revenueChart"></canvas>
      </div>
    </div>
    <div class="bg-gray-900 border border-gray-800 rounded-2xl p-5 flex flex-col">
      <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold mb-4">METODOS DE PAGO</h3>
      <div class="flex-1 flex items-center justify-center">
        <div class="relative w-48 h-48">
          <canvas id="paymentChart"></canvas>
          <div id="paymentCenter" class="absolute inset-0 flex flex-col items-center justify-center pointer-events-none">
            <div class="text-white font-extrabold text-lg tabular-nums" id="paymentTotal">—</div>
            <div class="text-gray-500 text-xs">total</div>
          </div>
        </div>
      </div>
      <div id="paymentLegend" class="mt-3 space-y-2"></div>
    </div>
  </div>

  <!-- Top productos + Por hora -->
  <div class="grid grid-cols-1 lg:grid-cols-2 gap-4">
    <div class="bg-gray-900 border border-gray-800 rounded-2xl p-5">
      <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold mb-4">TOP PRODUCTOS</h3>
      <div class="relative" style="height:260px">
        <canvas id="productsChart"></canvas>
      </div>
    </div>
    <div class="bg-gray-900 border border-gray-800 rounded-2xl p-5">
      <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold mb-1">PEDIDOS POR HORA</h3>
      <p class="text-gray-700 text-xs mb-4">Historico completo — pico de demanda</p>
      <div class="relative" style="height:260px">
        <canvas id="hourlyChart"></canvas>
      </div>
    </div>
  </div>

  <!-- Por categoría -->
  <div class="bg-gray-900 border border-gray-800 rounded-2xl p-5">
    <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold mb-4">INGRESOS POR CATEGORIA</h3>
    <div class="relative" style="height:180px">
      <canvas id="categoryChart"></canvas>
    </div>
  </div>

  <!-- Tabla de pedidos recientes -->
  <div class="bg-gray-900 border border-gray-800 rounded-2xl p-5">
    <div class="flex items-center justify-between mb-4">
      <h3 class="text-gray-400 text-[10px] tracking-[0.2em] font-bold">PEDIDOS RECIENTES</h3>
      <span id="recCount" class="text-gray-600 text-xs"></span>
    </div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm min-w-[520px]">
        <thead>
          <tr class="text-gray-600 text-[10px] tracking-widest border-b border-gray-800">
            <th class="text-left py-2 pr-4">FACTURA</th>
            <th class="text-left py-2 pr-4">CLIENTE</th>
            <th class="text-left py-2 pr-4">METODO</th>
            <th class="text-right py-2 pr-4">TOTAL</th>
            <th class="text-center py-2 pr-4">ESTADO</th>
            <th class="text-left py-2">FECHA / HORA</th>
            <th class="text-center py-2">VER</th>
          </tr>
        </thead>
        <tbody id="recTable"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- ═══ MODAL FACTURA ═══ -->
<div id="invModal" class="fixed inset-0 z-50 hidden items-center justify-center p-4 bg-black/85 backdrop-blur-sm">
  <div class="bg-gray-900 border border-purple-700/30 rounded-3xl w-full max-w-md shadow-2xl shadow-purple-950/50 overflow-hidden flex flex-col max-h-[90vh]">

    <!-- Cabecera factura -->
    <div class="bg-gradient-to-br from-purple-900 via-purple-800 to-indigo-900 px-6 pt-6 pb-5 shrink-0">
      <div class="flex items-start justify-between mb-4">
        <div class="flex items-center gap-2">
          <div class="w-8 h-8 rounded-xl bg-white/20 flex items-center justify-center text-base">🍹</div>
          <div>
            <div class="text-purple-200 text-[10px] font-bold tracking-[0.2em]">DRUNKS POS</div>
            <div class="text-white/50 text-[9px]">Sistema de Ventas</div>
          </div>
        </div>
        <button onclick="closeInv()" class="text-purple-300 hover:text-white text-2xl leading-none w-8 h-8 flex items-center justify-center rounded-lg hover:bg-white/10 transition-all">x</button>
      </div>
      <div class="text-purple-300 text-[10px] tracking-[0.25em] font-bold mb-0.5">FACTURA</div>
      <div id="inv-num" class="text-white font-extrabold text-3xl tracking-wider leading-none mb-3"></div>
      <div class="grid grid-cols-2 gap-x-4 gap-y-1 text-sm">
        <div><span class="text-purple-300/70 text-[10px]">CLIENTE</span><div id="inv-cli" class="text-white font-bold leading-snug"></div></div>
        <div><span class="text-purple-300/70 text-[10px]">MÉTODO</span><div id="inv-met" class="font-bold leading-snug"></div></div>
        <div><span class="text-purple-300/70 text-[10px]">FECHA Y HORA</span><div id="inv-dt" class="text-purple-100 text-xs leading-snug"></div></div>
        <div><span class="text-purple-300/70 text-[10px]">ESTADO</span><div id="inv-est" class="font-bold text-xs leading-snug"></div></div>
      </div>
    </div>

    <!-- Cuerpo -->
    <div class="flex-1 overflow-y-auto">
      <!-- Loading -->
      <div id="inv-loading" class="flex items-center justify-center py-12 text-gray-500 text-sm gap-2">
        <div class="w-3 h-3 rounded-full bg-purple-500 animate-ping"></div>Cargando factura...
      </div>
      <!-- Items -->
      <div id="inv-body" class="hidden px-6 py-4">
        <div class="text-gray-500 text-[10px] tracking-widest font-bold mb-3">PRODUCTOS PEDIDOS</div>
        <div id="inv-items" class="space-y-0 divide-y divide-gray-800/60"></div>
      </div>
    </div>

    <!-- Footer total -->
    <div id="inv-footer" class="hidden shrink-0 border-t border-gray-800 bg-gray-900/80 px-6 py-4">
      <div class="flex justify-between items-center mb-3">
        <span class="text-gray-400 text-sm font-semibold">TOTAL FACTURA</span>
        <span id="inv-total" class="text-white font-extrabold text-2xl tabular-nums"></span>
      </div>
      <div class="grid grid-cols-2 gap-2">
        <button onclick="printInv()" class="bg-gray-800 hover:bg-gray-700 text-gray-300 font-semibold py-2.5 rounded-xl text-sm transition-colors">🖨 Imprimir</button>
        <button onclick="closeInv()" class="bg-purple-700 hover:bg-purple-600 text-white font-bold py-2.5 rounded-xl text-sm transition-colors">Cerrar</button>
      </div>
    </div>
  </div>
</div>

<!-- PRINT AREA (hidden, only visible on print) -->
<div id="printArea" class="hidden"></div>
<style>
@media print {
  body > *:not(#printArea) { display: none !important; }
  #printArea { display: block !important; font-family: 'Courier New', monospace; max-width: 320px; margin: 0 auto; }
  #printArea h1 { font-size: 18px; font-weight: bold; text-align: center; margin-bottom: 4px; }
  #printArea .line { border-top: 1px dashed #000; margin: 6px 0; }
  #printArea .row { display: flex; justify-content: space-between; font-size: 12px; margin: 3px 0; }
  #printArea .total { font-size: 15px; font-weight: bold; }
}
</style>

<script>
// ── Globals ──
let currentRange = 'week';
const charts = {};
Chart.defaults.color = '#6b7280';
Chart.defaults.borderColor = '#1f2937';

const PALETTE = ['#9333ea','#3b82f6','#22c55e','#f59e0b','#ef4444','#ec4899','#14b8a6','#8b5cf6'];
const TOOLTIP = {
  backgroundColor:'#111827', borderColor:'#374151', borderWidth:1,
  titleColor:'#f9fafb', bodyColor:'#d1d5db', padding:10,
};

// ── Range selector ──
function setRange(r) {
  currentRange = r;
  document.querySelectorAll('.range-btn').forEach(b => {
    const active = b.dataset.r === r;
    b.className = 'range-btn px-4 py-1.5 rounded-full text-xs font-bold ' +
      (active ? 'bg-purple-600 text-white shadow-lg' : 'bg-gray-800 border border-gray-700 text-gray-400');
  });
  loadData();
}

function refresh() {
  const btn = document.getElementById('refreshBtn');
  btn.textContent = '↻';
  loadData().finally(() => { btn.textContent = 'Actualizar'; });
}

// ── Data loading ──
async function loadData() {
  document.getElementById('spinner').classList.remove('hidden');
  try {
    const r    = await fetch('/api/dashboard?range=' + currentRange);
    const data = await r.json();
    render(data);
    document.getElementById('lastUpdate').textContent =
      'Actualizado: ' + new Date().toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});
  } catch(e) {
    console.error(e);
  } finally {
    document.getElementById('spinner').classList.add('hidden');
  }
}

// ── Formatters ──
function fmt(n)  { return '$' + Math.round(n||0).toLocaleString('es-CO'); }
function fmtK(n) { return n>=1000000 ? '$'+(n/1000000).toFixed(1)+'M' : n>=1000 ? '$'+(n/1000).toFixed(0)+'k' : '$'+Math.round(n); }
function fmtDay(s) {
  return new Date(s+'T12:00:00').toLocaleDateString('es-CO',{weekday:'short',day:'numeric',month:'short'});
}
function esc(s) { return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

// ── Chart helper ──
function upsert(id, type, data, options) {
  if (charts[id]) {
    charts[id].data    = data;
    charts[id].options = { ...charts[id].options, ...options };
    charts[id].update('active');
  } else {
    charts[id] = new Chart(document.getElementById(id), { type, data, options });
  }
}

// ── Main render ──
function render(d) {
  renderKPIs(d);
  renderRevenue(d);
  renderPayment(d);
  renderProducts(d);
  renderHourly(d);
  renderCategory(d);
  renderTable(d);
}

// ── KPIs ──
function renderKPIs(d) {
  const rangeLabels = { today:'Hoy', week:'Esta semana', month:'Este mes', all:'Historico' };
  const lbl = rangeLabels[d.range] || '';
  const cards = [
    { icon:'💰', label:'Ventas ' + lbl, value:fmt(d.ventas_periodo), sub:d.pedidos_periodo+' pedidos completados', color:'from-purple-900/40 to-purple-900/10 border-purple-700/30' },
    { icon:'📅', label:'Ventas Hoy',    value:fmt(d.ventas_hoy),     sub:d.pedidos_hoy+' pedidos hoy',            color:'from-blue-900/40 to-blue-900/10 border-blue-700/30' },
    { icon:'🧾', label:'Ticket Prom.',  value:fmt(d.ticket_prom),    sub:'Promedio por pedido',                   color:'from-green-900/40 to-green-900/10 border-green-700/30' },
    { icon:'📦', label:'Total General', value:fmtK(d.total_todo),    sub:'Ingresos historicos',                   color:'from-amber-900/40 to-amber-900/10 border-amber-700/30' },
    { icon:'⏳', label:'En Espera',     value:d.pendientes_now,      sub:'pedidos pendientes ahora',              color:'from-red-900/40 to-red-900/10 border-red-700/30', raw:true },
  ];
  document.getElementById('kpiRow').innerHTML = cards.map(c => `
    <div class="kpi-card fade-in bg-gradient-to-br ${c.color} border rounded-2xl p-4">
      <div class="text-2xl mb-2">${c.icon}</div>
      <div class="text-gray-400 text-[10px] tracking-widest font-bold mb-1 leading-tight">${c.label.toUpperCase()}</div>
      <div class="text-white font-extrabold ${c.raw?'text-3xl':'text-2xl'} tabular-nums leading-none">${c.value}</div>
      <div class="text-gray-600 text-xs mt-1.5">${c.sub}</div>
    </div>`).join('');
}

// ── Revenue chart (line) ──
function renderRevenue(d) {
  const labels   = d.daily.map(x => fmtDay(x.dia));
  const values   = d.daily.map(x => x.total);
  const subtitle = d.daily.length ? d.daily.length+' dias con ventas' : 'Sin datos';
  document.getElementById('chartSubtitle').textContent = subtitle;

  upsert('revenueChart', 'bar', {
    labels,
    datasets: [{
      label: 'Ingresos',
      data: values,
      backgroundColor: ctx => {
        const g = ctx.chart.ctx.createLinearGradient(0,0,0,240);
        g.addColorStop(0,'rgba(147,51,234,.85)');
        g.addColorStop(1,'rgba(147,51,234,.1)');
        return g;
      },
      borderColor: '#9333ea',
      borderWidth: 1,
      borderRadius: 6,
      hoverBackgroundColor: '#a855f7',
    }]
  }, {
    responsive:true, maintainAspectRatio:false,
    plugins:{ legend:{display:false}, tooltip:{ ...TOOLTIP, callbacks:{ label: c=>' '+fmt(c.raw) }}},
    scales:{
      x:{ ticks:{color:'#6b7280',maxRotation:40,font:{size:10}}, grid:{color:'#111827'} },
      y:{ ticks:{color:'#6b7280',callback:v=>fmtK(v),font:{size:10}}, grid:{color:'#111827'} }
    }
  });
}

// ── Payment doughnut ──
function renderPayment(d) {
  if (!d.metodos || !d.metodos.length) {
    document.getElementById('paymentTotal').textContent = '—';
    document.getElementById('paymentLegend').innerHTML = '<p class="text-gray-700 text-xs text-center">Sin datos</p>';
    return;
  }
  const labels = d.metodos.map(m => m.metodo_pago);
  const counts = d.metodos.map(m => m.cnt);
  const totals = d.metodos.map(m => m.total);
  const colors = [d.metodos[0]?.metodo_pago==='Efectivo'?'#22c55e':'#3b82f6', '#3b82f6','#9333ea'];

  const grandTotal = totals.reduce((a,b)=>a+b,0);
  document.getElementById('paymentTotal').textContent = fmtK(grandTotal);

  upsert('paymentChart','doughnut',{
    labels,
    datasets:[{data:counts, backgroundColor:colors, borderColor:'#06060f', borderWidth:3, hoverOffset:6}]
  },{
    responsive:true, maintainAspectRatio:false, cutout:'68%',
    plugins:{
      legend:{display:false},
      tooltip:{...TOOLTIP, callbacks:{label:c=>' '+c.label+': '+c.raw+' pedidos'}}
    }
  });

  const grandCount = counts.reduce((a,b)=>a+b,0)||1;
  document.getElementById('paymentLegend').innerHTML = d.metodos.map((m,i)=>`
    <div class="flex items-center justify-between text-xs py-1">
      <div class="flex items-center gap-2">
        <span class="w-2.5 h-2.5 rounded-full shrink-0" style="background:${colors[i]}"></span>
        <span class="text-gray-300">${m.metodo_pago==='Efectivo'?'💵':'📱'} ${m.metodo_pago}</span>
      </div>
      <div class="text-right">
        <span class="text-white font-bold tabular-nums">${Math.round(m.cnt/grandCount*100)}%</span>
        <span class="text-gray-600 ml-2 tabular-nums">${fmt(m.total)}</span>
      </div>
    </div>`).join('');
}

// ── Top products (horizontal bar) ──
function renderProducts(d) {
  if (!d.top_prods || !d.top_prods.length) return;
  const labels = d.top_prods.map(p => p.nombre);
  const vend   = d.top_prods.map(p => p.vendidos);
  const rev    = d.top_prods.map(p => p.revenue);

  upsert('productsChart','bar',{
    labels,
    datasets:[
      { label:'Unidades', data:vend, backgroundColor:'rgba(147,51,234,.75)', borderColor:'#9333ea', borderWidth:1, borderRadius:4, yAxisID:'y' },
      { label:'Ingresos', data:rev,  backgroundColor:'rgba(59,130,246,.35)',  borderColor:'#3b82f6', borderWidth:1, borderRadius:4, yAxisID:'y2', type:'line', tension:0.4, pointRadius:3, pointBackgroundColor:'#3b82f6' },
    ]
  },{
    indexAxis:'y', responsive:true, maintainAspectRatio:false,
    plugins:{
      legend:{ labels:{color:'#9ca3af', boxWidth:10, padding:12}, position:'top' },
      tooltip:{ ...TOOLTIP, callbacks:{ label:c => c.datasetIndex===0 ? ' '+c.raw+' uds' : ' '+fmt(c.raw) }}
    },
    scales:{
      x:{ ticks:{color:'#6b7280',font:{size:10}}, grid:{color:'#111827'} },
      y:{ ticks:{color:'#9ca3af',font:{size:10}}, grid:{display:false} },
      y2:{display:false}
    }
  });
}

// ── Hourly bar ──
function renderHourly(d) {
  const all   = Array.from({length:24},(_,i)=>i);
  const cnts  = new Array(24).fill(0);
  const tots  = new Array(24).fill(0);
  (d.hourly||[]).forEach(h => { cnts[h.hora]=h.cnt; tots[h.hora]=h.total; });
  const labels = all.map(h=>(h<10?'0':'')+h+'h');

  upsert('hourlyChart','bar',{
    labels,
    datasets:[{
      label:'Pedidos',
      data:cnts,
      backgroundColor: cnts.map(v => {
        const m = Math.max(...cnts,1);
        const pct = v/m;
        if (pct>0.8) return 'rgba(239,68,68,.85)';
        if (pct>0.5) return 'rgba(245,158,11,.75)';
        return 'rgba(147,51,234,.6)';
      }),
      borderRadius:4,
    }]
  },{
    responsive:true, maintainAspectRatio:false,
    plugins:{
      legend:{display:false},
      tooltip:{ ...TOOLTIP, callbacks:{
        label:c=>` ${c.raw} pedidos · ${fmt(tots[c.dataIndex])}`
      }}
    },
    scales:{
      x:{ticks:{color:'#6b7280',font:{size:9},maxRotation:0}, grid:{display:false}},
      y:{ticks:{color:'#6b7280',font:{size:10},stepSize:1}, grid:{color:'#111827'}}
    }
  });
}

// ── Category bar ──
function renderCategory(d) {
  if (!d.by_cat || !d.by_cat.length) return;
  const labels = d.by_cat.map(c => c.nombre);
  const rev    = d.by_cat.map(c => c.revenue);
  const vend   = d.by_cat.map(c => c.vendidos);

  upsert('categoryChart','bar',{
    labels,
    datasets:[
      { label:'Ingresos',  data:rev,  backgroundColor:PALETTE.map(c=>c+'cc'), borderColor:PALETTE, borderWidth:1, borderRadius:6 },
      { label:'Unidades',  data:vend, backgroundColor:'transparent', borderColor:PALETTE, borderWidth:2, borderRadius:6, type:'line', tension:0.4, pointRadius:5, pointBackgroundColor:PALETTE, yAxisID:'y2' },
    ]
  },{
    responsive:true, maintainAspectRatio:false,
    plugins:{
      legend:{labels:{color:'#9ca3af',boxWidth:10,padding:14}, position:'top'},
      tooltip:{ ...TOOLTIP, callbacks:{ label:c => c.datasetIndex===0 ? ' '+fmt(c.raw) : ' '+c.raw+' uds' }}
    },
    scales:{
      x:{ticks:{color:'#9ca3af',font:{size:11}}, grid:{display:false}},
      y:{ticks:{color:'#6b7280',callback:v=>fmtK(v),font:{size:10}}, grid:{color:'#111827'}},
      y2:{position:'right', ticks:{color:'#6b7280',font:{size:10}}, grid:{display:false}}
    }
  });
}

// ── Recent orders table ──
function renderTable(d) {
  const tbody = document.getElementById('recTable');
  document.getElementById('recCount').textContent = d.recientes.length + ' pedidos';
  if (!d.recientes.length) {
    tbody.innerHTML = '<tr><td colspan="7" class="py-10 text-center text-gray-700 text-sm">Sin pedidos en este periodo</td></tr>';
    return;
  }
  tbody.innerHTML = d.recientes.map(r => {
    const isPend = r.estado === 'pendiente';
    const dt     = new Date(r.fecha);
    const fecha  = dt.toLocaleDateString('es-CO', {day:'2-digit', month:'2-digit', year:'2-digit'});
    const hora   = dt.toLocaleTimeString('es-CO', {hour:'2-digit', minute:'2-digit'});
    const factura = r.numero_factura || ('#' + String(r.id).padStart(5,'0'));
    return `<tr class="border-b border-gray-800/30 hover:bg-purple-900/10 transition-colors cursor-pointer" onclick="openInv(${r.id})">
      <td class="py-2.5 pr-4">
        <span class="text-purple-400 text-xs font-extrabold tabular-nums tracking-wide font-mono">${factura}</span>
      </td>
      <td class="py-2.5 pr-4 text-white text-sm font-semibold">${esc(r.cliente)}</td>
      <td class="py-2.5 pr-4 text-gray-400 text-xs whitespace-nowrap">${r.metodo_pago==='Efectivo'?'💵':'📱'} ${r.metodo_pago}</td>
      <td class="py-2.5 pr-4 text-right">
        <span class="text-green-400 text-sm font-extrabold tabular-nums">${fmt(r.total)}</span>
      </td>
      <td class="py-2.5 pr-4 text-center">
        <span class="text-[11px] font-bold px-2.5 py-1 rounded-full ${isPend?'bg-yellow-900/50 text-yellow-400':'bg-green-900/50 text-green-400'}">
          ${isPend?'Pendiente':'Entregado'}
        </span>
      </td>
      <td class="py-2.5 pr-4 text-gray-500 text-xs tabular-nums whitespace-nowrap">${fecha} ${hora}</td>
      <td class="py-2.5 text-center">
        <span class="text-purple-500 hover:text-purple-300 text-sm transition-colors">👁</span>
      </td>
    </tr>`;
  }).join('');
}

// ── Factura Modal ──
let currentInvData = null;

async function openInv(id) {
  // Reset & show modal
  document.getElementById('inv-loading').classList.remove('hidden');
  document.getElementById('inv-body').classList.add('hidden');
  document.getElementById('inv-footer').classList.add('hidden');
  document.getElementById('inv-num').textContent = '...';
  document.getElementById('invModal').classList.remove('hidden');
  document.getElementById('invModal').classList.add('flex');

  try {
    const r = await fetch('/api/pedidos/' + id);
    if (!r.ok) throw new Error('Not found');
    const o = await r.json();
    currentInvData = o;

    const factura = o.numero_factura || ('#' + String(o.id).padStart(5,'0'));
    const dt      = new Date(o.fecha);
    const entregado = o.estado === 'entregado';

    document.getElementById('inv-num').textContent  = factura;
    document.getElementById('inv-cli').textContent  = o.cliente;

    const metEl = document.getElementById('inv-met');
    metEl.textContent  = (o.metodo_pago==='Efectivo'?'💵 ':'📱 ') + o.metodo_pago;
    metEl.className    = 'font-bold leading-snug ' + (o.metodo_pago==='Efectivo'?'text-green-400':'text-blue-400');

    document.getElementById('inv-dt').textContent  =
      dt.toLocaleDateString('es-CO',{day:'2-digit',month:'long',year:'numeric'}) +
      '  ' + dt.toLocaleTimeString('es-CO',{hour:'2-digit',minute:'2-digit'});

    const estEl = document.getElementById('inv-est');
    estEl.textContent = entregado ? '✓ Entregado' : '⏳ Pendiente';
    estEl.className   = 'font-bold text-xs leading-snug ' + (entregado?'text-green-400':'text-yellow-400');

    // Items
    document.getElementById('inv-items').innerHTML = (o.items||[]).map(item => {
      const parts = (item.observaciones||'').split(' | ');
      const isBase = s => s && /^[^:]+: /.test(s);
      const base   = isBase(parts[0]) ? parts[0] : null;
      const notas  = base ? parts.slice(1).join(' | ') : (item.observaciones||'');
      return `<div class="flex items-start gap-3 py-3">
        <span class="bg-purple-800/60 text-purple-200 text-xs font-extrabold px-2.5 py-1 rounded-xl shrink-0 min-w-[2.5rem] text-center">${item.cantidad}x</span>
        <div class="flex-1 min-w-0">
          <div class="text-white font-semibold text-sm leading-snug">${esc(item.nombre)}</div>
          ${base  ? '<div class="text-amber-400 text-xs mt-0.5 font-bold">📌 ' + base + '</div>' : ''}
          ${notas ? '<div class="text-yellow-300 text-xs mt-0.5">📝 ' + esc(notas) + '</div>' : ''}
          <div class="text-gray-600 text-xs mt-1 tabular-nums">${fmt(item.precio)} por unidad</div>
        </div>
        <div class="text-right shrink-0">
          <div class="text-green-400 font-extrabold text-sm tabular-nums">${fmt(item.subtotal)}</div>
        </div>
      </div>`;
    }).join('');

    document.getElementById('inv-total').textContent = fmt(o.total);

    document.getElementById('inv-loading').classList.add('hidden');
    document.getElementById('inv-body').classList.remove('hidden');
    document.getElementById('inv-footer').classList.remove('hidden');
  } catch(e) {
    document.getElementById('inv-loading').textContent = 'Error al cargar la factura';
  }
}

function closeInv() {
  document.getElementById('invModal').classList.add('hidden');
  document.getElementById('invModal').classList.remove('flex');
  currentInvData = null;
}

function printInv() {
  if (!currentInvData) return;
  const o = currentInvData;
  const factura = o.numero_factura || ('#' + String(o.id).padStart(5,'0'));
  const dt = new Date(o.fecha).toLocaleString('es-CO',{day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'});
  const items = (o.items||[]).map(i => {
    const obs = i.observaciones ? ' (' + i.observaciones + ')' : '';
    return `<div class="row"><span>${i.cantidad}x ${i.nombre}${obs}</span><span>$${Math.round(i.subtotal).toLocaleString('es-CO')}</span></div>`;
  }).join('');
  document.getElementById('printArea').innerHTML = `
    <h1>🍹 DRUNKS POS</h1>
    <div class="row"><span>Factura:</span><span>${factura}</span></div>
    <div class="row"><span>Cliente:</span><span>${o.cliente}</span></div>
    <div class="row"><span>Método:</span><span>${o.metodo_pago}</span></div>
    <div class="row"><span>Fecha:</span><span>${dt}</span></div>
    <div class="line"></div>
    ${items}
    <div class="line"></div>
    <div class="row total"><span>TOTAL</span><span>$${Math.round(o.total).toLocaleString('es-CO')}</span></div>
    <div class="line"></div>
    <p style="text-align:center;font-size:10px;margin-top:8px">¡Gracias por tu compra!</p>
  `;
  document.getElementById('printArea').classList.remove('hidden');
  window.print();
  document.getElementById('printArea').classList.add('hidden');
}

// Close modal on backdrop click
document.addEventListener('click', e => {
  if (e.target.id === 'invModal') closeInv();
});

// ── Export Excel ──
async function exportExcel() {
  const btn  = document.getElementById('exportBtn');
  const orig = btn.innerHTML;
  btn.innerHTML = '<span>⏳</span><span>Generando...</span>';
  btn.disabled  = true;
  try {
    const r = await fetch('/api/export/excel');
    if (!r.ok) { alert('Error al generar el archivo'); return; }
    const blob  = await r.blob();
    const cd    = r.headers.get('Content-Disposition') || '';
    const match = cd.match(/filename="([^"]+)"/);
    const fname = match ? match[1] : 'Drunks_Reporte.xlsx';
    const url   = URL.createObjectURL(blob);
    const a     = document.createElement('a');
    a.href = url; a.download = fname; a.click();
    URL.revokeObjectURL(url);
  } catch(e) {
    alert('Error de conexion al exportar');
  } finally {
    btn.innerHTML = orig;
    btn.disabled  = false;
  }
}

// ── Settings DB ──
function saveSetting(key, value) {
  localStorage.setItem(key, value);
  fetch('/api/settings/'+key, {method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).catch(()=>{});
}
async function loadSettings() {
  try {
    const s = await fetch('/api/settings').then(r=>r.json());
    if (s.palette || s.drunksPalette) applyTheme(s.palette || s.drunksPalette);
    Object.entries(s).forEach(([k,v]) => localStorage.setItem(k, v));
  } catch(e) { applyTheme(localStorage.getItem('drunksPalette') || 'dark'); }
}
// ── Theme System ──
function applyTheme(t) {
  const theme = t||'dark';
  if (theme.startsWith('custom:')) { applyCustomTheme(theme.split(':')[1]); return; }
  document.documentElement.setAttribute('data-theme', theme);
  saveSetting('drunksPalette', theme); saveSetting('palette', theme);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b => {
    const on = b.dataset.theme === theme;
    b.style.background = on ? 'rgba(147,51,234,.2)' : '';
    b.style.outline    = on ? '1px solid rgba(147,51,234,.5)' : '';
  });
  const picker = document.getElementById('customColorPicker');
  if (picker) picker.style.display = 'none';
}
function applyCustomTheme(hex) {
  let style = document.getElementById('custom-theme-style');
  if (!style) { style = document.createElement('style'); style.id='custom-theme-style'; document.head.appendChild(style); }
  const h = hex||'#9333ea';
  const ri=parseInt(h.slice(1,3),16),gi=parseInt(h.slice(3,5),16),bi=parseInt(h.slice(5,7),16);
  style.textContent = `html[data-theme="custom"] body{background:#06060f}
    html[data-theme="custom"] .bg-purple-600,.bg-purple-700{background-color:${h}!important}
    html[data-theme="custom"] .text-purple-400,.text-purple-300{color:${h}!important}
    html[data-theme="custom"] .border-purple-500{border-color:${h}!important}`;
  document.documentElement.setAttribute('data-theme','custom');
  saveSetting('drunksPalette','custom:'+h); saveSetting('palette','custom:'+h);
  document.querySelectorAll('.theme-opt[data-theme]').forEach(b=>{b.style.background='';b.style.outline='';});
  const picker = document.getElementById('customColorPicker');
  if (picker) { picker.style.display='inline-block'; picker.value=h; }
}
function toggleThemePicker() {
  document.getElementById('themePicker').classList.toggle('hidden');
}
document.addEventListener('click', e => {
  if (!e.target.closest('#themePicker') && !e.target.closest('[onclick*="toggleThemePicker"]'))
    document.getElementById('themePicker')?.classList.add('hidden');
});

loadSettings().then(() => { loadData(); setInterval(loadData, 60000); });

// ── Auto-update ──
async function checkUpdate() {
  try {
    const d = await fetch('/api/update/check').then(r => r.json());
    if (d.has_update && d.latest) {
      document.getElementById('updateVersion').textContent = 'v'+d.current+' → v'+d.latest;
      document.getElementById('updateBanner').classList.remove('hidden');
    }
  } catch(e) {}
}
async function applyUpdate() {
  const btn = document.getElementById('updateBtn');
  btn.textContent = 'Descargando...'; btn.disabled = true;
  try {
    await fetch('/api/update/apply', {method:'POST'});
    btn.textContent = 'Reiniciando...';
  } catch(e) { btn.textContent = 'Error — intenta de nuevo'; btn.disabled = false; }
}
checkUpdate();
setInterval(checkUpdate, 30 * 60 * 1000);
</script>

<!-- ═══ THEME PICKER ═══ -->
<div id="themePicker" class="hidden fixed z-[90]" style="top:4.5rem;right:1rem">
  <div class="bg-gray-900 border border-gray-800 rounded-2xl p-4 shadow-2xl shadow-black/60 w-52">
    <div class="text-gray-500 text-[10px] tracking-[.2em] font-bold mb-3">PALETA DE COLORES</div>
    <div class="space-y-0.5">
      <button data-theme="dark"    onclick="applyTheme('dark')"    class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors">
        <span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#06060f"></span>
        <span class="text-white text-sm font-semibold">Oscuro</span>
      </button>
      <button data-theme="light"   onclick="applyTheme('light')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors">
        <span class="w-5 h-5 rounded-full shrink-0 border border-gray-600" style="background:#f0ebff"></span>
        <span class="text-white text-sm font-semibold">Claro</span>
      </button>
      <button data-theme="ocean"   onclick="applyTheme('ocean')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors">
        <span class="w-5 h-5 rounded-full shrink-0" style="background:#2563eb"></span>
        <span class="text-white text-sm font-semibold">Océano</span>
      </button>
      <button data-theme="emerald" onclick="applyTheme('emerald')" class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors">
        <span class="w-5 h-5 rounded-full shrink-0" style="background:#0d9488"></span>
        <span class="text-white text-sm font-semibold">Esmeralda</span>
      </button>
      <button data-theme="amber"   onclick="applyTheme('amber')"   class="theme-opt w-full flex items-center gap-3 px-3 py-2 rounded-xl hover:bg-gray-800 transition-colors">
        <span class="w-5 h-5 rounded-full shrink-0" style="background:#d97706"></span>
        <span class="text-white text-sm font-semibold">Ámbar</span>
      </button>
    </div>
  </div>
</div>
<!-- ═══ BANNER DE ACTUALIZACIÓN ═══ -->
<div id="updateBanner" class="hidden fixed bottom-0 left-0 right-0 z-[100] bg-purple-950/97 border-t border-purple-500/50 backdrop-blur px-5 py-3 flex items-center justify-between gap-3 shadow-2xl">
  <div class="flex items-center gap-3">
    <span class="text-2xl">🔄</span>
    <div>
      <div class="text-white font-bold text-sm">Nueva versión disponible</div>
      <div id="updateVersion" class="text-purple-300 text-xs font-mono"></div>
    </div>
  </div>
  <div class="flex gap-2 shrink-0">
    <button id="updateBtn" onclick="applyUpdate()"
      class="bg-purple-600 hover:bg-purple-500 active:bg-purple-700 text-white text-sm font-extrabold px-4 py-2 rounded-xl transition-colors shadow-lg">
      Actualizar ahora
    </button>
    <button onclick="document.getElementById('updateBanner').classList.add('hidden')"
      class="text-purple-400 hover:text-white text-sm px-3 py-2 rounded-xl transition-colors">
      Después
    </button>
  </div>
</div>
</body>
</html>"""

# ── Páginas ──
@app.get("/vendedor", response_class=HTMLResponse)
def page_vendedor(): return HTMLResponse(VENDEDOR_HTML)

@app.get("/cocina", response_class=HTMLResponse)
def page_cocina(): return HTMLResponse(COCINA_HTML)

@app.get("/dashboard", response_class=HTMLResponse)
def page_dashboard(): return HTMLResponse(DASHBOARD_HTML)

LAUNCHER_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Drunks POS</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{background:#0d0d1a;font-family:'Segoe UI',sans-serif;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:100vh;color:#fff}
h1{font-size:2.2rem;font-weight:800;margin-bottom:.4rem;letter-spacing:-1px}
.sub{color:#8b8ba8;margin-bottom:3rem;font-size:1rem}
.cards{display:flex;gap:1.5rem;flex-wrap:wrap;justify-content:center}
.card{background:#1a1a2e;border:1px solid #2d2d4a;border-radius:1.25rem;padding:2.2rem 2.8rem;cursor:pointer;text-decoration:none;color:#fff;transition:.18s;text-align:center;min-width:180px}
.card:hover{background:#2d2d4a;border-color:#7c3aed;transform:translateY(-3px);box-shadow:0 8px 32px rgba(124,58,237,.3)}
.card .icon{font-size:2.6rem;margin-bottom:.8rem}
.card .name{font-size:1.1rem;font-weight:700}
.card .desc{font-size:.8rem;color:#8b8ba8;margin-top:.3rem}
.badge{display:inline-block;background:#7c3aed22;color:#a78bfa;font-size:.7rem;font-weight:600;padding:.2rem .7rem;border-radius:99px;margin-top:.5rem;border:1px solid #7c3aed44}
</style>
</head>
<body>
<h1>🍺 Drunks POS</h1>
<p class="sub">¿A dónde quieres ir?</p>
<div class="cards">
  <a class="card" href="/vendedor">
    <div class="icon">🛒</div>
    <div class="name">Vendedor</div>
    <div class="desc">Tomar pedidos y cobrar</div>
    <span class="badge">Caja</span>
  </a>
  <a class="card" href="/cocina">
    <div class="icon">👨‍🍳</div>
    <div class="name">Cocina</div>
    <div class="desc">Ver y gestionar pedidos</div>
    <span class="badge">Producción</span>
  </a>
  <a class="card" href="/dashboard">
    <div class="icon">📊</div>
    <div class="name">Dashboard</div>
    <div class="desc">Reportes y estadísticas</div>
    <span class="badge">Admin</span>
  </a>
</div>
</body>
</html>"""

APP_SHELL_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Drunks POS</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
html,body{height:100%;overflow:hidden;background:#0d0d1a}
.shell{display:flex;flex-direction:column;height:100vh}
.topbar{
  background:#0d0d1a;border-bottom:1px solid #1e1e35;
  height:46px;display:flex;align-items:center;padding:0 14px;gap:6px;
  flex-shrink:0;user-select:none;-webkit-user-select:none
}
.logo{font-family:'Segoe UI',sans-serif;font-weight:800;font-size:.95rem;
  color:#fff;display:flex;align-items:center;gap:6px;margin-right:14px;white-space:nowrap}
.nav-btn{
  font-family:'Segoe UI',sans-serif;font-size:.8rem;font-weight:600;
  color:#8b8ba8;background:transparent;border:1px solid transparent;
  border-radius:8px;padding:5px 13px;cursor:pointer;transition:.15s;
  display:flex;align-items:center;gap:5px;white-space:nowrap
}
.nav-btn:hover{background:#1e1e35;color:#fff}
.nav-btn.active{background:#2d1f6e;color:#c4b5fd;border-color:#7c3aed55}
.spacer{flex:1}
.upd-pill{
  display:none;align-items:center;gap:7px;
  background:#7c3aed22;border:1px solid #7c3aed55;border-radius:99px;
  padding:4px 13px 4px 9px;cursor:pointer;
  animation:pulse 2s infinite
}
.upd-pill:hover{background:#7c3aed44}
.upd-dot{width:8px;height:8px;background:#7c3aed;border-radius:50%;flex-shrink:0}
.upd-txt{font-size:.75rem;font-weight:700;color:#c4b5fd;font-family:'Segoe UI',sans-serif}
@keyframes pulse{
  0%,100%{box-shadow:0 0 0 0 rgba(124,58,237,.4)}
  50%{box-shadow:0 0 0 6px rgba(124,58,237,0)}
}
iframe{flex:1;border:none;width:100%}
</style>
</head>
<body>
<div class="shell">
  <div class="topbar">
    <div class="logo">🍺 Drunks POS</div>
    <button class="nav-btn active" id="btn-vendedor" onclick="go('/vendedor','vendedor')">🛒 Vendedor</button>
    <button class="nav-btn" id="btn-cocina"   onclick="go('/cocina','cocina')">👨‍🍳 Cocina</button>
    <button class="nav-btn" id="btn-dashboard" onclick="go('/dashboard','dashboard')">📊 Dashboard</button>
    <div class="spacer"></div>
    <div class="upd-pill" id="upd-pill" onclick="reqUpdate()">
      <div class="upd-dot"></div>
      <span class="upd-txt" id="upd-txt">Actualización disponible</span>
    </div>
  </div>
  <iframe id="frame" src="/vendedor"></iframe>
</div>
<script>
var _cur = 'vendedor';
function go(url, name) {
  if (_cur === name) return;
  _cur = name;
  document.querySelectorAll('.nav-btn').forEach(function(b){b.classList.remove('active')});
  document.getElementById('btn-' + name).classList.add('active');
  document.getElementById('frame').src = url;
}
// Sync active tab when iframe navigates internally
document.getElementById('frame').addEventListener('load', function() {
  try {
    var p = this.contentWindow.location.pathname;
    if (p.includes('vendedor')) { _cur='vendedor'; }
    else if (p.includes('cocina')) { _cur='cocina'; }
    else if (p.includes('dashboard')) { _cur='dashboard'; }
    else return;
    document.querySelectorAll('.nav-btn').forEach(function(b){b.classList.remove('active')});
    document.getElementById('btn-' + _cur).classList.add('active');
  } catch(e) {}
});
function showUpdate(ver) {
  document.getElementById('upd-txt').textContent = 'v' + ver + ' disponible';
  document.getElementById('upd-pill').style.display = 'flex';
}
function reqUpdate() {
  if (window.pywebview && window.pywebview.api)
    window.pywebview.api.open_update_window();
}
</script>
</body>
</html>"""

@app.get("/app", response_class=HTMLResponse)
def page_app(): return HTMLResponse(APP_SHELL_HTML)

@app.get("/")
def root(): return HTMLResponse(LAUNCHER_HTML)

# ─────────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run(app, host=HOST, port=PORT, reload=False, log_level="info")
