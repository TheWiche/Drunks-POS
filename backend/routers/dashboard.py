import io
from datetime import datetime

from fastapi import APIRouter, HTTPException, Response

from ..database import get_conn

router = APIRouter()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


@router.get("/dashboard")
def get_dashboard(range: str = "week"):
    with get_conn() as conn:
        today = datetime.now().strftime("%Y-%m-%d")
        if   range == "today": df = f"AND DATE(fecha)='{today}'"
        elif range == "month": df = "AND DATE(fecha)>=DATE('now','-29 days')"
        elif range == "all":   df = ""
        else:                  df = "AND DATE(fecha)>=DATE('now','-6 days')"

        def q(sql, *a): return conn.execute(sql, a).fetchone()[0]
        def qa(sql):    return [dict(r) for r in conn.execute(sql).fetchall()]

        ventas_hoy      = q(f"SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado' AND DATE(fecha)='{today}'")
        pedidos_hoy     = q(f"SELECT COUNT(*) FROM pedidos WHERE DATE(fecha)='{today}'")
        pendientes_now  = q("SELECT COUNT(*) FROM pedidos WHERE estado='pendiente'")
        ventas_periodo  = q(f"SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado' {df}")
        pedidos_periodo = q(f"SELECT COUNT(*) FROM pedidos WHERE estado='entregado' {df}")
        ticket_prom     = q("SELECT COALESCE(AVG(total),0) FROM pedidos WHERE estado='entregado'")
        total_todo      = q("SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado'")

        if range == "all":
            daily = qa("""
                SELECT DATE(fecha) as dia, COALESCE(SUM(total),0) as total, COUNT(*) as cnt
                FROM pedidos WHERE estado='entregado'
                GROUP BY DATE(fecha) ORDER BY dia DESC LIMIT 30""")
            daily.reverse()
        else:
            daily = qa(f"""
                SELECT DATE(fecha) as dia, COALESCE(SUM(total),0) as total, COUNT(*) as cnt
                FROM pedidos WHERE estado='entregado' {df}
                GROUP BY DATE(fecha) ORDER BY dia""")

        hourly = qa("""
            SELECT CAST(strftime('%H',fecha) AS INTEGER) as hora,
                   COUNT(*) as cnt, COALESCE(SUM(total),0) as total
            FROM pedidos WHERE estado='entregado'
            GROUP BY hora ORDER BY hora""")

        top_prods = qa(f"""
            SELECT pr.nombre, SUM(dp.cantidad) as vendidos, SUM(dp.cantidad*pr.precio) as revenue
            FROM detalle_pedidos dp
            JOIN productos pr ON dp.producto_id=pr.id
            JOIN pedidos p ON dp.pedido_id=p.id
            WHERE p.estado='entregado' {df}
            GROUP BY pr.id, pr.nombre ORDER BY vendidos DESC LIMIT 8""")

        metodos = qa(f"""
            SELECT metodo_pago, COUNT(*) as cnt, COALESCE(SUM(total),0) as total
            FROM pedidos WHERE estado='entregado' {df}
            GROUP BY metodo_pago""")

        by_cat = qa(f"""
            SELECT COALESCE(c.nombre,'Sin Categoria') as nombre,
                   SUM(dp.cantidad) as vendidos, SUM(dp.cantidad*pr.precio) as revenue
            FROM detalle_pedidos dp
            JOIN productos pr ON dp.producto_id=pr.id
            JOIN pedidos p ON dp.pedido_id=p.id
            LEFT JOIN categorias c ON pr.categoria_id=c.id
            WHERE p.estado='entregado' {df}
            GROUP BY c.id, c.nombre ORDER BY revenue DESC""")

        recientes = qa(f"""
            SELECT id, cliente, metodo_pago, total, estado, fecha
            FROM pedidos WHERE 1=1 {df.replace("estado='entregado' AND","").replace("AND estado='entregado'","")}
            ORDER BY id DESC LIMIT 30""")

    return {
        "range":           range,
        "ventas_hoy":      float(ventas_hoy),
        "pedidos_hoy":     int(pedidos_hoy),
        "pendientes_now":  int(pendientes_now),
        "ventas_periodo":  float(ventas_periodo),
        "pedidos_periodo": int(pedidos_periodo),
        "ticket_prom":     float(ticket_prom),
        "total_todo":      float(total_todo),
        "daily":           daily,
        "hourly":          hourly,
        "top_prods":       top_prods,
        "metodos":         metodos,
        "by_cat":          by_cat,
        "recientes":       recientes,
    }


@router.get("/export/excel")
def export_excel():
    if not XLSX_AVAILABLE:
        raise HTTPException(503, "openpyxl no está instalado")

    now = datetime.now()

    C_PURPLE   = "4C1D95"
    C_PURP_LT  = "F0EBFF"
    C_PURP_MED = "7C3AED"
    C_GREEN    = "166534"
    C_RED      = "991B1B"
    C_AMBER    = "92400E"

    def h_font():  return Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    def h_fill():  return PatternFill("solid", fgColor=C_PURPLE)
    def h_align(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def row_fill(i): return PatternFill("solid", fgColor=C_PURP_LT if i % 2 == 0 else "FFFFFF")
    def thin(color="D8B4FE"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def apply_header(ws, cols, row=1):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = h_font(); cell.fill = h_fill(); cell.alignment = h_align()
        ws.row_dimensions[row].height = 22

    def auto_col(ws, mn=10, mx=55):
        for col in ws.columns:
            w = mn
            for cell in col:
                if cell.value is not None:
                    w = max(w, min(len(str(cell.value)) + 4, mx))
            ws.column_dimensions[get_column_letter(col[0].column)].width = w

    with get_conn() as conn:
        pedidos = conn.execute(
            "SELECT id, cliente, metodo_pago, total, estado, fecha FROM pedidos ORDER BY fecha DESC"
        ).fetchall()

        detalle = conn.execute("""
            SELECT p.id AS pedido_id, p.cliente, p.metodo_pago, p.fecha,
                   pr.nombre AS producto, COALESCE(c.nombre,'Sin Categoria') AS categoria,
                   dp.cantidad, pr.precio, dp.cantidad*pr.precio AS subtotal, dp.observaciones
            FROM detalle_pedidos dp
            JOIN pedidos  p  ON dp.pedido_id   = p.id
            JOIN productos pr ON dp.producto_id = pr.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            ORDER BY p.fecha DESC, p.id, pr.nombre
        """).fetchall()

        productos = conn.execute("""
            SELECT p.id, p.nombre, COALESCE(c.nombre,'Sin Categoria') AS categoria,
                   p.precio, p.disponible, p.tiene_base
            FROM productos p LEFT JOIN categorias c ON p.categoria_id=c.id
            ORDER BY c.nombre, p.nombre
        """).fetchall()

        kpi_ingresos  = conn.execute("SELECT COALESCE(SUM(total),0) FROM pedidos WHERE estado='entregado'").fetchone()[0]
        kpi_pedidos   = conn.execute("SELECT COUNT(*) FROM pedidos").fetchone()[0]
        kpi_ticket    = conn.execute("SELECT COALESCE(AVG(total),0) FROM pedidos WHERE estado='entregado'").fetchone()[0]
        kpi_pendiente = conn.execute("SELECT COUNT(*) FROM pedidos WHERE estado='pendiente'").fetchone()[0]

        metodos_sum = conn.execute("""
            SELECT metodo_pago, COUNT(*) AS cnt, COALESCE(SUM(total),0) AS total
            FROM pedidos WHERE estado='entregado' GROUP BY metodo_pago ORDER BY total DESC
        """).fetchall()

        top5 = conn.execute("""
            SELECT pr.nombre, COALESCE(c.nombre,'Sin Categoria') AS cat,
                   SUM(dp.cantidad) AS uds, SUM(dp.cantidad*pr.precio) AS rev
            FROM detalle_pedidos dp
            JOIN productos  pr ON dp.producto_id = pr.id
            JOIN pedidos    p  ON dp.pedido_id   = p.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            WHERE p.estado='entregado'
            GROUP BY pr.id ORDER BY rev DESC LIMIT 5
        """).fetchall()

        by_cat = conn.execute("""
            SELECT COALESCE(c.nombre,'Sin Categoria') AS cat,
                   COUNT(DISTINCT p.id) AS pedidos, SUM(dp.cantidad) AS uds,
                   SUM(dp.cantidad*pr.precio) AS rev
            FROM detalle_pedidos dp
            JOIN productos  pr ON dp.producto_id = pr.id
            JOIN pedidos    p  ON dp.pedido_id   = p.id
            LEFT JOIN categorias c ON pr.categoria_id = c.id
            WHERE p.estado='entregado'
            GROUP BY c.id ORDER BY rev DESC
        """).fetchall()

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "DRUNKS POS  —  REPORTE DE VENTAS"
    t.font      = Font(bold=True, size=20, color=C_PURP_MED, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill      = PatternFill("solid", fgColor="FAF5FF")
    ws.row_dimensions[1].height = 46

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = f"Generado el {now.strftime('%d/%m/%Y a las %H:%M:%S')}"
    s.font      = Font(italic=True, size=10, color="9CA3AF", name="Calibri")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    r = 4
    ws.cell(r, 1).value = "ESTADÍSTICAS GENERALES"
    ws.cell(r, 1).font  = Font(bold=True, size=13, color=C_PURP_MED, name="Calibri")
    ws.row_dimensions[r].height = 22
    r += 1

    kpis = [
        ("Total de Pedidos Registrados",  int(kpi_pedidos),    False, C_PURPLE),
        ("Ingresos Totales (entregados)", float(kpi_ingresos), True,  C_GREEN),
        ("Ticket Promedio",               float(kpi_ticket),   True,  C_PURP_MED),
        ("Pedidos Pendientes Actualmente",int(kpi_pendiente),  False, C_AMBER),
    ]
    for label, value, money, color in kpis:
        cl = ws.cell(r, 1, label)
        cl.font = Font(size=11, color="374151", name="Calibri")
        cl.fill = PatternFill("solid", fgColor="F9FAFB")
        cv = ws.cell(r, 2, round(value) if money else value)
        cv.font          = Font(bold=True, size=13, color=color, name="Calibri")
        cv.number_format = "#,##0" if money else "0"
        cv.alignment     = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws.cell(r, 1).value = "VENTAS POR MÉTODO DE PAGO"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    ws.row_dimensions[r].height = 20
    r += 1
    apply_header(ws, ["Método de Pago", "# Pedidos", "Total Ingresos"], row=r)
    r += 1
    for i, m in enumerate(metodos_sum):
        ws.cell(r+i, 1, m["metodo_pago"]).fill = row_fill(i)
        ws.cell(r+i, 2, m["cnt"]).fill          = row_fill(i)
        tc = ws.cell(r+i, 3, round(m["total"]))
        tc.number_format = "#,##0"; tc.font = Font(bold=True, color=C_GREEN); tc.fill = row_fill(i)
    r += len(metodos_sum) + 2

    ws.cell(r, 1).value = "INGRESOS POR CATEGORÍA"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    r += 1
    apply_header(ws, ["Categoría", "Pedidos", "Unidades", "Ingresos"], row=r)
    r += 1
    for i, cat in enumerate(by_cat):
        ws.cell(r+i, 1, cat["cat"]).fill       = row_fill(i)
        ws.cell(r+i, 2, cat["pedidos"]).fill    = row_fill(i)
        ws.cell(r+i, 3, cat["uds"]).fill        = row_fill(i)
        rc = ws.cell(r+i, 4, round(cat["rev"]))
        rc.number_format = "#,##0"; rc.font = Font(bold=True, color=C_GREEN); rc.fill = row_fill(i)
    r += len(by_cat) + 2

    ws.cell(r, 1).value = "TOP 5 PRODUCTOS MÁS VENDIDOS"
    ws.cell(r, 1).font  = Font(bold=True, size=12, color=C_PURP_MED, name="Calibri")
    r += 1
    apply_header(ws, ["Producto", "Categoría", "Unidades Vendidas", "Ingresos Generados"], row=r)
    r += 1
    for i, p in enumerate(top5):
        ws.cell(r+i, 1, p["nombre"]).fill = row_fill(i)
        ws.cell(r+i, 2, p["cat"]).fill    = row_fill(i)
        ws.cell(r+i, 3, p["uds"]).fill    = row_fill(i)
        ri = ws.cell(r+i, 4, round(p["rev"]))
        ri.number_format = "#,##0"; ri.font = Font(bold=True, color=C_GREEN); ri.fill = row_fill(i)

    auto_col(ws)

    ws2 = wb.create_sheet("Pedidos")
    ws2.sheet_view.showGridLines = False
    cols2 = ["#", "Cliente", "Método de Pago", "Total (COP)", "Estado", "Fecha", "Hora"]
    apply_header(ws2, cols2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
    ws2.row_dimensions[1].height = 22

    for i, p in enumerate(pedidos):
        row_n = i + 2
        try:    dt = datetime.fromisoformat(str(p["fecha"]))
        except: dt = now
        bg = row_fill(i)
        entregado = p["estado"] == "entregado"
        ws2.cell(row_n, 1, p["id"]).fill      = bg
        ws2.cell(row_n, 2, p["cliente"]).fill  = bg
        ws2.cell(row_n, 3, p["metodo_pago"]).fill = bg
        tc = ws2.cell(row_n, 4, round(p["total"]))
        tc.number_format = "#,##0"; tc.font = Font(bold=True, color=C_GREEN); tc.fill = bg
        ec = ws2.cell(row_n, 5, "Entregado" if entregado else "Pendiente")
        ec.font = Font(bold=True, color=C_GREEN if entregado else C_AMBER)
        ec.fill = PatternFill("solid", fgColor="D1FAE5" if entregado else "FEF3C7")
        ec.alignment = Alignment(horizontal="center")
        ws2.cell(row_n, 6, dt.strftime("%d/%m/%Y")).fill = bg
        ws2.cell(row_n, 7, dt.strftime("%H:%M")).fill    = bg
        ws2.row_dimensions[row_n].height = 18

    auto_col(ws2)

    ws3 = wb.create_sheet("Detalle de Pedidos")
    ws3.sheet_view.showGridLines = False
    cols3 = ["Pedido #", "Cliente", "Método", "Fecha y Hora", "Producto",
             "Categoría", "Cantidad", "Precio Unit.", "Subtotal", "Base / Notas"]
    apply_header(ws3, cols3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"
    ws3.row_dimensions[1].height = 22

    for i, d in enumerate(detalle):
        row_n = i + 2
        try:    dt = datetime.fromisoformat(str(d["fecha"]))
        except: dt = now
        bg = row_fill(i)
        ws3.cell(row_n, 1, d["pedido_id"]).fill  = bg
        ws3.cell(row_n, 2, d["cliente"]).fill     = bg
        ws3.cell(row_n, 3, d["metodo_pago"]).fill = bg
        ws3.cell(row_n, 4, dt.strftime("%d/%m/%Y %H:%M")).fill = bg
        ws3.cell(row_n, 5, d["producto"]).fill    = bg
        ws3.cell(row_n, 6, d["categoria"]).fill   = bg
        ws3.cell(row_n, 7, d["cantidad"]).fill     = bg
        pu = ws3.cell(row_n, 8, round(d["precio"]))
        pu.number_format = "#,##0"; pu.fill = bg
        sb = ws3.cell(row_n, 9, round(d["subtotal"]))
        sb.number_format = "#,##0"; sb.font = Font(bold=True, color=C_GREEN); sb.fill = bg
        ws3.cell(row_n, 10, d["observaciones"] or "").fill = bg
        ws3.row_dimensions[row_n].height = 18

    auto_col(ws3)

    ws4 = wb.create_sheet("Productos")
    ws4.sheet_view.showGridLines = False
    cols4 = ["#", "Nombre del Producto", "Categoría", "Precio (COP)", "Estado", "Requiere Base"]
    apply_header(ws4, cols4)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"

    for i, p in enumerate(productos):
        row_n = i + 2
        bg     = row_fill(i)
        activo = bool(p["disponible"])
        base   = bool(p["tiene_base"])
        ws4.cell(row_n, 1, p["id"]).fill      = bg
        ws4.cell(row_n, 2, p["nombre"]).fill   = bg
        ws4.cell(row_n, 3, p["categoria"]).fill = bg
        pc = ws4.cell(row_n, 4, round(p["precio"]))
        pc.number_format = "#,##0"; pc.font = Font(bold=True, color=C_GREEN); pc.fill = bg
        dc = ws4.cell(row_n, 5, "Activo" if activo else "Agotado")
        dc.font = Font(bold=True, color=C_GREEN if activo else C_RED)
        dc.fill = PatternFill("solid", fgColor="D1FAE5" if activo else "FEE2E2")
        dc.alignment = Alignment(horizontal="center")
        bc = ws4.cell(row_n, 6, "Sí" if base else "No")
        bc.font = Font(bold=True, color="D97706" if base else "6B7280")
        bc.fill = PatternFill("solid", fgColor="FEF3C7" if base else "F9FAFB")
        bc.alignment = Alignment(horizontal="center")
        ws4.row_dimensions[row_n].height = 18

    auto_col(ws4)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"Drunks_Reporte_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
